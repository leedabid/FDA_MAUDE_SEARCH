# -*- coding: utf-8 -*-
"""
FDA MAUDE CGM 부작용 데이터 수집기
=================================================
Dexcom / FreeStyle Libre CGM 관련 부작용(이상사례) 보고를
FDA openFDA API(MAUDE)에서 수집하여 SQLite DB와 Excel 파일로 저장합니다.

[주요 기능]
1. openFDA Device Event API(MAUDE)에서 Dexcom, FreeStyle Libre 관련
   부작용 보고를 날짜 범위로 조회
2. 주요 필드(MAUDE 번호, EVENT_TYPE, BRAND_NAME, 제조사, 환자 문제,
   이벤트 설명, 제조사 내러티브 등)를 SQLite DB에 저장 (중복 방지)
3. 매 실행마다 최신 Excel 파일을 작업 폴더에 생성 (필드별 정렬·필터 가능)
4. 각 보고서를 MAUDE 번호와 함께 "소비자 불만사항 / 제조사 대응 / 결론"
   3개 섹션으로 요약 (FDA 원문 + 한글 키워드 기반 간단 요약)

[실행 방법]
  # 최초 실행 (지난 1년치 수집)
  python fda_maude_collector.py --initial

  # 매일 실행 (전날 데이터 증분 수집)
  python fda_maude_collector.py

  # 특정 기간만 수집
  python fda_maude_collector.py --start 20260101 --end 20260422

  # API 키 사용 (권장, 속도 제한 완화)
  python fda_maude_collector.py --api-key YOUR_API_KEY

[필요 패키지]
  pip install requests pandas openpyxl
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sqlite3
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    import requests
except ImportError:  # pragma: no cover
    print("requests 라이브러리가 필요합니다. 'pip install requests' 실행하세요.")
    sys.exit(1)

# Windows 콘솔(CP949)에서도 한글이 깨지지 않도록 표준 출력을 UTF-8 로 재설정
try:  # Python 3.7+
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
except Exception:
    pass


_ILLEGAL_XLSX_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def _sanitize_xlsx_value(value: Any) -> Any:
    """openpyxl 이 거부하는 제어문자를 제거한다.

    MAUDE 원문에는 보이지 않는 제어문자나 잘못된 문자가 섞일 수 있어
    Excel 스트리밍 저장 시 IllegalCharacterError 를 일으킨다.
    """
    if isinstance(value, str) and value:
        return _ILLEGAL_XLSX_CONTROL_CHARS.sub("", value)
    return value


def _sanitize_xlsx_row(
    row: Iterable[Any],
    *,
    report_number: Optional[str] = None,
    warn_limit: int = 20,
    warned: Optional[Dict[str, int]] = None,
) -> List[Any]:
    """행 단위로 Excel 안전값으로 정리하고, 정리 사실을 제한적으로 로그한다."""
    cleaned: List[Any] = []
    changed = False
    for idx, value in enumerate(row):
        new_value = _sanitize_xlsx_value(value)
        if new_value != value:
            changed = True
            if warned is not None and warned.get("count", 0) < warn_limit:
                col_name = _EXCEL_MAIN_COLUMNS[idx][0] if idx < len(_EXCEL_MAIN_COLUMNS) else f"col_{idx+1}"
                log.warning(
                    "Excel 금지문자 정리: report_number=%s, column=%s",
                    report_number or "(unknown)",
                    col_name,
                )
                warned["count"] = warned.get("count", 0) + 1
        cleaned.append(new_value)
    if changed and warned is not None and warned.get("count", 0) == warn_limit:
        warned["count"] = warn_limit + 1
        log.warning("Excel 금지문자 정리 로그가 %d건을 넘어서 추가 로그를 생략합니다.", warn_limit)
    return cleaned

# ===========================================================================
# ★ 사용자 설정 영역 — 여기만 고치면 검색 대상/출력 파일 이름 등을 바꿀 수 있음 ★
# ===========================================================================

# ---------------------------------------------------------------------------
# [설정 1] 검색할 당뇨 디바이스 조건 (제조사 + 브랜드)
# ---------------------------------------------------------------------------
# 제조사/브랜드를 OR 로 묶어 검색합니다.
#   (manufacturer_d_name IN SEARCH_MANUFACTURERS)
#   OR (brand_name IN SEARCH_BRANDS)
SEARCH_MANUFACTURERS: List[str] = [
    "DEXCOM",
    "ABBOTT",
    "ABBOTT DIABETES CARE",
    "TANDEM",
    "TANDEM DIABETES CARE",
    "INSULET",
]
SEARCH_BRANDS: List[str] = [
    "FREESTYLE LIBRE",
    "DEXCOM",
    "OMNIPOD",
    "T:SLIM",
    "MINIMED 780G",
]

# 체크포인트 키/기존 함수 시그니처 호환을 위해 유지합니다.
CGM_BRANDS: List[str] = SEARCH_MANUFACTURERS + SEARCH_BRANDS

# ---------------------------------------------------------------------------
# [설정 1-b] 브랜드 → 디바이스 카테고리 매핑
# ---------------------------------------------------------------------------
# 위 CGM_BRANDS 로 수집된 각 보고서를 device_category 칼럼("CGM" / "Insulin Pump")
# 으로 분류해 DB 에 저장합니다. brand_name 에 대해 위에서부터 차례로 부분일치
# 검사하여 첫 매치를 사용합니다 (대소문자 무관).
#
# 주의: 순서가 중요합니다.
#   - DEXCOM 이 MOBI 보다 먼저 와야 "DEXCOM MOBILE APP" 류가 MOBI=Pump 로
#     오분류되지 않습니다.
BRAND_CATEGORY_MAP: List[Tuple[str, str]] = [
    ("DEXCOM",          "CGM"),
    ("FREESTYLE LIBRE", "CGM"),
    ("LIBRE",           "CGM"),
    ("SENSEN",          "CGM"),
    ("MINIMED 780G",    "Insulin Pump"),
    ("OMNIPOD",         "Insulin Pump"),
    ("T:SLIM",          "Insulin Pump"),
    ("TANDEM",          "Insulin Pump"),
    ("INSULET",         "Insulin Pump"),
]


def _resolve_device_category(brand_name: Optional[str]) -> Optional[str]:
    """brand_name 으로부터 디바이스 카테고리("CGM" / "Insulin Pump") 추론.
    매치 안 되면 None (DB 에는 NULL).
    """
    if not brand_name:
        return None
    upper = str(brand_name).upper()
    for needle, cat in BRAND_CATEGORY_MAP:
        if needle in upper:
            return cat
    return None

# ---------------------------------------------------------------------------
# [설정 2] 수집할 EVENT_TYPE (★★★ 볼륨이 너무 크면 여기부터 조정 ★★★)
# ---------------------------------------------------------------------------
# FDA MAUDE 의 event_type 은 5가지: Death / Injury / Malfunction / Other /
# "No answer provided". Dexcom · Libre 는 "Malfunction"(단순 센서 오작동)
# 보고가 연간 수십만 건이라, 환자 피해 사례(Death/Injury/Other)만 수집
# 하는 것을 기본값으로 합니다.
#
# Malfunction 까지 포함하고 싶으면 아래 리스트에 "Malfunction" 을 추가하세요.
EVENT_TYPES: List[str] = [
    "Death",
    "Injury",
    "Other",
    # "Malfunction",   # ← 주석 해제 시 오작동 보고까지 포함 (건수 폭증 주의)
]

# ---------------------------------------------------------------------------
# [설정 3] adverse_event_flag=Y 만 수집할지 여부
# ---------------------------------------------------------------------------
# True 로 하면 "환자에게 실제로 부작용이 발생한 보고"만 가져옵니다.
# (adverse_event_flag = Y). 단순 기기 결함 보고는 제외됩니다.
# False 로 하면 플래그 무관 전체 수집.
ONLY_ADVERSE_EVENTS: bool = True

# ---------------------------------------------------------------------------
# [설정 4] 보조 필드로도 검색할지 여부
# ---------------------------------------------------------------------------
# device.brand_name 에서 0건이 나오면, device.generic_name 이나
# device.manufacturer_d_name 으로도 한 번 더 찾습니다. (안전장치)
USE_FALLBACK_FIELDS: bool = True

# 보조 검색에 쓸 일반명 / 제조사명 (위 CGM_BRANDS 가 모두 실패했을 때만 작동)
FALLBACK_GENERIC_NAMES: List[str] = [
    "CONTINUOUS GLUCOSE MONITOR",
    "GLUCOSE SENSOR",
]
FALLBACK_MANUFACTURERS: List[str] = [
    "DEXCOM",
    "ABBOTT",
    "ABBOTT DIABETES CARE",
    "TANDEM",
    "TANDEM DIABETES CARE",
    "INSULET",
    "MEDTRONIC",
]

# ---------------------------------------------------------------------------
# [설정 5] 출력 파일 이름
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "fda_maude_cgm.db"           # 누적 저장 SQLite DB
EXCEL_PATH = BASE_DIR / "fda_maude_cgm.xlsx"      # 사람이 보는 Excel
LOG_PATH = BASE_DIR / "fda_maude_collector.log"   # 실행 로그

# API 키 파일 후보 (순서대로 탐색, 먼저 발견되는 것 사용)
API_KEY_FILES: List[Path] = [
    BASE_DIR / "FDA_MAUDE_API_KEY.txt",  # 사용자 선호 파일명
    BASE_DIR / "api_key.txt",            # 기본 파일명
]

# ---------------------------------------------------------------------------
# [설정 6] openFDA API 상수 (일반적으로 수정 불필요)
# ---------------------------------------------------------------------------
# openFDA 는 미국 정부가 운영하는 무료 공공 API (과금 없음).
# API 키 발급 URL: https://open.fda.gov/apis/authentication/
#
# 속도 제한 (키 없음 → 키 있음):
#   - 요청/분    : 40     → 240 (6배)
#   - 요청/일(시):  1,000 → 120,000 (사실상 무제한)
#
# 서버측 하드 리밋 (키와 무관, 변경 불가):
#   - limit (1회 요청당 최대 레코드) = 1,000
#   - skip  (페이지네이션 누적 최대) = 25,000
#   → skip 한도 초과 시 스크립트가 자동으로 월 단위 분할 수집
OPENFDA_URL = "https://api.fda.gov/device/event.json"
PAGE_SIZE = 1000          # 한 번 요청 최대 건수 (openFDA 하드 상한)
MAX_SKIP = 25000          # 페이지네이션 누적 상한 (openFDA 하드 제약)
REQUEST_TIMEOUT = 60      # 초

# 요청 간 대기시간 (초) — 키 유무에 따라 자동 선택
# 키 없음: 분당 40 제한 → 1.6초 간격
# 키 있음: 분당 240 제한 → 0.25초로도 버퍼 충분, 실측으로는 0.05도 안전
SLEEP_WITH_KEY: float = 0.05
SLEEP_WITHOUT_KEY: float = 1.6

# ===========================================================================
# ▲ 사용자 설정 영역 끝 ▲
# ===========================================================================

# ---------------------------------------------------------------------------
# 로깅
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("fda_maude")


# ---------------------------------------------------------------------------
# DB 스키마
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# 스키마는 2단계로 분리:
#   1) SCHEMA_TABLES_SQL : 테이블만 (CREATE TABLE IF NOT EXISTS)
#   2) _migrate_schema() : 구 DB 에 신규 칼럼 ALTER TABLE 로 추가
#   3) SCHEMA_INDEXES_SQL: 인덱스 (칼럼이 모두 존재함이 보장된 뒤 실행)
# → 한 블록에 몰아넣으면, 구 DB 에서 신규 칼럼 인덱스 생성이 ALTER 보다
#   먼저 실행돼 "no such column" 오류로 전체가 실패한다.
# ---------------------------------------------------------------------------

SCHEMA_TABLES_SQL = """
CREATE TABLE IF NOT EXISTS maude_reports (
    report_number          TEXT PRIMARY KEY,    -- MAUDE 번호 (MDR Report Key)
    event_type             TEXT,                -- 사건 유형 (Death / Injury / Malfunction / Other)
    date_received          TEXT,                -- FDA 접수일 (YYYY-MM-DD)
    date_of_event          TEXT,                -- 실제 사건 발생일
    date_report            TEXT,                -- 제조사 보고일
    brand_name             TEXT,                -- 제품명 (예: DEXCOM G7)
    device_category        TEXT,                -- 디바이스 카테고리 (CGM / Insulin Pump)
    generic_name           TEXT,                -- 일반명 (예: Continuous Glucose Monitor)
    manufacturer_name      TEXT,                -- 제조사
    manufacturer_country   TEXT,                -- 제조사 소재 국가
    model_number           TEXT,
    product_code           TEXT,                -- FDA 제품 코드
    source_type            TEXT,                -- 보고 주체 (Manufacturer / Consumer / Health Professional 등)
    report_source_code     TEXT,
    type_of_report         TEXT,                -- Initial submission / Followup 등
    patient_age            TEXT,                -- 환자 나이 (예: "65 YR")
    patient_sex            TEXT,                -- 환자 성별 (Male / Female / Unknown)
    patient_ethnicity      TEXT,                -- 환자 민족 (Hispanic / Not Hispanic)
    patient_race           TEXT,                -- 환자 인종 (White / Black / Asian 등)
    patient_weight         TEXT,                -- 환자 체중 (있을 경우)
    patient_problems       TEXT,                -- 환자 문제 (세미콜론 구분)
    product_problems       TEXT,                -- 제품 문제 (세미콜론 구분)
    event_description      TEXT,                -- 소비자/보고자 기술 내용 (원문)
    manufacturer_narrative TEXT,                -- 제조사 서술 (원문, 대응·조사결과)
    additional_manufacturer_narrative TEXT,     -- 추가 제조사 서술 (원문, 후속 대응)
    summary_complaint_kr   TEXT,                -- 한글 요약: 소비자 불만사항
    summary_response_kr    TEXT,                -- 한글 요약: 제조사 대응
    summary_conclusion_kr  TEXT,                -- 한글 요약: 결론
    raw_json               TEXT,                -- 원본 JSON (디버깅용)
    collected_at           TEXT                 -- 수집 시각
);

-- 증분 수집용 체크포인트 — 브랜드 조합별로 "어디까지 수집했는지" 기록
CREATE TABLE IF NOT EXISTS collection_checkpoint (
    brand_key           TEXT PRIMARY KEY,  -- 정규화된 CGM_BRANDS (예: "DEXCOM|FREESTYLE LIBRE")
    last_search_end     TEXT,              -- 마지막으로 검색한 종료일 (YYYY-MM-DD)
    last_run_at         TEXT,              -- 마지막 실행 시각
    last_inserted       INTEGER,           -- 마지막 실행 시 신규/갱신 건수
    total_collected     INTEGER,           -- 누적 수집 건수
    first_run_at        TEXT,              -- 최초 실행 시각 (참고)
    note                TEXT               -- 비고 (예: "최초 2년치 수집")
);
"""

SCHEMA_INDEXES_SQL = """
CREATE INDEX IF NOT EXISTS idx_event_type      ON maude_reports(event_type);
CREATE INDEX IF NOT EXISTS idx_brand           ON maude_reports(brand_name);
CREATE INDEX IF NOT EXISTS idx_device_category ON maude_reports(device_category);
CREATE INDEX IF NOT EXISTS idx_manufacturer    ON maude_reports(manufacturer_name);
CREATE INDEX IF NOT EXISTS idx_mfr_country     ON maude_reports(manufacturer_country);
CREATE INDEX IF NOT EXISTS idx_date_received   ON maude_reports(date_received);
CREATE INDEX IF NOT EXISTS idx_product_code    ON maude_reports(product_code);
CREATE INDEX IF NOT EXISTS idx_patient_sex     ON maude_reports(patient_sex);
CREATE INDEX IF NOT EXISTS idx_type_of_report  ON maude_reports(type_of_report);
CREATE INDEX IF NOT EXISTS idx_reports_date
    ON maude_reports(date_received);
CREATE INDEX IF NOT EXISTS idx_reports_event
    ON maude_reports(event_type);
CREATE INDEX IF NOT EXISTS idx_reports_category_date
    ON maude_reports(device_category, date_received);
CREATE INDEX IF NOT EXISTS idx_reports_brand_date
    ON maude_reports(brand_name, date_received);
CREATE INDEX IF NOT EXISTS idx_reports_mfr_date
    ON maude_reports(manufacturer_name, date_received);
CREATE INDEX IF NOT EXISTS idx_reports_product_code
    ON maude_reports(product_code);
"""

# 하위 호환: 예전 스크립트/테스트가 SCHEMA_SQL 을 임포트할 수도 있으므로 합본도 제공.
# 주의: 이걸 그대로 executescript 로 돌리면 구 DB 에서 오류. init_db() 를 써야 함.
SCHEMA_SQL = SCHEMA_TABLES_SQL + "\n" + SCHEMA_INDEXES_SQL


LEGACY_FTS_TRIGGER_DROP_SQL = """
DROP TRIGGER IF EXISTS maude_reports_ai;
DROP TRIGGER IF EXISTS maude_reports_ad;
DROP TRIGGER IF EXISTS maude_reports_au;
"""


# ---------------------------------------------------------------------------
# 한글 키워드 요약 (LLM 없이, 규칙 기반)
# ---------------------------------------------------------------------------

# CGM 부작용에서 자주 등장하는 영문 키워드 → 한글 매핑
_SYMPTOM_MAP: List[Tuple[re.Pattern, str]] = [
    (re.compile(r"\bsensor\s+(fail|error|malfunction)", re.I), "센서 오류/고장"),
    (re.compile(r"\b(inaccurate|false|incorrect)\s+reading", re.I), "수치 부정확"),
    (re.compile(r"\bno\s+reading|signal\s+loss|loss\s+of\s+signal", re.I), "신호 끊김"),
    (re.compile(r"\bearly\s+(fail|shutdown)|premature", re.I), "조기 종료"),
    (re.compile(r"\b(skin|adhesive|rash|irritation|allergic|itch|burn)", re.I), "피부 반응/접착부 문제"),
    (re.compile(r"\b(bleed|bleeding|insertion\s+pain|pain)", re.I), "삽입부 통증/출혈"),
    (re.compile(r"\b(hypo|low\s+blood\s+sugar|hypoglyc)", re.I), "저혈당"),
    (re.compile(r"\b(hyper|high\s+blood\s+sugar|hyperglyc|dka|ketoacidosis)", re.I), "고혈당/케토산증"),
    (re.compile(r"\b(transmitter|receiver)\s+(fail|error|malfunction|not\s+charg)", re.I), "송수신기 문제"),
    (re.compile(r"\b(app|application|phone|bluetooth|pairing|connect)", re.I), "앱/블루투스 연결 문제"),
    (re.compile(r"\balarm|alert(\s+fail|\s+not)", re.I), "알람 미작동"),
    (re.compile(r"\b(fell\s+off|fall\s+off|detached|dislodge)", re.I), "센서 탈락"),
    (re.compile(r"\bhospital|er\b|emergency\s+room|admit", re.I), "응급실/입원"),
    (re.compile(r"\bdeath|died|fatal", re.I), "사망"),
    (re.compile(r"\bseizure|unconscious|coma|syncope", re.I), "발작/의식저하"),
]

# 제조사 대응/조사 결과에서 자주 등장하는 영문 패턴 → 한글
_RESPONSE_MAP: List[Tuple[re.Pattern, str]] = [
    (re.compile(r"replaced|replacement\s+(was\s+)?(sent|provided|shipped)", re.I), "교체품 제공"),
    (re.compile(r"refund", re.I), "환불 처리"),
    (re.compile(r"no\s+(device\s+)?returned|device\s+not\s+returned|not\s+available\s+for\s+(evaluation|analysis)", re.I), "제품 미반납 → 분석 불가"),
    (re.compile(r"evaluation\s+(was\s+)?(completed|performed)|analysis\s+(was\s+)?(completed|performed)", re.I), "제조사 분석 수행"),
    (re.compile(r"no\s+(anomaly|defect|malfunction|root\s+cause)\s+(was\s+)?(found|identified|determined)", re.I), "이상/결함 미발견"),
    (re.compile(r"root\s+cause\s+(was\s+)?(identified|determined)", re.I), "원인 규명"),
    (re.compile(r"investigation\s+(is\s+)?(ongoing|pending|continuing)", re.I), "조사 진행 중"),
    (re.compile(r"corrective\s+action|capa\b", re.I), "시정조치(CAPA)"),
    (re.compile(r"user\s+error|improper\s+use|instructions?\s+for\s+use|ifu\s+", re.I), "사용자 오사용 가능성"),
    (re.compile(r"manufacturing\s+(defect|issue)", re.I), "제조 결함 확인"),
    (re.compile(r"consistent\s+with\s+(expected|known)|within\s+specification", re.I), "사양 범위 내"),
    (re.compile(r"further\s+information\s+(is\s+)?(requested|needed)|awaiting", re.I), "추가 정보 대기"),
]


def _find_keywords(text: Optional[str], mapping: List[Tuple[re.Pattern, str]]) -> List[str]:
    """정규식 매핑에서 텍스트에 나타나는 모든 한글 태그를 중복 없이 반환."""
    if not text:
        return []
    found: List[str] = []
    for pattern, ko in mapping:
        if pattern.search(text):
            if ko not in found:
                found.append(ko)
    return found


def summarize_korean(
    event_type: Optional[str],
    patient_problems: Optional[str],
    product_problems: Optional[str],
    event_description: Optional[str],
    manufacturer_narrative: Optional[str],
) -> Tuple[str, str, str]:
    """FDA 원문에서 소비자 불만/제조사 대응/결론 3줄 한글 요약을 생성."""

    # 1) 소비자 불만사항
    complaint_parts: List[str] = []
    if patient_problems:
        complaint_parts.append(f"환자 문제: {patient_problems}")
    if product_problems:
        complaint_parts.append(f"제품 문제: {product_problems}")
    symptom_tags = _find_keywords(event_description, _SYMPTOM_MAP)
    if symptom_tags:
        complaint_parts.append("요약 키워드: " + ", ".join(symptom_tags))
    complaint = " | ".join(complaint_parts) if complaint_parts else "상세 불만 내용 없음"

    # 2) 제조사 대응
    response_tags = _find_keywords(manufacturer_narrative, _RESPONSE_MAP)
    if response_tags:
        response = ", ".join(response_tags)
    elif manufacturer_narrative:
        # 너무 길면 앞 200자만
        response = (manufacturer_narrative[:200] + "…") if len(manufacturer_narrative) > 200 else manufacturer_narrative
    else:
        response = "제조사 내러티브 없음"

    # 3) 결론 (event_type 기반 + 간이 판정)
    et = (event_type or "").lower()
    if "death" in et:
        verdict = "사망 사례 – FDA 등급 최상위, 즉시 검토 필요"
    elif "injury" in et:
        verdict = "상해 사례 – 임상적 영향 확인 필요"
    elif "malfunction" in et:
        if "이상/결함 미발견" in response_tags:
            verdict = "기기 오작동 보고 – 제조사 분석 결과 결함 미확인"
        elif "조사 진행 중" in response_tags or "추가 정보 대기" in response_tags:
            verdict = "기기 오작동 보고 – 조사/추가정보 대기 중"
        else:
            verdict = "기기 오작동 보고"
    elif "other" in et:
        verdict = "기타 유형 보고"
    else:
        verdict = "분류 미상"

    return complaint, response, verdict


# ---------------------------------------------------------------------------
# FDA API 호출
# ---------------------------------------------------------------------------

def _build_or_clause(field: str, values: List[str]) -> str:
    """field:"v1" OR field:"v2" ... 를 괄호로 묶어 반환."""
    return "(" + " OR ".join(f'{field}:"{v}"' for v in values) + ")"


def _date_clause(start: datetime, end: datetime) -> str:
    return f"date_received:[{start.strftime('%Y%m%d')} TO {end.strftime('%Y%m%d')}]"


def _filter_clauses() -> List[str]:
    """EVENT_TYPES / ONLY_ADVERSE_EVENTS 기반 부작용 필터 절을 반환."""
    clauses: List[str] = []
    if EVENT_TYPES:
        clauses.append(_build_or_clause("event_type", EVENT_TYPES))
    if ONLY_ADVERSE_EVENTS:
        clauses.append("adverse_event_flag:Y")
    return clauses


def _compose_query(primary_clause: str, start: datetime, end: datetime) -> str:
    parts = [primary_clause, _date_clause(start, end)] + _filter_clauses()
    return " AND ".join(parts).replace(" ", "+")


def build_search_query(start: datetime, end: datetime) -> str:
    """메인 검색 쿼리: (manufacturer OR brand) + 부작용 필터."""
    clauses: List[str] = []
    if SEARCH_MANUFACTURERS:
        clauses.append(_build_or_clause("device.manufacturer_d_name", SEARCH_MANUFACTURERS))
    if SEARCH_BRANDS:
        clauses.append(_build_or_clause("device.brand_name", SEARCH_BRANDS))
    primary = " OR ".join(f"({c})" for c in clauses if c)
    return _compose_query(primary, start, end)


def build_fallback_queries(start: datetime, end: datetime) -> List[Tuple[str, str]]:
    """메인 쿼리가 0건일 때 시도할 보조 쿼리 목록.
    각 항목: (레이블, search 문자열)
    """
    queries: List[Tuple[str, str]] = []

    # (1) openfda.brand_name — openFDA 가 정규화한 브랜드명 필드
    queries.append((
        "openfda.brand_name",
        _compose_query(_build_or_clause("openfda.brand_name", CGM_BRANDS), start, end),
    ))

    # (2) 제조사명
    if FALLBACK_MANUFACTURERS:
        queries.append((
            "device.manufacturer_d_name",
            _compose_query(
                _build_or_clause("device.manufacturer_d_name", FALLBACK_MANUFACTURERS),
                start, end,
            ),
        ))

    # (3) 일반명 (CGM 전체)
    if FALLBACK_GENERIC_NAMES:
        queries.append((
            "device.generic_name",
            _compose_query(
                _build_or_clause("device.generic_name", FALLBACK_GENERIC_NAMES),
                start, end,
            ),
        ))

    return queries


def _probe_query(search: str, api_key: Optional[str], verbose: bool = False) -> int:
    """해당 search 로 총 몇 건이 있는지 probe (limit=1). 0이면 결과 없음."""
    url = f"{OPENFDA_URL}?search={search}&limit=1"
    if api_key:
        url += f"&api_key={api_key}"
    if verbose:
        log.info("[probe] URL: %s", url)
    try:
        resp = requests.get(url, timeout=REQUEST_TIMEOUT)
    except requests.RequestException as e:
        log.error("[probe] 요청 실패: %s", e)
        return -1
    if resp.status_code == 404:
        return 0
    if resp.status_code != 200:
        log.warning("[probe] HTTP %d: %s", resp.status_code, resp.text[:200])
        return -1
    total = resp.json().get("meta", {}).get("results", {}).get("total", 0)
    return total


class APIKeyRequired(Exception):
    """openFDA 가 API 키를 요구할 때 (속도 제한 / 요청 차단) 던지는 예외."""
    pass


def _parse_date_yyyymmdd(s: Optional[str]) -> Optional[datetime]:
    """openFDA 의 'YYYYMMDD' 문자열을 datetime 으로."""
    if not s:
        return None
    s = str(s).strip()
    if len(s) == 8 and s.isdigit():
        try:
            return datetime.strptime(s, "%Y%m%d")
        except ValueError:
            return None
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        try:
            return datetime.strptime(s, "%Y-%m-%d")
        except ValueError:
            return None
    return None


def _fetch_page(
    session: requests.Session,
    search: str,
    skip: int,
    api_key: Optional[str],
    verbose: bool,
    sort: str = "date_received:asc",
) -> Tuple[int, List[Dict[str, Any]]]:
    """단일 페이지 요청. (status_code, results) 반환.
    429/5xx/네트워크 오류는 내부 재시도.
    403 API_KEY_MISSING 은 APIKeyRequired 예외로 올림.
    """
    url = f"{OPENFDA_URL}?search={search}&limit={PAGE_SIZE}&skip={skip}&sort={sort}"
    if api_key:
        url += f"&api_key={api_key}"
    if verbose:
        log.info("  URL: %s", url)
    retries = 0
    while True:
        try:
            resp = session.get(url, timeout=REQUEST_TIMEOUT)
        except requests.RequestException as e:
            retries += 1
            if retries > 3:
                log.error("  요청 실패 3회 초과: %s", e)
                return 0, []
            log.warning("  요청 실패 (%d/3): %s — 5초 후 재시도", retries, e)
            time.sleep(5)
            continue

        if resp.status_code == 200:
            return 200, resp.json().get("results", []) or []
        if resp.status_code == 404:
            # openFDA 는 결과 없음 = 404 반환
            return 404, []
        if resp.status_code == 429:
            log.warning("  속도 제한(429). 60초 대기.")
            time.sleep(60)
            continue
        if resp.status_code == 403:
            txt = resp.text[:300]
            if "API_KEY" in txt or "api_key" in txt or "rate limit" in txt.lower():
                raise APIKeyRequired(txt)
            log.error("  HTTP 403: %s", txt)
            return 403, []
        if 500 <= resp.status_code < 600:
            retries += 1
            if retries > 3:
                log.error("  HTTP %d 지속: %s", resp.status_code, resp.text[:200])
                return resp.status_code, []
            log.warning("  HTTP %d (%d/3). 10초 후 재시도.", resp.status_code, retries)
            time.sleep(10)
            continue
        log.error("  HTTP %d: %s", resp.status_code, resp.text[:300])
        return resp.status_code, []


def _fetch_cursor_window(
    primary_clause: str,
    window_start: datetime,
    window_end: datetime,
    api_key: Optional[str],
    verbose: bool,
    session: requests.Session,
) -> Iterable[Dict[str, Any]]:
    """[window_start, window_end] 구간을 커서(date_received:asc) 기반으로 수집.

    페이지네이션 방식:
      - openFDA 는 skip 누적 25,000 까지만 허용.
      - 한 창(window) 안에서 sort=date_received:asc 로 1,000씩 끌어오며
        skip 가 25,000 에 근접하면 '마지막 레코드 날짜 + 1일' 로 창을 재설정해
        skip 을 0 으로 리셋 (커서 이동).
      - 같은 날짜에 겹치는 레코드가 있어도 PRIMARY KEY(report_number) 가
        중복을 자동으로 막아줌.
    """
    page_sleep = SLEEP_WITH_KEY if api_key else SLEEP_WITHOUT_KEY
    cur_start = window_start
    safe_skip_limit = MAX_SKIP - PAGE_SIZE  # 24,000 — 25,000 도달 직전 cursor 이동

    while cur_start <= window_end:
        search = _compose_query(primary_clause, cur_start, window_end)
        skip = 0
        last_date_received: Optional[str] = None
        page_count_in_window = 0
        yielded_in_window = 0

        while True:
            log.info("  [cursor %s~%s] skip=%d",
                     cur_start.strftime("%Y-%m-%d"),
                     window_end.strftime("%Y-%m-%d"),
                     skip)
            status, results = _fetch_page(
                session, search, skip, api_key, verbose and skip == 0
            )
            if status == 404 or not results:
                # 이 창은 여기까지
                return
            if status != 200:
                # 회복 불가 에러: 바깥 for 문이 다음 창으로 진행하지 않도록 종료
                return

            for ev in results:
                yield ev
                yielded_in_window += 1
                last_date_received = ev.get("date_received") or last_date_received

            page_count_in_window += 1
            if len(results) < PAGE_SIZE:
                # 이 창 완전히 소진 — 다음 창 불필요
                return

            skip += PAGE_SIZE

            # skip 한도에 근접하면 커서 이동 (새 창 열기)
            if skip > safe_skip_limit:
                parsed = _parse_date_yyyymmdd(last_date_received)
                if parsed is None:
                    log.warning(
                        "  ⚠ skip 한도 도달했지만 last_date_received 파싱 실패. 중단."
                    )
                    return
                # 같은 날짜 레코드가 더 있을 수 있으니 +0일이 아닌 현재 날짜부터 재시작.
                # (PRIMARY KEY 로 중복 자동 배제됨)
                next_start = parsed
                if next_start <= cur_start:
                    # 진전이 없으면 무한 루프 방지를 위해 하루 전진
                    next_start = cur_start + timedelta(days=1)
                log.info(
                    "  ↻ skip 한도 근접 — 커서 이동: %s → %s (이 창 %d건 수집)",
                    cur_start.strftime("%Y-%m-%d"),
                    next_start.strftime("%Y-%m-%d"),
                    yielded_in_window,
                )
                cur_start = next_start
                break  # inner while: 새 창으로 재진입

            time.sleep(page_sleep)
        # continue outer while: 새 window 로 재진입


def fetch_events(
    start: datetime,
    end: datetime,
    api_key: Optional[str] = None,
    verbose: bool = False,
) -> Iterable[Dict[str, Any]]:
    """openFDA MAUDE API 를 호출해 이벤트를 순차 반환.

    로직 (커서 기반):
      1) 메인 쿼리(device.brand_name + 부작용 필터)로 총 건수 확인
      2) 0건이면 fallback 쿼리(openfda.brand_name / manufacturer_d_name /
         generic_name) 순차 시도
      3) sort=date_received:asc 로 페이지네이션. skip 한도 근접 시
         "마지막 레코드의 date_received 부터" 로 창을 다시 열어 커서 이동.
         → 월 단위 분할 없이도 25,000건 한도 우회 가능.
      4) 403 API_KEY_MISSING 은 APIKeyRequired 예외로 상위에 전달.
    """
    # 1) 메인 쿼리 먼저 시도
    main_search = build_search_query(start, end)
    log.info("메인 쿼리 시도 ((device.manufacturer_d_name OR device.brand_name) + 부작용 필터)")
    if verbose:
        log.info("search=%s", main_search)
    total = _probe_query(main_search, api_key, verbose=verbose)
    log.info("메인 쿼리 결과 총 건수: %s", f"{total:,}" if total >= 0 else "조회 실패")

    chosen_primary: Optional[str] = None
    if total > 0:
        chosen_primary = " OR ".join([
            f"({_build_or_clause('device.manufacturer_d_name', SEARCH_MANUFACTURERS)})",
            f"({_build_or_clause('device.brand_name', SEARCH_BRANDS)})",
        ])

    # 2) 0건이면 fallback
    if total == 0 and USE_FALLBACK_FIELDS:
        log.warning("메인 쿼리 0건 → fallback 쿼리 시도")
        fallback_primaries = [
            ("openfda.brand_name", _build_or_clause("openfda.brand_name", SEARCH_BRANDS)),
        ]
        if FALLBACK_MANUFACTURERS:
            fallback_primaries.append((
                "device.manufacturer_d_name",
                _build_or_clause("device.manufacturer_d_name", FALLBACK_MANUFACTURERS),
            ))
        if FALLBACK_GENERIC_NAMES:
            fallback_primaries.append((
                "device.generic_name",
                _build_or_clause("device.generic_name", FALLBACK_GENERIC_NAMES),
            ))
        for label, primary in fallback_primaries:
            q = _compose_query(primary, start, end)
            n = _probe_query(q, api_key, verbose=verbose)
            log.info("Fallback [%s] → 총 %s 건", label,
                     f"{n:,}" if n >= 0 else "조회 실패")
            if n > 0:
                chosen_primary = primary
                total = n
                break

    if total <= 0 or chosen_primary is None:
        log.warning("=" * 60)
        log.warning("검색 결과 0건. 가능한 원인:")
        log.warning("  1) FDA MAUDE 는 접수 후 DB 반영까지 수 개월 지연")
        log.warning("     (조회 범위: %s ~ %s)",
                    start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d"))
        log.warning("  2) 부작용 필터가 너무 엄격: EVENT_TYPES / ONLY_ADVERSE_EVENTS 확인")
        log.warning("  3) 브랜드명 오탈자 — CGM_BRANDS 확인")
        log.warning("=" * 60)
        return

    session = requests.Session()
    log.info("커서 기반 수집 시작 (sort=date_received:asc, 한도 우회 가능)")

    yield from _fetch_cursor_window(
        chosen_primary, start, end, api_key, verbose, session
    )

    log.info("수집 종료.")


# ---------------------------------------------------------------------------
# 필드 평탄화
# ---------------------------------------------------------------------------

def _first(lst: Optional[List[Any]]) -> Optional[Any]:
    if lst and isinstance(lst, list):
        return lst[0]
    return None


def _join_unique(values: Iterable[Any]) -> str:
    seen = []
    for v in values:
        if v is None:
            continue
        if isinstance(v, list):
            for x in v:
                if x and x not in seen:
                    seen.append(str(x))
        else:
            if v and v not in seen:
                seen.append(str(v))
    # 코드명 내부에 쉼표가 포함될 수 있어 구분자는 세미콜론(;)을 사용한다.
    return "; ".join(seen)


def _normalize_date(d: Optional[str]) -> Optional[str]:
    """openFDA 는 'YYYYMMDD' 형태 문자열. 'YYYY-MM-DD' 로 변환."""
    if not d:
        return None
    d = str(d).strip()
    if len(d) == 8 and d.isdigit():
        return f"{d[0:4]}-{d[4:6]}-{d[6:8]}"
    return d


def _patient_age(p: Dict[str, Any]) -> Optional[str]:
    """환자 나이 필드를 사람이 읽을 수 있는 문자열로 변환."""
    age = p.get("patient_age")
    unit = p.get("patient_age_unit") or p.get("age_unit")
    if age is None or age == "":
        return None
    if unit:
        return f"{age} {unit}".strip()
    return str(age)


def _patient_weight(p: Dict[str, Any]) -> Optional[str]:
    w = p.get("patient_weight")
    unit = p.get("patient_weight_unit")
    if w is None or w == "":
        return None
    return f"{w} {unit}".strip() if unit else str(w)


def _manufacturer_country(dev: Dict[str, Any]) -> Optional[str]:
    """openFDA device 블록에서 제조사 국가 추출 (여러 후보 필드 탐색)."""
    for key in (
        "manufacturer_d_country",
        "manufacturer_g1_country",
        "manufacturer_country",
    ):
        v = dev.get(key)
        if v:
            return str(v)
    return None


def _join_narrative_parts(parts: List[str]) -> Optional[str]:
    cleaned = [str(p).strip() for p in parts if str(p or "").strip()]
    if not cleaned:
        return None
    if len(cleaned) == 1:
        return cleaned[0]
    return "\n\n".join(f"[{i}] {txt}" for i, txt in enumerate(cleaned, start=1))


def flatten_event(ev: Dict[str, Any]) -> Dict[str, Any]:
    """openFDA MAUDE 이벤트 JSON 을 평탄화된 dict 로 변환."""
    devices = ev.get("device", []) or []
    dev = _first(devices) or {}

    patients = ev.get("patient", []) or []
    patient_problems = _join_unique(p.get("patient_problems") for p in patients)

    # 환자 인구통계 (첫 번째 환자 기준)
    pat0: Dict[str, Any] = patients[0] if patients else {}
    patient_sex = pat0.get("patient_sex") or ev.get("patient_sex")
    patient_ethnicity = pat0.get("patient_ethnicity") or ev.get("patient_ethnicity")
    patient_race = pat0.get("patient_race") or ev.get("patient_race")

    product_problems = _join_unique([ev.get("product_problems")])

    # mdr_text may contain repeated narrative blocks as the case is updated.
    event_desc_parts: List[str] = []
    mfr_narr_parts: List[str] = []
    add_mfr_narr_parts: List[str] = []
    for t in ev.get("mdr_text", []) or []:
        if not isinstance(t, dict):
            continue
        code = str(t.get("text_type_code") or t.get("type") or "").lower()
        txt = str(t.get("text") or t.get("text_value") or "").strip()
        if not txt:
            continue
        if "additional" in code and "manufacturer" in code:
            add_mfr_narr_parts.append(txt)
        elif "manufacturer" in code:
            mfr_narr_parts.append(txt)
        else:
            event_desc_parts.append(txt)

    report_number = ev.get("report_number") or ev.get("mdr_report_key") or ""

    return {
        "report_number": str(report_number),
        "event_type": ev.get("event_type"),
        "date_received": _normalize_date(ev.get("date_received")),
        "date_of_event": _normalize_date(ev.get("date_of_event")),
        "date_report": _normalize_date(ev.get("date_report")),
        "brand_name": dev.get("brand_name"),
        "device_category": _resolve_device_category(dev.get("brand_name")),
        "generic_name": dev.get("generic_name"),
        "manufacturer_name": (
            dev.get("manufacturer_d_name")
            or ev.get("manufacturer_name")
            or dev.get("manufacturer_g1_name")
        ),
        "manufacturer_country": _manufacturer_country(dev),
        "model_number": dev.get("model_number"),
        "product_code": dev.get("device_report_product_code"),
        "source_type": _first(ev.get("source_type")) if isinstance(ev.get("source_type"), list) else ev.get("source_type"),
        "report_source_code": ev.get("report_source_code"),
        "type_of_report": _join_unique([ev.get("type_of_report")]),
        "patient_age": _patient_age(pat0),
        "patient_sex": patient_sex,
        "patient_ethnicity": patient_ethnicity,
        "patient_race": patient_race,
        "patient_weight": _patient_weight(pat0),
        "patient_problems": patient_problems,
        "product_problems": product_problems,
        "event_description": _join_narrative_parts(event_desc_parts),
        "manufacturer_narrative": _join_narrative_parts(mfr_narr_parts),
        "additional_manufacturer_narrative": _join_narrative_parts(add_mfr_narr_parts),
        "raw_json": json.dumps(ev, ensure_ascii=False),
    }


# ---------------------------------------------------------------------------
# SQLite 저장
# ---------------------------------------------------------------------------

# 최신 스키마에서 maude_reports 테이블이 가져야 하는 칼럼 전체 목록.
# 구 DB 마이그레이션 시, 여기에 있는데 실제 DB 에 없는 칼럼은 ALTER TABLE 로 추가한다.
# (report_number 은 PRIMARY KEY 이므로 여기에 넣지 않는다 — 반드시 있어야 함)
_TARGET_COLUMNS: List[Tuple[str, str]] = [
    ("event_type",             "TEXT"),
    ("date_received",          "TEXT"),
    ("date_of_event",          "TEXT"),
    ("date_report",            "TEXT"),
    ("brand_name",             "TEXT"),
    ("device_category",        "TEXT"),
    ("generic_name",           "TEXT"),
    ("manufacturer_name",      "TEXT"),
    ("manufacturer_country",   "TEXT"),
    ("model_number",           "TEXT"),
    ("product_code",           "TEXT"),
    ("source_type",            "TEXT"),
    ("report_source_code",     "TEXT"),
    ("type_of_report",         "TEXT"),
    ("patient_age",            "TEXT"),
    ("patient_sex",            "TEXT"),
    ("patient_ethnicity",      "TEXT"),
    ("patient_race",           "TEXT"),
    ("patient_weight",         "TEXT"),
    ("patient_problems",       "TEXT"),
    ("product_problems",       "TEXT"),
    ("event_description",      "TEXT"),
    ("manufacturer_narrative", "TEXT"),
    ("additional_manufacturer_narrative", "TEXT"),
    ("summary_complaint_kr",   "TEXT"),
    ("summary_response_kr",    "TEXT"),
    ("summary_conclusion_kr",  "TEXT"),
    ("raw_json",               "TEXT"),
    ("collected_at",           "TEXT"),
]

# 하위 호환
_ADDED_COLUMNS = _TARGET_COLUMNS


def _migrate_schema(conn: sqlite3.Connection) -> None:
    """구 DB 에 누락된 칼럼을 모두 ALTER TABLE 로 추가 (비파괴 마이그레이션).

    어떤 버전의 구 스키마에서 올라오든 안전하게 동작하도록 '목표 스키마' 의
    모든 칼럼을 대상으로 검사한다. 이미 존재하는 칼럼은 건너뛰므로 비용은 거의 없다.
    """
    existing = {row[1] for row in conn.execute("PRAGMA table_info(maude_reports)").fetchall()}
    for col, ctype in _TARGET_COLUMNS:
        if col not in existing:
            log.info("DB 마이그레이션: maude_reports.%s 칼럼 추가", col)
            conn.execute(f"ALTER TABLE maude_reports ADD COLUMN {col} {ctype}")
    conn.commit()


def _normalize_match_text(value: Optional[object]) -> str:
    return str(value or "").strip().upper()


def _is_cgm_brand_like(row: sqlite3.Row) -> bool:
    fields = [row["brand_name"], row["generic_name"], row["manufacturer_name"], row["product_code"]]
    combined = " | ".join(_normalize_match_text(v) for v in fields if _normalize_match_text(v))
    if not combined:
        return False
    needles = ["DEXCOM", "FREESTYLE LIBRE", "FREE STYLE LIBRE", "LIBRE", "SENSEN"]
    return any(needle in combined for needle in needles)


def _log_null_category_remnants(conn: sqlite3.Connection, limit: int = 30) -> None:
    rows = conn.execute(
        """
        SELECT COALESCE(NULLIF(brand_name,''), '미상') AS brand_name,
               COALESCE(NULLIF(manufacturer_name,''), '미상') AS manufacturer_name,
               COALESCE(NULLIF(product_code,''), '미상') AS product_code,
               COUNT(*) AS cnt
          FROM maude_reports
         WHERE device_category IS NULL
         GROUP BY 1, 2, 3
         ORDER BY cnt DESC, brand_name, manufacturer_name, product_code
         LIMIT ?
        """,
        (limit,),
    ).fetchall()
    if not rows:
        log.info("남은 device_category NULL row 없음")
        return
    log.info("남은 device_category NULL row 상위 %d개:", limit)
    for brand_name, manufacturer_name, product_code, cnt in rows:
        log.info("  - brand=%s | manufacturer=%s | product_code=%s | count=%d", brand_name, manufacturer_name, product_code, cnt)


def backfill_device_category(conn: sqlite3.Connection) -> None:
    """과거 NULL device_category 를 CGM / Pump 으로 비파괴 백필한다."""
    before_null = conn.execute(
        "SELECT COUNT(*) FROM maude_reports WHERE device_category IS NULL"
    ).fetchone()[0]
    log.info("백필 전 device_category NULL count: %d", before_null)

    params = [
        "%DEXCOM%", "%DEXCOM%", "%DEXCOM%", "%DEXCOM%",
        "%FREESTYLE LIBRE%", "%FREESTYLE LIBRE%", "%FREESTYLE LIBRE%", "%FREESTYLE LIBRE%",
        "%FREE STYLE LIBRE%", "%FREE STYLE LIBRE%", "%FREE STYLE LIBRE%", "%FREE STYLE LIBRE%",
        "%LIBRE%", "%LIBRE%", "%LIBRE%", "%LIBRE%",
        "%SENSEN%", "%SENSEN%", "%SENSEN%", "%SENSEN%",
    ]
    conn.execute(
        """
        UPDATE maude_reports
           SET device_category = 'CGM'
         WHERE device_category IS NULL
           AND (
                UPPER(COALESCE(brand_name, '')) LIKE ?
                OR UPPER(COALESCE(generic_name, '')) LIKE ?
                OR UPPER(COALESCE(manufacturer_name, '')) LIKE ?
                OR UPPER(COALESCE(product_code, '')) LIKE ?
                OR UPPER(COALESCE(brand_name, '')) LIKE ?
                OR UPPER(COALESCE(generic_name, '')) LIKE ?
                OR UPPER(COALESCE(manufacturer_name, '')) LIKE ?
                OR UPPER(COALESCE(product_code, '')) LIKE ?
                OR UPPER(COALESCE(brand_name, '')) LIKE ?
                OR UPPER(COALESCE(generic_name, '')) LIKE ?
                OR UPPER(COALESCE(manufacturer_name, '')) LIKE ?
                OR UPPER(COALESCE(product_code, '')) LIKE ?
                OR UPPER(COALESCE(brand_name, '')) LIKE ?
                OR UPPER(COALESCE(generic_name, '')) LIKE ?
                OR UPPER(COALESCE(manufacturer_name, '')) LIKE ?
                OR UPPER(COALESCE(product_code, '')) LIKE ?
                OR UPPER(COALESCE(brand_name, '')) LIKE ?
                OR UPPER(COALESCE(generic_name, '')) LIKE ?
                OR UPPER(COALESCE(manufacturer_name, '')) LIKE ?
                OR UPPER(COALESCE(product_code, '')) LIKE ?
           )
        """,
        params,
    )
    conn.commit()
    after_null = conn.execute(
        "SELECT COUNT(*) FROM maude_reports WHERE device_category IS NULL"
    ).fetchone()[0]
    counts = conn.execute(
        """
        SELECT COALESCE(device_category, 'NULL') AS device_category, COUNT(*)
          FROM maude_reports
         GROUP BY COALESCE(device_category, 'NULL')
         ORDER BY COUNT(*) DESC
        """
    ).fetchall()
    log.info("백필 후 device_category 별 count:")
    for category, cnt in counts:
        log.info("  - %s: %d", category, cnt)
    log.info("백필 후 device_category NULL count: %d", after_null)
    _log_null_category_remnants(conn)


def init_db(path: Path = DB_PATH) -> sqlite3.Connection:
    """DB 초기화.

    순서가 중요하다:
      1) 테이블을 먼저 보장한다 (신규 DB 에서만 실제로 생성, 구 DB 는 스킵)
      2) 구 DB 에 없는 신규 칼럼을 ALTER TABLE 로 추가 (마이그레이션)
      3) 그 다음에야 인덱스를 만든다 — 이제는 모든 칼럼이 존재함이 보장됨

    과거에는 세 가지를 한 블록에 합쳐 `executescript` 로 돌렸는데,
    구 DB 에서 `CREATE INDEX ... ON maude_reports(manufacturer_country)` 가
    ALTER 보다 먼저 실행돼 "no such column" 으로 전체가 실패했다.
    """
    conn = sqlite3.connect(path, timeout=30)
    conn.execute("PRAGMA busy_timeout = 30000")
    try:
        conn.execute("PRAGMA journal_mode=WAL")
    except Exception:
        pass
    # 1) 테이블
    conn.executescript(SCHEMA_TABLES_SQL)
    # 2) 구 DB 마이그레이션 (신규 칼럼 추가)
    _migrate_schema(conn)
    # 3) 인덱스 (모든 칼럼 존재 보장됨)
    conn.executescript(SCHEMA_INDEXES_SQL)
    # 4) Legacy FTS triggers are disabled: the dashboard uses SQL filters, not FTS.
    # Keeping these triggers active duplicates large text/raw_json data and can create
    # multi-GB rollback journals during startup sync.
    conn.executescript(LEGACY_FTS_TRIGGER_DROP_SQL)
    conn.commit()
    backfill_device_category(conn)
    conn.execute("PRAGMA optimize;")
    conn.commit()
    return conn


# ---------------------------------------------------------------------------
# 체크포인트 관리 (증분 수집용)
# ---------------------------------------------------------------------------

def _brand_key(brands: List[str]) -> str:
    """현재 CGM_BRANDS 리스트를 정규화된 키 문자열로 변환.
    브랜드 목록이 바뀌면 키가 달라져 새 체크포인트가 쓰인다 (= 자동 재초기화).
    """
    return "|".join(sorted(b.strip().upper() for b in brands))


def get_checkpoint(conn: sqlite3.Connection, brands: List[str]) -> Optional[Dict[str, Any]]:
    """해당 브랜드 조합의 체크포인트를 반환. 없으면 None (= 최초 실행)."""
    key = _brand_key(brands)
    row = conn.execute(
        "SELECT brand_key, last_search_end, last_run_at, last_inserted, "
        "total_collected, first_run_at, note "
        "FROM collection_checkpoint WHERE brand_key = ?",
        (key,),
    ).fetchone()
    if not row:
        return None
    return {
        "brand_key": row[0],
        "last_search_end": row[1],
        "last_run_at": row[2],
        "last_inserted": row[3],
        "total_collected": row[4],
        "first_run_at": row[5],
        "note": row[6],
    }


def save_checkpoint(
    conn: sqlite3.Connection,
    brands: List[str],
    search_end: datetime,
    inserted_count: int,
    note: str = "",
) -> None:
    """이번 실행의 결과를 체크포인트에 저장.
    - search_end: 이번에 조회한 종료일 (다음 실행의 기준점)
    - inserted_count: 이번에 upsert 된 건수
    """
    key = _brand_key(brands)
    now = datetime.now().isoformat(timespec="seconds")
    end_str = search_end.strftime("%Y-%m-%d")

    existing = get_checkpoint(conn, brands)
    if existing is None:
        conn.execute(
            "INSERT INTO collection_checkpoint "
            "(brand_key, last_search_end, last_run_at, last_inserted, "
            " total_collected, first_run_at, note) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (key, end_str, now, inserted_count, inserted_count, now, note),
        )
    else:
        total = (existing.get("total_collected") or 0) + inserted_count
        conn.execute(
            "UPDATE collection_checkpoint SET "
            "last_search_end = ?, last_run_at = ?, "
            "last_inserted = ?, total_collected = ?, note = ? "
            "WHERE brand_key = ?",
            (end_str, now, inserted_count, total, note, key),
        )
    conn.commit()


def get_latest_date_received(conn: sqlite3.Connection) -> Optional[datetime]:
    """DB 에 저장된 date_received 의 최댓값(YYYY-MM-DD)을 datetime 으로 반환."""
    row = conn.execute(
        "SELECT MAX(date_received) FROM maude_reports "
        "WHERE date_received IS NOT NULL AND date_received GLOB '????-??-??'"
    ).fetchone()
    if not row or not row[0]:
        return None
    try:
        return datetime.strptime(row[0], "%Y-%m-%d")
    except ValueError:
        log.warning("DB 최신 date_received 파싱 실패: %s", row[0])
        return None


def upsert_report(conn: sqlite3.Connection, row: Dict[str, Any]) -> bool:
    """단일 보고서를 DB 에 upsert. 신규 삽입이면 True, 이미 있으면 False."""
    if not row.get("report_number"):
        return False

    # 한글 요약
    manufacturer_text_for_summary = "\n\n".join(
        p for p in [
            row.get("manufacturer_narrative"),
            row.get("additional_manufacturer_narrative"),
        ] if p
    ) or None
    complaint_kr, response_kr, conclusion_kr = summarize_korean(
        row.get("event_type"),
        row.get("patient_problems"),
        row.get("product_problems"),
        row.get("event_description"),
        manufacturer_text_for_summary,
    )

    row["summary_complaint_kr"] = complaint_kr
    row["summary_response_kr"] = response_kr
    row["summary_conclusion_kr"] = conclusion_kr
    row["collected_at"] = datetime.now().isoformat(timespec="seconds")

    columns = [
        "report_number", "event_type", "date_received", "date_of_event", "date_report",
        "brand_name", "device_category", "generic_name", "manufacturer_name", "manufacturer_country",
        "model_number", "product_code",
        "source_type", "report_source_code",
        "type_of_report",
        "patient_age", "patient_sex", "patient_ethnicity", "patient_race", "patient_weight",
        "patient_problems", "product_problems",
        "event_description", "manufacturer_narrative", "additional_manufacturer_narrative",
        "summary_complaint_kr", "summary_response_kr", "summary_conclusion_kr",
        "raw_json", "collected_at",
    ]
    placeholders = ",".join(["?"] * len(columns))
    colnames = ",".join(columns)
    updates = ",".join(f"{c}=excluded.{c}" for c in columns if c != "report_number")

    sql = (
        f"INSERT INTO maude_reports ({colnames}) VALUES ({placeholders}) "
        f"ON CONFLICT(report_number) DO UPDATE SET {updates}"
    )
    cur = conn.execute(sql, [row.get(c) for c in columns])
    # cur.rowcount 는 UPDATE 면 1, INSERT 면 1 이라 구분이 어렵다.
    # 신규 여부는 별도 SELECT 로 판단하지 않고, 대신 반환값을 수집 통계용으로 쓰지 않음.
    return cur.rowcount > 0


# ---------------------------------------------------------------------------
# Excel 내보내기
# ---------------------------------------------------------------------------

# 전체 보고서 시트 칼럼 정의 (라벨, SQL 표현식)
_EXCEL_MAIN_COLUMNS: List[Tuple[str, str]] = [
    ("MAUDE 번호", "report_number"),
    ("EVENT_TYPE", "event_type"),
    ("DATE_RECEIVED", "date_received"),
    ("DATE_OF_EVENT", "date_of_event"),
    ("BRAND_NAME", "brand_name"),
    ("DEVICE_CATEGORY", "device_category"),
    ("GENERIC_NAME", "generic_name"),
    ("MANUFACTURER", "manufacturer_name"),
    ("MANUFACTURER_COUNTRY", "manufacturer_country"),
    ("MODEL_NUMBER", "model_number"),
    ("PRODUCT_CODE", "product_code"),
    ("SOURCE_TYPE", "source_type"),
    ("PATIENT_AGE", "patient_age"),
    ("PATIENT_SEX", "patient_sex"),
    ("PATIENT_ETHNICITY", "patient_ethnicity"),
    ("PATIENT_RACE", "patient_race"),
    ("PATIENT_WEIGHT", "patient_weight"),
    ("PATIENT_PROBLEMS", "patient_problems"),
    ("PRODUCT_PROBLEMS", "product_problems"),
    ("[요약] 소비자 불만", "summary_complaint_kr"),
    ("[요약] 제조사 대응", "summary_response_kr"),
    ("[요약] 결론", "summary_conclusion_kr"),
    ("EVENT_DESCRIPTION (원문)", "event_description"),
    ("MANUFACTURER_NARRATIVE (원문)", "manufacturer_narrative"),
    ("ADDITIONAL_MANUFACTURER_NARRATIVE (원문)", "additional_manufacturer_narrative"),
    ("수집시각", "collected_at"),
]


def _stream_main_sheet(ws, conn: sqlite3.Connection) -> int:
    """전체 보고서 시트를 커서 스트리밍으로 채움. 행 수 반환."""
    labels = [lbl for lbl, _ in _EXCEL_MAIN_COLUMNS]
    select_cols = ", ".join(expr for _, expr in _EXCEL_MAIN_COLUMNS)
    ws.append(labels)

    cur = conn.execute(
        f"SELECT {select_cols} FROM maude_reports "
        f"ORDER BY date_received DESC, report_number DESC"
    )
    rows = 0
    warned: Dict[str, int] = {"count": 0}
    for row in cur:
        report_number = row[0] if row else None
        ws.append(_sanitize_xlsx_row(row, report_number=report_number, warned=warned))
        rows += 1
    return rows


def _append_query_sheet(ws, conn: sqlite3.Connection, sql: str, headers: List[str]) -> None:
    ws.append(headers)
    for row in conn.execute(sql):
        ws.append([_sanitize_xlsx_value(v) for v in row])


def export_excel(conn: sqlite3.Connection, path: Path = EXCEL_PATH) -> None:
    """DB 전체를 Excel 로 내보냄 (openpyxl write_only 모드, 스트리밍).

    수만~수십만 건이어도 메모리 사용이 작고, 각 시트에 자동 필터와 상단 고정이
    적용됩니다. 기존 pandas 기반과 동일한 시트 구성 + 인구통계/국가 칼럼 추가.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("Excel 내보내기에는 openpyxl 이 필요합니다. 'pip install openpyxl' 실행.")
        return

    wb = Workbook(write_only=True)

    # 1) 전체 보고서
    ws_main = wb.create_sheet("전체 보고서")
    total_rows = _stream_main_sheet(ws_main, conn)
    # write_only 모드에서는 auto_filter / freeze_panes 설정 가능
    # (저장 직전까지 cells 를 메모리에 들고 있지 않음)
    if total_rows > 0:
        last_col_letter = get_column_letter(len(_EXCEL_MAIN_COLUMNS))
        ws_main.auto_filter.ref = f"A1:{last_col_letter}{total_rows + 1}"
    ws_main.freeze_panes = "A2"

    # 2) EVENT_TYPE 집계
    ws_evt = wb.create_sheet("EVENT_TYPE 집계")
    _append_query_sheet(
        ws_evt, conn,
        "SELECT event_type, COUNT(*) FROM maude_reports "
        "GROUP BY event_type ORDER BY COUNT(*) DESC",
        ["EVENT_TYPE", "건수"],
    )
    ws_evt.freeze_panes = "A2"

    # 3) BRAND 집계
    ws_brand = wb.create_sheet("BRAND 집계")
    _append_query_sheet(
        ws_brand, conn,
        "SELECT brand_name, COUNT(*) FROM maude_reports "
        "GROUP BY brand_name ORDER BY COUNT(*) DESC",
        ["BRAND_NAME", "건수"],
    )
    ws_brand.freeze_panes = "A2"

    # 4) MANUFACTURER 집계 (국가 포함)
    ws_mfr = wb.create_sheet("제조사 집계")
    _append_query_sheet(
        ws_mfr, conn,
        "SELECT manufacturer_name, manufacturer_country, COUNT(*) FROM maude_reports "
        "GROUP BY manufacturer_name, manufacturer_country ORDER BY COUNT(*) DESC",
        ["MANUFACTURER", "COUNTRY", "건수"],
    )
    ws_mfr.freeze_panes = "A2"

    # 5) 환자 인구통계 집계
    ws_demo = wb.create_sheet("환자 인구통계")
    _append_query_sheet(
        ws_demo, conn,
        """
        SELECT
            COALESCE(NULLIF(patient_sex,''), '미상') AS sex,
            COALESCE(NULLIF(patient_race,''), '미상') AS race,
            COALESCE(NULLIF(patient_ethnicity,''), '미상') AS ethnicity,
            COUNT(*) AS cnt
        FROM maude_reports
        GROUP BY sex, race, ethnicity
        ORDER BY cnt DESC
        """,
        ["PATIENT_SEX", "PATIENT_RACE", "PATIENT_ETHNICITY", "건수"],
    )
    ws_demo.freeze_panes = "A2"

    # 6) 월별 추이
    ws_mon = wb.create_sheet("월별 추이")
    _append_query_sheet(
        ws_mon, conn,
        """
        SELECT substr(date_received,1,7),
               COUNT(*),
               SUM(CASE WHEN event_type='Death'       THEN 1 ELSE 0 END),
               SUM(CASE WHEN event_type='Injury'      THEN 1 ELSE 0 END),
               SUM(CASE WHEN event_type='Malfunction' THEN 1 ELSE 0 END),
               SUM(CASE WHEN event_type='Other'       THEN 1 ELSE 0 END)
        FROM maude_reports
        WHERE date_received IS NOT NULL
        GROUP BY substr(date_received,1,7)
        ORDER BY 1 DESC
        """,
        ["월", "건수", "사망", "상해", "오작동", "기타"],
    )
    ws_mon.freeze_panes = "A2"

    # 열 너비 대략 지정 (write_only 모드에서도 dimensions 설정은 가능)
    _MAIN_WIDTHS = {
        "MAUDE 번호": 14, "EVENT_TYPE": 11, "DATE_RECEIVED": 12, "DATE_OF_EVENT": 12,
        "BRAND_NAME": 28, "GENERIC_NAME": 26, "MANUFACTURER": 26,
        "MANUFACTURER_COUNTRY": 10, "MODEL_NUMBER": 14, "PRODUCT_CODE": 10,
        "SOURCE_TYPE": 14, "PATIENT_AGE": 10, "PATIENT_SEX": 10,
        "PATIENT_ETHNICITY": 14, "PATIENT_RACE": 12, "PATIENT_WEIGHT": 12,
        "PATIENT_PROBLEMS": 30, "PRODUCT_PROBLEMS": 30,
        "[요약] 소비자 불만": 40, "[요약] 제조사 대응": 30, "[요약] 결론": 28,
        "EVENT_DESCRIPTION (원문)": 50, "MANUFACTURER_NARRATIVE (원문)": 50, "ADDITIONAL_MANUFACTURER_NARRATIVE (원문)": 50,
        "수집시각": 18,
    }
    for idx, (label, _) in enumerate(_EXCEL_MAIN_COLUMNS, start=1):
        ws_main.column_dimensions[get_column_letter(idx)].width = _MAIN_WIDTHS.get(label, 14)

    wb.save(str(path))
    log.info("Excel 저장 완료: %s  (전체 보고서 %d건)", path, total_rows)


def export_excel_only(db_path: Path, excel_path: Path) -> int:
    """기존 DB를 열어 Excel만 다시 생성한다."""
    conn = init_db(db_path)
    try:
        log.info("Excel 재생성 모드: DB=%s, Excel=%s", db_path, excel_path)
        export_excel(conn, excel_path)
    finally:
        conn.close()
    return 0


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------

def parse_date(s: str) -> datetime:
    return datetime.strptime(s, "%Y%m%d")


def load_api_key(cli_value: Optional[str]) -> Optional[str]:
    """API 키 우선순위:
       1) --api-key CLI 인자
       2) 환경변수 OPENFDA_API_KEY
       3) 파일: FDA_MAUDE_API_KEY.txt → api_key.txt 순서
          (첫 줄 기준, 공백/주석(#) 제거)
    """
    if cli_value:
        return cli_value.strip()
    env = os.environ.get("OPENFDA_API_KEY")
    if env:
        return env.strip()
    for path in API_KEY_FILES:
        if not path.exists():
            continue
        try:
            for line in path.read_text(encoding="utf-8").splitlines():
                s = line.strip()
                if not s or s.startswith("#"):
                    continue
                log.info("API 키 파일 사용: %s", path.name)
                return s
        except Exception as e:
            log.warning("%s 읽기 실패: %s", path.name, e)
    return None


def run_test_mode(api_key: Optional[str]) -> int:
    """API 연결 및 간단한 쿼리가 동작하는지 확인하는 진단 모드.
    - 브랜드별로 날짜 제한 없이 총 건수 출력
    - 연도별 건수 출력 (FDA MAUDE 데이터 지연 확인용)
    """
    log.info("=" * 60)
    log.info("[TEST MODE] openFDA API 연결 및 쿼리 검증")
    log.info("=" * 60)

    # 1) 단순 브랜드 probe — 날짜 제한 없이 총 건수
    log.info("\n[1] 브랜드별 총 보고서 수 (날짜 무관)")
    for brand in CGM_BRANDS:
        q = f'device.brand_name:"{brand}"'.replace(" ", "+")
        total = _probe_query(q, api_key, verbose=True)
        log.info("  %s: %s 건", brand, f"{total:,}" if total >= 0 else "조회 실패")

    # 2) 연도별 probe — 전체 vs 부작용 필터 적용 비교
    log.info("\n[2] 연도별 건수 (Dexcom + Libre)")
    log.info("  (전체 = 모든 보고서 / 부작용 = EVENT_TYPE %s + adverse_event_flag=%s)",
             EVENT_TYPES or "전체", "Y" if ONLY_ADVERSE_EVENTS else "전체")
    brand_or = _build_or_clause("device.brand_name", SEARCH_BRANDS)
    filter_clause = ""
    fc = _filter_clauses()
    if fc:
        filter_clause = " AND " + " AND ".join(fc)
    years = [2022, 2023, 2024, 2025, 2026]
    for y in years:
        q_all = f'{brand_or} AND date_received:[{y}0101 TO {y}1231]'.replace(" ", "+")
        q_filt = (f'{brand_or} AND date_received:[{y}0101 TO {y}1231]{filter_clause}'
                  ).replace(" ", "+")
        n_all = _probe_query(q_all, api_key, verbose=False)
        n_filt = _probe_query(q_filt, api_key, verbose=False) if fc else n_all
        log.info("  %d년: 전체 %s건 / 부작용 %s건", y,
                 f"{n_all:,}" if n_all >= 0 else "?",
                 f"{n_filt:,}" if n_filt >= 0 else "?")

    # 3) 가장 최근 보고서 1건의 date_received 확인 (DB 최신 데이터 지점)
    log.info("\n[3] 가장 최근 보고서의 date_received (MAUDE DB 최신 지점)")
    q = f'{brand_or}'.replace(" ", "+")
    url = f"{OPENFDA_URL}?search={q}&limit=1&sort=date_received:desc"
    if api_key:
        url += f"&api_key={api_key}"
    try:
        resp = requests.get(url, timeout=REQUEST_TIMEOUT)
        if resp.status_code == 200:
            r = resp.json().get("results", [{}])[0]
            log.info("  최신 date_received: %s", r.get("date_received"))
            log.info("  최신 report_number: %s", r.get("report_number"))
            dev = (r.get("device") or [{}])[0]
            log.info("  brand_name: %s", dev.get("brand_name"))
    except Exception as e:
        log.error("  실패: %s", e)

    log.info("\n" + "=" * 60)
    log.info("진단 완료. 연도별 건수가 0이 아닌 구간에서 --start/--end 를 지정해 수집하세요.")
    log.info("=" * 60)
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="FDA MAUDE CGM 수집기 (체크포인트 기반 자동 증분 수집)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
자동 수집 로직:
  1) 최초 실행 (체크포인트 없음) → 최근 2년치 조회 + 체크포인트 저장
  2) 이후 실행 → 체크포인트 저장일 - 1일(겹침) 부터 오늘까지 조회 → 체크포인트 갱신
  3) CGM_BRANDS 목록이 바뀌면 자동으로 "최초 실행" 모드로 전환됨
     (새 브랜드 조합 = 새 체크포인트)

예시:
  python fda_maude_collector.py              # 위 자동 로직 (권장)
  python fda_maude_collector.py --test       # API 연결 진단
  python fda_maude_collector.py --initial    # 강제로 최초 모드 (2년 재수집)
  python fda_maude_collector.py --start 20240101 --end 20251130  # 기간 수동 지정
  python fda_maude_collector.py --export-only # DB만으로 Excel 재생성
  python fda_maude_collector.py --verbose    # 상세 URL 로그
""")
    parser.add_argument("--start", help="조회 시작일 YYYYMMDD (수동 지정, 체크포인트 갱신 안 함)")
    parser.add_argument("--end", help="조회 종료일 YYYYMMDD (수동 지정)")
    parser.add_argument("--initial", action="store_true",
                        help="강제 최초 모드: 체크포인트 무시하고 최근 2년 재수집")
    parser.add_argument("--initial-years", type=int, default=5,
                        help="최초 실행 시 과거 N년치 수집 (기본 2)")
    parser.add_argument("--overlap-days", type=int, default=1,
                        help="증분 시 이전 체크포인트와 겹칠 일수 (기본 1)")
    parser.add_argument("--api-key", default=None,
                        help="openFDA API 키 (환경변수 OPENFDA_API_KEY 또는 api_key.txt 파일도 가능)")
    parser.add_argument("--db", default=str(DB_PATH), help="SQLite 경로")
    parser.add_argument("--excel", default=str(EXCEL_PATH), help="Excel 출력 경로")
    parser.add_argument("--export-only", action="store_true",
                        help="DB 수집 없이 현재 DB로 Excel만 재생성")
    parser.add_argument("--test", action="store_true",
                        help="진단 모드: API 연결 및 쿼리 검증만 수행 (DB 저장 안 함)")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="요청 URL 등 상세 로그 출력")
    args = parser.parse_args()

    # API 키 로딩 (CLI > 환경변수 > FDA_MAUDE_API_KEY.txt > api_key.txt)
    api_key = load_api_key(args.api_key)
    if api_key:
        log.info("openFDA API 키 사용 (끝 4자리: ...%s)", api_key[-4:])
        log.info("  → 속도 제한 완화: 240요청/분, 120,000요청/시간, 페이지 간 %.2fs",
                 SLEEP_WITH_KEY)
    else:
        log.warning("openFDA API 키 없음 — 분당 40 / 일 1,000 요청 제한 (페이지 간 %.1fs)",
                    SLEEP_WITHOUT_KEY)
        log.warning("  무료 키 발급 (과금 없음): https://open.fda.gov/apis/authentication/")
        log.warning("  발급 후 이 폴더에 'FDA_MAUDE_API_KEY.txt' 또는 'api_key.txt' 로 저장")

    # --test 모드: 진단만 하고 종료
    if args.test:
        return run_test_mode(api_key)

    if args.export_only:
        return export_excel_only(Path(args.db), Path(args.excel))

    db_path = Path(args.db)
    excel_path = Path(args.excel)
    today = datetime.now()

    # DB 초기화 후 체크포인트 확인
    conn = init_db(db_path)
    checkpoint = get_checkpoint(conn, CGM_BRANDS)

    # 고장난 체크포인트 자동 무효화:
    # total_collected=0 이면 이전 실행에서 실질적으로 아무 것도 못 가져온 것.
    # (예: API 키 부재 → 403, 필터 에러, 네트워크 실패 등)
    # 이 체크포인트를 따라가면 "하루 2일치만 재시도" 같은 허무한 실행이 반복됨.
    # → 강제로 최초 수집 모드로 떨어뜨리고 기존 레코드는 삭제하여 재시도.
    if checkpoint and (checkpoint.get("total_collected") or 0) == 0:
        log.warning("=" * 60)
        log.warning("이전 체크포인트가 비정상 (수집 0건):")
        log.warning("  last_search_end = %s", checkpoint.get("last_search_end"))
        log.warning("  last_run_at     = %s", checkpoint.get("last_run_at"))
        log.warning("  note            = %s", checkpoint.get("note"))
        log.warning("→ 이 체크포인트를 무효화하고 최초 수집 모드로 재시도합니다.")
        log.warning("=" * 60)
        key = _brand_key(CGM_BRANDS)
        conn.execute("DELETE FROM collection_checkpoint WHERE brand_key = ?", (key,))
        conn.commit()
        checkpoint = None

    # --- 조회 기간 결정 ---
    manual_range = False
    is_initial_run = False
    mode_note = ""

    if args.start and args.end:
        # (A) 사용자 수동 지정: 체크포인트 건드리지 않음
        start = parse_date(args.start)
        end = parse_date(args.end)
        manual_range = True
        mode_note = f"수동 범위: {args.start} ~ {args.end}"

    elif args.initial or checkpoint is None:
        # (B) 강제 최초 모드 OR 체크포인트 없음 → 최근 N년치
        end = today
        start = today - timedelta(days=365 * args.initial_years)
        is_initial_run = True
        if args.initial:
            mode_note = f"강제 초기 수집 ({args.initial_years}년)"
        else:
            mode_note = f"최초 실행: {args.initial_years}년치 일괄 수집"

    else:
        # (C) 체크포인트 있음 → 증분 수집 (이전 종료일 - overlap 부터 오늘까지)
        last_end = datetime.strptime(checkpoint["last_search_end"], "%Y-%m-%d")
        start = last_end - timedelta(days=args.overlap_days)
        end = today
        mode_note = (
            f"증분 수집 (이전 종료일 {checkpoint['last_search_end']} "
            f"- {args.overlap_days}일 겹침 → 오늘)"
        )

    log.info("=" * 60)
    log.info("대상 브랜드 (key=%s):", _brand_key(CGM_BRANDS))
    for b in CGM_BRANDS:
        log.info("  - %s", b)
    log.info("모드: %s", mode_note)
    if checkpoint:
        log.info("이전 체크포인트: last_search_end=%s, total_collected=%d, last_run_at=%s",
                 checkpoint["last_search_end"],
                 checkpoint["total_collected"] or 0,
                 checkpoint["last_run_at"])
    else:
        log.info("이전 체크포인트: 없음 (최초 실행)")
    log.info("조회 기간: %s ~ %s", start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d"))
    log.info("부작용 필터: EVENT_TYPE in %s | adverse_event_flag=%s",
             EVENT_TYPES or "(제한 없음)",
             "Y 만" if ONLY_ADVERSE_EVENTS else "전체")
    log.info("Fallback 쿼리: %s", "활성" if USE_FALLBACK_FIELDS else "비활성")
    log.info("=" * 60)

    inserted = 0
    errors = 0
    api_key_issue = False
    try:
        try:
            for ev in fetch_events(start, end, api_key=api_key, verbose=args.verbose):
                try:
                    row = flatten_event(ev)
                    if not row["report_number"]:
                        continue
                    upsert_report(conn, row)
                    inserted += 1
                    if inserted % 100 == 0:
                        conn.commit()
                        log.info("진행: %d건 저장", inserted)
                except Exception as e:  # noqa: BLE001
                    errors += 1
                    log.exception("이벤트 처리 실패: %s", e)
        except APIKeyRequired as e:
            api_key_issue = True
            log.error("=" * 60)
            log.error("openFDA 가 API 키를 요구하거나 속도 제한에 걸렸습니다.")
            log.error("서버 메시지: %s", str(e)[:200])
            log.error("해결 방법:")
            log.error("  1) https://open.fda.gov/apis/authentication/ 에서 무료 키 발급")
            log.error("  2) 발급받은 키를 이 폴더 'api_key.txt' 파일에 저장")
            log.error("  3) run_collector.bat 다시 실행")
            log.error("지금까지 받은 %d건은 DB 에 저장합니다.", inserted)
            log.error("=" * 60)
        conn.commit()
    finally:
        log.info("DB 저장 완료 (신규/갱신: %d, 오류: %d)", inserted, errors)

        # 체크포인트 갱신 정책:
        #  (A) 수동 범위 모드 → 저장 안 함
        #  (B) API 키 문제 발생 → 저장 안 함 (다음 실행에서 같은 범위 재시도)
        #  (C) 최초 수집인데 0건 → 저장 안 함 (실질 실패로 간주, 다음에 재시도)
        #  (D) 증분인데 0건 → 저장함 (정상적으로 "새 보고서 없음" 일 수 있음)
        skip_reason: Optional[str] = None
        if manual_range:
            skip_reason = "수동 범위 모드"
        elif api_key_issue:
            skip_reason = "API 키 문제"
        elif is_initial_run and inserted == 0:
            skip_reason = "최초 수집에서 0건 (실질 실패로 간주)"

        if skip_reason:
            log.warning("체크포인트 갱신 안 함 (%s). 다음 실행 시 같은 범위 재시도.", skip_reason)
        else:
            checkpoint_end = end
            note_text = "최초 수집" if is_initial_run else "증분 수집"

            # 증분 수집에서 0건이면 "오늘"로 전진하지 않고,
            # DB 에 실제 존재하는 마지막 date_received 를 체크포인트로 유지.
            if (not is_initial_run) and inserted == 0:
                latest_db_date = get_latest_date_received(conn)
                if latest_db_date is not None:
                    checkpoint_end = latest_db_date
                    note_text = "증분 수집 (0건: DB 마지막 날짜 유지)"
                    log.info(
                        "증분 0건: 체크포인트를 조회 종료일(%s) 대신 DB 마지막 날짜(%s)로 유지",
                        end.strftime("%Y-%m-%d"),
                        checkpoint_end.strftime("%Y-%m-%d"),
                    )
                else:
                    log.warning(
                        "증분 0건 + DB 마지막 date_received 없음: 체크포인트를 조회 종료일(%s)로 저장",
                        end.strftime("%Y-%m-%d"),
                    )

            save_checkpoint(conn, CGM_BRANDS, checkpoint_end, inserted, note=note_text)
            log.info("체크포인트 저장: last_search_end=%s (다음 실행은 %s 부터)",
                     checkpoint_end.strftime("%Y-%m-%d"),
                     (checkpoint_end - timedelta(days=args.overlap_days)).strftime("%Y-%m-%d"))

        # Excel 내보내기
        try:
            export_excel(conn, excel_path)
        except Exception as e:  # noqa: BLE001
            log.exception("Excel 내보내기 실패: %s", e)
        conn.close()

    total = sqlite3.connect(db_path).execute("SELECT COUNT(*) FROM maude_reports").fetchone()[0]
    log.info("DB 총 보고서 수: %d", total)
    if total == 0:
        log.warning("\n[힌트] 0건이면 먼저 'python fda_maude_collector.py --test' 로 진단하세요.\n"
                    "       연도별 건수를 확인하고, 데이터가 있는 구간으로 --start/--end 를 지정합니다.")
    log.info("완료.")
    return 0


if __name__ == "__main__":
    sys.exit(main())