# -*- coding: utf-8 -*-
"""
FDA MAUDE CGM 부작용 대시보드 (Streamlit)
=========================================
`fda_maude_collector.py` 가 수집한 `fda_maude_cgm.db` 를 읽어
브라우저에서 필터링·검색·다운로드할 수 있는 대시보드를 제공합니다.

실행:
  streamlit run maude_dashboard.py
  (또는 run_dashboard.bat 더블클릭)

필터:
  - 브랜드 (다중 선택)
  - EVENT_TYPE (Death / Injury / Malfunction / Other)
  - 날짜 범위 (달력 위젯)
  - 키워드 검색 (event_description, manufacturer_narrative, 요약 공통)
  - 문제 코드 검색 (Health Effect / Medical Device Problem Code 번호 일치)
  - MDR 보고 구분 (초기 보고 / 후속 수정)

탭:
  1) 인사이트        — 경쟁사 위협 선제 대응 (A+C: 개발 리스크 회피 + 규제 대응)
                       · Death/Injury 드릴다운 · 에스컬레이션 위험도
                       · 제조사 사각지대(source_type 갭) · 규제 선행 지표
                       · 급증 신호 · 신규 코드 · 리포트 xlsx 발행 버튼
  2) 전체 보고서     — 필터된 표 + Excel 다운로드
  3) EVENT_TYPE 분포 — 파이/막대
  4) 문제 코드       — Health Effect - Clinical Code / Medical Device Problem Code Top-N + 브랜드 교차표
  5) 환자 인구통계   — 성별/인종/민족/연령
  6) 제조사 · 국가    — Top 10
  7) 월별 추이       — 이벤트 유형별 추이
"""

from __future__ import annotations

import io
import hashlib
import json
import re
import sqlite3
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import altair as alt
import pandas as pd
import streamlit as st

# ---------------------------------------------------------------------------
# 기본 설정
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "fda_maude_cgm.db"
BRAND_GROUPS_PATH = BASE_DIR / "brand_groups.json"
PROBLEM_CODE_MAP_PATH = BASE_DIR / "problem_code_map.json"
ANNEX_XLSX_PATH = BASE_DIR / "fda-annexes-a-g-2025.xlsx"
BRAND_GROUPS: Dict[str, List[str]] = {}
CANONICAL_TO_MEMBERS: Dict[str, List[str]] = {}
MEMBER_TO_CANONICAL: Dict[str, str] = {}

PATIENT_PROBLEM_LABEL = "Health Effect - Clinical Code"
DEVICE_PROBLEM_LABEL = "Medical Device Problem Code"
CODE_FILTER_TOKEN = "__MAUDE_CODE_FILTER__="
REPORT_STAGE_TOKEN = "__MAUDE_REPORT_STAGE__="
ID_FILTER_TOKEN = "__MAUDE_ID_FILTER__="

st.set_page_config(
    page_title="FDA MAUDE CGM 부작용 대시보드",
    page_icon="📊",
    layout="wide",
)


# ---------------------------------------------------------------------------
# DB 접근 유틸 (캐시 적용)
# ---------------------------------------------------------------------------

@st.cache_data(show_spinner=False)
def _db_last_modified() -> float:
    if DB_PATH.exists():
        return DB_PATH.stat().st_mtime
    return 0.0


def _connect() -> sqlite3.Connection:
    """읽기 전용 connection. 캐시 무효화를 위해 파일 mtime 을 키로 사용."""
    _ = _db_last_modified()  # mtime 체크만
    return sqlite3.connect(f"file:{DB_PATH}?mode=ro", uri=True)


def _parse_code_numbers(raw: str) -> List[str]:
    """코드 입력 문자열에서 숫자 코드 목록(2~6자리)을 추출."""
    text = str(raw or "").strip()
    if not text:
        return []
    codes: List[str] = []
    seen: Set[str] = set()
    for m in re.finditer(r"\d{2,6}", text):
        code = m.group(0)
        if code not in seen:
            seen.add(code)
            codes.append(code)
    return codes


def _normalize_report_stage(raw: str) -> str:
    s = str(raw or "").strip().lower()
    if s in ("initial", "followup"):
        return s
    return "all"


def _parse_id_search_terms(raw: str) -> List[str]:
    """ID 검색 입력에서 토큰 목록(쉼표/공백/줄바꿈 구분, 중복 제거) 추출."""
    text = str(raw or "").strip()
    if not text:
        return []
    seen: Set[str] = set()
    out: List[str] = []
    for term in re.split(r"[,\s]+", text):
        token = term.strip()
        if not token or token in seen:
            continue
        seen.add(token)
        out.append(token)
    return out


def _is_number_like_id(term: str) -> bool:
    """숫자열 또는 하이픈 숫자 조합(예: 2954323-2024-14003)인지 판별."""
    t = str(term or "").strip()
    if not t:
        return False
    return bool(re.fullmatch(r"\d+(?:-\d+)*", t))


def _compose_search_keyword(
    keyword: str,
    code_search: str,
    report_stage: str = "all",
    id_search: str = "",
) -> str:
    """기존 keyword 파이프라인을 유지하면서 코드 필터를 함께 전달."""
    kw = str(keyword or "").strip()
    code_text = str(code_search or "").strip()
    id_text = str(id_search or "").strip()
    stage = _normalize_report_stage(report_stage)
    out = kw
    if code_text:
        out = (f"{out} " if out else "") + f"{CODE_FILTER_TOKEN}{code_text}"
    if id_text:
        out = (f"{out} " if out else "") + f"{ID_FILTER_TOKEN}{id_text}"
    if stage != "all":
        out = (f"{out} " if out else "") + f"{REPORT_STAGE_TOKEN}{stage}"
    return out.strip()


def _split_search_keyword(raw_keyword: str) -> Tuple[str, List[str], str, List[str]]:
    """합성된 keyword 문자열에서 (키워드, 코드목록, 보고구분, ID목록) 분리."""
    text = str(raw_keyword or "").strip()
    if not text:
        return "", [], "all", []

    keyword = text
    stage = "all"
    codes: List[str] = []
    ids: List[str] = []

    stage_idx = keyword.find(REPORT_STAGE_TOKEN)
    if stage_idx >= 0:
        stage_text = keyword[stage_idx + len(REPORT_STAGE_TOKEN):].strip().split()[0] if keyword[stage_idx + len(REPORT_STAGE_TOKEN):].strip() else ""
        stage = _normalize_report_stage(stage_text)
        keyword = (keyword[:stage_idx] + " " + keyword[stage_idx + len(REPORT_STAGE_TOKEN) + len(stage_text):]).strip()

    id_idx = keyword.find(ID_FILTER_TOKEN)
    if id_idx >= 0:
        id_text = keyword[id_idx + len(ID_FILTER_TOKEN):].strip()
        keyword = keyword[:id_idx].strip()
        if CODE_FILTER_TOKEN in id_text:
            id_text = id_text.split(CODE_FILTER_TOKEN, 1)[0].strip()
        if REPORT_STAGE_TOKEN in id_text:
            id_text = id_text.split(REPORT_STAGE_TOKEN, 1)[0].strip()
        ids = _parse_id_search_terms(id_text)

    code_idx = keyword.find(CODE_FILTER_TOKEN)
    if code_idx >= 0:
        code_text = keyword[code_idx + len(CODE_FILTER_TOKEN):].strip()
        keyword = keyword[:code_idx].strip()
        # code_text 안에 다른 토큰이 섞인 경우 방어
        if ID_FILTER_TOKEN in code_text:
            code_text = code_text.split(ID_FILTER_TOKEN, 1)[0].strip()
        if REPORT_STAGE_TOKEN in code_text:
            code_text = code_text.split(REPORT_STAGE_TOKEN, 1)[0].strip()
        codes = _parse_code_numbers(code_text)

    return keyword, codes, stage, ids


@st.cache_data(show_spinner=False)
def _code_to_problem_terms(db_mtime: float, map_mtime: float) -> Dict[str, List[str]]:
    """코드번호 -> 문제용어(정규화) 역매핑."""
    _ = db_mtime
    # 속도: 검색용 역매핑은 수동 JSON 맵만 사용한다.
    # (DB 전수 스캔 기반 추론은 비용이 커서 필터 응답성을 떨어뜨림)
    merged = load_problem_code_map(map_mtime)
    buckets: Dict[str, Set[str]] = defaultdict(set)
    for scope in ("common", "patient", "device"):
        section = merged.get(scope, {})
        if not isinstance(section, dict):
            continue
        for norm_term, code in section.items():
            c = str(code or "").strip()
            t = re.sub(r"\s+", " ", str(norm_term or "").strip()).lower()
            if c and t:
                buckets[c].add(t)
    return {k: sorted(v) for k, v in buckets.items()}


@st.cache_data(show_spinner=False)
def detect_columns(mtime: float) -> List[str]:
    """maude_reports 테이블에 실제로 존재하는 칼럼명을 반환.
    구 버전 DB 에 신규 칼럼(manufacturer_country, patient_* 등)이 없을 수 있어
    쿼리 작성 시 이 목록을 참고해 NULL 로 대체한다.
    """
    if not DB_PATH.exists():
        return []
    with _connect() as conn:
        return [r[1] for r in conn.execute("PRAGMA table_info(maude_reports)").fetchall()]


def _safe_col(available: List[str], col: str) -> str:
    """해당 칼럼이 있으면 col, 없으면 NULL 리터럴을 반환 (alias 는 호출부에서 붙임)."""
    return col if col in available else "NULL"


def load_brand_groups() -> Dict[str, List[str]]:
    """저장된 브랜드 그룹(대표명 -> 멤버 목록) 로드."""
    if not BRAND_GROUPS_PATH.exists():
        return {}
    try:
        payload = json.loads(BRAND_GROUPS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}
    if not isinstance(payload, dict):
        return {}
    groups = payload.get("groups", {})
    if not isinstance(groups, dict):
        return {}
    cleaned: Dict[str, List[str]] = {}
    for canonical, members in groups.items():
        if not isinstance(canonical, str) or not canonical.strip():
            continue
        if not isinstance(members, list):
            continue
        unique_members: List[str] = []
        seen: Set[str] = set()
        for m in members:
            if isinstance(m, str) and m.strip() and m not in seen:
                unique_members.append(m)
                seen.add(m)
        if unique_members:
            cleaned[canonical.strip()] = unique_members
    return cleaned


def save_brand_groups(groups: Dict[str, List[str]]) -> None:
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "groups": groups,
    }
    BRAND_GROUPS_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def build_brand_alias_maps(groups: Dict[str, List[str]]) -> Tuple[Dict[str, List[str]], Dict[str, str]]:
    """(대표명->멤버들), (멤버 대문자->대표명) 반환."""
    canonical_to_members: Dict[str, List[str]] = {}
    member_to_canonical: Dict[str, str] = {}
    for canonical, members in groups.items():
        norm_canonical = canonical.strip()
        if not norm_canonical:
            continue
        all_members: List[str] = []
        seen_members: Set[str] = set()
        for raw_member in list(members) + [norm_canonical]:
            member = str(raw_member).strip()
            if not member or member in seen_members:
                continue
            all_members.append(member)
            seen_members.add(member)
            member_to_canonical[member.upper()] = norm_canonical
        canonical_to_members[norm_canonical] = all_members
    return canonical_to_members, member_to_canonical


def normalize_brand_value(value: object, member_to_canonical: Dict[str, str]) -> object:
    if value is None:
        return value
    s = str(value).strip()
    if not s:
        return value
    return member_to_canonical.get(s.upper(), value)


def apply_brand_aliases(df: pd.DataFrame, member_to_canonical: Dict[str, str]) -> pd.DataFrame:
    """브랜드 칼럼이 있으면 대표명으로 치환해 집계 일관성 확보."""
    if df is None or df.empty or not member_to_canonical:
        return df
    alias_cols = ["brand_name", "BRAND_NAME", "brand", "BRAND"]
    out = df.copy()
    for col in alias_cols:
        if col in out.columns:
            out[col] = out[col].apply(lambda v: normalize_brand_value(v, member_to_canonical))
    return out


@st.cache_data(show_spinner=False)
def load_metadata(mtime: float) -> Tuple[List[str], List[str], Optional[date], Optional[date], int]:
    """필터 위젯 기본값 계산: 브랜드 / event_type / 날짜 범위 / 총건수."""
    if not DB_PATH.exists():
        return [], [], None, None, 0
    with _connect() as conn:
        brands = [r[0] for r in conn.execute(
            "SELECT DISTINCT brand_name FROM maude_reports "
            "WHERE brand_name IS NOT NULL AND brand_name <> '' "
            "ORDER BY brand_name"
        )]
        event_types = [r[0] for r in conn.execute(
            "SELECT DISTINCT event_type FROM maude_reports "
            "WHERE event_type IS NOT NULL AND event_type <> '' "
            "ORDER BY event_type"
        )]
        row = conn.execute(
            "SELECT MIN(date_received), MAX(date_received) FROM maude_reports "
            "WHERE date_received IS NOT NULL AND date_received <> ''"
        ).fetchone()
        min_d = _parse_iso_date(row[0]) if row and row[0] else None
        max_d = _parse_iso_date(row[1]) if row and row[1] else None
        total = conn.execute("SELECT COUNT(*) FROM maude_reports").fetchone()[0]
    return brands, event_types, min_d, max_d, total


def _parse_iso_date(s: str) -> Optional[date]:
    if not s:
        return None
    try:
        s = str(s).strip().replace("-", "")[:8]
        if len(s) == 8 and s.isdigit():
            return datetime.strptime(s, "%Y%m%d").date()
    except Exception:
        pass
    return None


def _build_where(
    brands: List[str],
    event_types: List[str],
    date_from: Optional[date],
    date_to: Optional[date],
    keyword: str,
) -> Tuple[str, List]:
    """필터로부터 WHERE 절과 파라미터 리스트를 만든다.

    브랜드 매칭은 기본적으로 '부분 일치(LIKE)' 로 수행합니다. 예를 들어 사용자가
    'DEXCOM' 을 선택하면 'DEXCOM G7 CONTINUOUS GLUCOSE MONITORING SYSTEM',
    'DEXCOM G6 CGM' 등이 모두 매칭됩니다. (FDA MAUDE 에는 같은 제품이
    짧은 이름과 긴 이름으로 혼재되어 저장되기 때문.)
    단, 사용자가 그룹 대표명(예: Libre 2)을 선택한 경우에는 해당 그룹 멤버
    명칭에 대해 정확 매칭을 수행하여, 'Libre 2+' 같은 다른 그룹으로의
    과도한 포함을 방지합니다.
    """
    clauses: List[str] = []
    params: List = []
    keyword_text, code_filters, report_stage, id_filters = _split_search_keyword(keyword)

    if brands:
        # 대표명 그룹이 있으면 실제 멤버 브랜드명으로 확장해서 검색.
        # 그룹 선택이나 개별 브랜드 선택 모두 정확 매칭으로 처리한다.
        expanded_brands: List[str] = []
        seen: Set[str] = set()
        for chosen in brands:
            members = CANONICAL_TO_MEMBERS.get(chosen)
            if members is not None:
                for member in members:
                    if member not in seen:
                        expanded_brands.append(member)
                        seen.add(member)
            else:
                if chosen not in seen:
                    expanded_brands.append(chosen)
                    seen.add(chosen)
        if len(expanded_brands) == 1:
            clauses.append("brand_name COLLATE NOCASE = ?")
            params.append(expanded_brands[0])
        else:
            placeholders = ",".join(["?"] * len(expanded_brands))
            clauses.append(f"brand_name COLLATE NOCASE IN ({placeholders})")
            params.extend(expanded_brands)

    if event_types:
        placeholders = ",".join(["?"] * len(event_types))
        clauses.append(f"event_type IN ({placeholders})")
        params.extend(event_types)

    if date_from:
        clauses.append("REPLACE(SUBSTR(date_received, 1, 10), '-', '') >= ?")
        params.append(date_from.strftime("%Y%m%d"))
    if date_to:
        clauses.append("REPLACE(SUBSTR(date_received, 1, 10), '-', '') <= ?")
        params.append(date_to.strftime("%Y%m%d"))

    if keyword_text:
        like = f"%{keyword_text}%"
        mdr_key_like = f'%\"mdr_report_key\"%{keyword_text}%'
        # summary 칼럼은 구 DB 에도 존재하므로 항상 포함 가능
        clauses.append(
            "("
            "event_description LIKE ? OR "
            "manufacturer_narrative LIKE ? OR "
            "summary_complaint_kr LIKE ? OR "
            "summary_response_kr LIKE ? OR "
            "patient_problems LIKE ? OR "
            "product_problems LIKE ? OR "
            "COALESCE(report_number, '') LIKE ? OR "
            "COALESCE(raw_json, '') LIKE ?"
            ")"
        )
        params.extend([like] * 6 + [like, mdr_key_like])

    if id_filters:
        per_id_clauses: List[str] = []
        for term in id_filters:
            if _is_number_like_id(term):
                per_id_clauses.append(
                    "("
                    "COALESCE(report_number, '') = ? COLLATE NOCASE OR "
                    "COALESCE(raw_json, '') LIKE ? OR "
                    "COALESCE(raw_json, '') LIKE ? OR "
                    "COALESCE(raw_json, '') LIKE ?"
                    ")"
                )
                params.extend([
                    term,
                    f'%\"mdr_report_key\":\"{term}\"%',
                    f'%\"mdr_report_key\": \"{term}\"%',
                    f'%\"mdr_report_key\": {term}%',
                ])
            else:
                like = f"%{term}%"
                mdr_key_like = f'%\"mdr_report_key\"%{term}%'
                per_id_clauses.append(
                    "("
                    "COALESCE(report_number, '') LIKE ? COLLATE NOCASE OR "
                    "COALESCE(raw_json, '') LIKE ? COLLATE NOCASE"
                    ")"
                )
                params.extend([like, mdr_key_like])
        clauses.append("(" + " OR ".join(per_id_clauses) + ")")

    if code_filters:
        code_term_map = _code_to_problem_terms(_db_last_modified(), _problem_code_map_mtime())
        per_code_clauses: List[str] = []
        for code in code_filters:
            one_code_clauses: List[str] = []
            like_patterns: List[str] = [
                f"%({code})%",
                f"%{code}:%",
                f"%{code} -%",
                f"%{code}|%",
            ]

            # 화면 표시용 코드(수동 매핑)만 있고 원문에 숫자가 없는 경우도 포함
            terms = code_term_map.get(str(code), [])
            for term in terms[:40]:
                like_patterns.append(f"%{term}%")

            for pat in like_patterns:
                one_code_clauses.append(
                    "(COALESCE(patient_problems, '') LIKE ? COLLATE NOCASE "
                    "OR COALESCE(product_problems, '') LIKE ? COLLATE NOCASE)"
                )
                params.extend([pat, pat])

            per_code_clauses.append("(" + " OR ".join(one_code_clauses) + ")")

        clauses.append("(" + " OR ".join(per_code_clauses) + ")")

    if report_stage in ("initial", "followup"):
        pat = "%Initial submission%" if report_stage == "initial" else "%Followup%"
        available = set(detect_columns(_db_last_modified()))
        if "type_of_report" in available:
            clauses.append(
                "("
                "COALESCE(type_of_report, '') LIKE ? COLLATE NOCASE "
                "OR (TRIM(COALESCE(type_of_report, '')) = '' AND COALESCE(raw_json, '') LIKE ? COLLATE NOCASE)"
                ")"
            )
            params.extend([pat, pat])
        else:
            # 구 스키마 호환
            clauses.append("COALESCE(raw_json, '') LIKE ? COLLATE NOCASE")
            params.append(pat)

    where = " AND ".join(clauses) if clauses else "1=1"
    return where, params


# 메인 테이블 표시용 칼럼 정의: (SQL 식별자, 엑셀/표 라벨)
#   - 실제 DB 에 칼럼이 없으면 NULL 로 자동 대체됨.
_REPORT_COLS: List[Tuple[str, str]] = [
    ("report_number",          "MAUDE 번호"),
    ("event_type",             "EVENT_TYPE"),
    ("type_of_report",         "REPORT_STAGE"),
    ("date_received",          "DATE_RECEIVED"),
    ("date_report",            "DATE_REPORT"),
    ("date_of_event",          "DATE_OF_EVENT"),
    ("brand_name",             "BRAND_NAME"),
    ("manufacturer_name",      "MANUFACTURER"),
    ("manufacturer_country",   "COUNTRY"),
    ("patient_sex",            "SEX"),
    ("patient_age",            "AGE"),
    ("patient_race",           "RACE"),
    ("patient_ethnicity",      "ETHNICITY"),
    ("patient_problems",       "HEALTH_EFFECT_CLINICAL_CODE"),
    ("product_problems",       "MEDICAL_DEVICE_PROBLEM_CODE"),
    ("summary_complaint_kr",   "[요약] 소비자 불만"),
    ("summary_response_kr",    "[요약] 제조사 대응"),
    ("summary_conclusion_kr",  "[요약] 결론"),
    ("event_description",      "EVENT_DESCRIPTION (원문)"),
    ("manufacturer_narrative", "MANUFACTURER_NARRATIVE (원문)"),
    ("raw_json",               "__RAW_JSON"),
]


@st.cache_data(show_spinner=False)
def query_reports(
    mtime: float,
    brands: List[str],
    event_types: List[str],
    date_from: Optional[date],
    date_to: Optional[date],
    keyword: str,
    limit: int = 5000,
) -> pd.DataFrame:
    """필터 적용된 리포트 DataFrame 반환 (최대 limit 건).
    DB 에 없는 칼럼은 NULL 로 대체하여, 구 버전 DB 에서도 동작한다.
    """
    available = detect_columns(mtime)
    where, params = _build_where(brands, event_types, date_from, date_to, keyword)
    select_parts = []
    for col, label in _REPORT_COLS:
        expr = col if col in available else "NULL"
        # label 은 한글/공백 포함이므로 작은따옴표 대신 큰따옴표로 묶음
        select_parts.append(f'{expr} AS "{label}"')
    select_clause = ",\n            ".join(select_parts)
    sql = f"""
        SELECT
            {select_clause}
        FROM maude_reports
        WHERE {where}
        ORDER BY date_received DESC, report_number DESC
        LIMIT {int(limit)}
    """
    with _connect() as conn:
        df = pd.read_sql_query(sql, conn, params=params)
    if "__RAW_JSON" in df.columns:
        def _extract_additional(raw: Optional[str]) -> str:
            if not isinstance(raw, str) or "Additional Manufacturer Narrative" not in raw:
                return ""
            return _extract_narrative_sections(raw).get("additional_manufacturer_narrative", "")
        df["ADDITIONAL_MANUFACTURER_NARRATIVE (원문)"] = df["__RAW_JSON"].apply(_extract_additional)
        df = df.drop(columns=["__RAW_JSON"])
    elif "ADDITIONAL_MANUFACTURER_NARRATIVE (원문)" not in df.columns:
        df["ADDITIONAL_MANUFACTURER_NARRATIVE (원문)"] = ""
    return format_problem_columns(apply_brand_aliases(df, MEMBER_TO_CANONICAL))


@st.cache_data(show_spinner=False)
def query_filtered_count(
    mtime: float,
    brands: List[str],
    event_types: List[str],
    date_from: Optional[date],
    date_to: Optional[date],
    keyword: str,
) -> int:
    where, params = _build_where(brands, event_types, date_from, date_to, keyword)
    with _connect() as conn:
        return conn.execute(
            f"SELECT COUNT(*) FROM maude_reports WHERE {where}", params
        ).fetchone()[0]


def _extract_narrative_sections(raw_json: Optional[str]) -> Dict[str, str]:
    """raw_json.mdr_text 를 파싱해 주요 내러티브 섹션을 분리한다."""
    out = {
        "event_description": "",
        "manufacturer_narrative": "",
        "additional_manufacturer_narrative": "",
    }
    if not raw_json:
        return out
    try:
        payload = json.loads(raw_json)
    except Exception:
        return out

    event_parts: List[str] = []
    mfr_parts: List[str] = []
    add_mfr_parts: List[str] = []
    for item in payload.get("mdr_text", []) or []:
        if not isinstance(item, dict):
            continue
        txt = str(item.get("text") or "").strip()
        if not txt:
            continue
        tcode = str(item.get("text_type_code") or "").strip().lower()
        if "additional manufacturer narrative" in tcode:
            add_mfr_parts.append(txt)
        elif "manufacturer" in tcode:
            mfr_parts.append(txt)
        else:
            event_parts.append(txt)

    out["event_description"] = "\n\n".join(event_parts).strip()
    out["manufacturer_narrative"] = "\n\n".join(mfr_parts).strip()
    out["additional_manufacturer_narrative"] = "\n\n".join(add_mfr_parts).strip()
    return out


@st.cache_data(show_spinner=False)
def query_report_narrative_sections(mtime: float, report_number: str) -> Dict[str, str]:
    """개별 MAUDE 번호의 raw_json 에서 내러티브 섹션을 추출."""
    _ = mtime
    if not report_number:
        return _extract_narrative_sections(None)
    with _connect() as conn:
        row = conn.execute(
            "SELECT raw_json FROM maude_reports WHERE report_number = ? LIMIT 1",
            (str(report_number),),
        ).fetchone()
    raw_json = row[0] if row else None
    return _extract_narrative_sections(raw_json)


@st.cache_data(show_spinner=False)
def query_group(
    mtime: float,
    brands: List[str],
    event_types: List[str],
    date_from: Optional[date],
    date_to: Optional[date],
    keyword: str,
    group_sql: str,
    header: List[str],
) -> pd.DataFrame:
    """필터 적용된 GROUP BY 집계. group_sql 은 SELECT ... FROM maude_reports WHERE {where} 형태."""
    where, params = _build_where(brands, event_types, date_from, date_to, keyword)
    sql = group_sql.format(where=where)
    with _connect() as conn:
        df = pd.read_sql_query(sql, conn, params=params)
    if len(df.columns) == len(header):
        df.columns = header
    return format_problem_columns(apply_brand_aliases(df, MEMBER_TO_CANONICAL))


@st.cache_data(show_spinner=False)
def query_code_fields(
    mtime: float,
    brands: List[str],
    event_types: List[str],
    date_from: Optional[date],
    date_to: Optional[date],
    keyword: str,
) -> pd.DataFrame:
    """필터 적용된 brand_name / event_type / patient_problems / product_problems
    네 칼럼만 로드. 탭 6(문제 코드) 전용 — 쉼표 구분 코드를 Python 쪽에서
    split/explode 로 펼쳐 Top-N 집계한다. 대량(수십만 건) 에서도 가볍다."""
    where, params = _build_where(brands, event_types, date_from, date_to, keyword)
    sql = f"""
        SELECT brand_name, event_type, patient_problems, product_problems
        FROM maude_reports
        WHERE {where}
    """
    with _connect() as conn:
        df = pd.read_sql_query(sql, conn, params=params)
    return format_problem_columns(apply_brand_aliases(df, MEMBER_TO_CANONICAL))


def _explode_codes(series: pd.Series, domain: Optional[str] = None) -> pd.Series:
    """세미콜론(;) 구분 문자열을 개별 코드 단위로 분리·정리.

    예) "Hypoglycemia; Dizziness; Nausea" → ["Hypoglycemia","Dizziness","Nausea"]
    주의: 코드명 내부의 쉼표(,)는 코드 텍스트로 유지한다.
    공백·빈문자열·None 은 모두 제외.
    """
    exploded = (
        series.dropna()
        .astype(str)
        .apply(_split_problem_terms)
        .explode()
        .str.strip()
        .replace("", pd.NA)
        .dropna()
    )
    if exploded.empty:
        return exploded

    # 같은 집합 내에 "Term (1234)"가 일부 존재하면, 동일 term의 코드 누락 항목도 보정한다.
    local_code_votes: Dict[str, Counter[str]] = defaultdict(Counter)
    parsed: List[Tuple[str, str]] = []
    for raw in exploded.tolist():
        term, code = _extract_problem_term_code(raw)
        parsed.append((term, code))
        if term and code:
            local_code_votes[_normalize_problem_term(term)][code] += 1

    local_best: Dict[str, str] = {}
    for norm_term, votes in local_code_votes.items():
        if not votes:
            continue
        local_best[norm_term] = votes.most_common(1)[0][0]

    out: List[str] = []
    for term, code in parsed:
        if not term:
            out.append("")
            continue
        if code:
            out.append(f"{term} ({code})")
            continue
        local_code = local_best.get(_normalize_problem_term(term), "")
        if local_code:
            out.append(f"{term} ({local_code})")
            continue
        out.append(_format_problem_term_with_code(term, domain=domain))
    return pd.Series(out, index=exploded.index)


def _problem_code_map_mtime() -> float:
    if PROBLEM_CODE_MAP_PATH.exists():
        return PROBLEM_CODE_MAP_PATH.stat().st_mtime
    return 0.0


@st.cache_data(show_spinner=False)
def load_problem_code_map(mtime: float) -> Dict[str, Dict[str, str]]:
    """문제 코드 매핑 로드.

    파일 형식(`problem_code_map.json`) 지원:
      1) 구형(flat): {"Term": "1234", ...}
      2) 권장형:
         {
           "common": {"Term": "1234"},
           "patient": {"Term": "2418"},
           "device": {"Term": "1535"}
         }
    """
    if not PROBLEM_CODE_MAP_PATH.exists():
        return {"common": {}, "patient": {}, "device": {}}
    try:
        payload = json.loads(PROBLEM_CODE_MAP_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {"common": {}, "patient": {}, "device": {}}
    if not isinstance(payload, dict):
        return {"common": {}, "patient": {}, "device": {}}

    def _norm_section(section: object) -> Dict[str, str]:
        out: Dict[str, str] = {}
        if not isinstance(section, dict):
            return out
        for k, v in section.items():
            if not isinstance(k, str):
                continue
            key = _normalize_problem_term(k)
            if not key:
                continue
            code = str(v).strip() if v is not None else ""
            if code:
                out[key] = code
        return out

    # 구형(flat) 파일 호환
    if not any(k in payload for k in ("common", "patient", "device")):
        return {
            "common": _norm_section(payload),
            "patient": {},
            "device": {},
        }

    return {
        "common": _norm_section(payload.get("common", {})),
        "patient": _norm_section(payload.get("patient", {})),
        "device": _norm_section(payload.get("device", {})),
    }


def _normalize_problem_term(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip()).upper()


def _extract_problem_term_code(raw: object) -> Tuple[str, str]:
    """문자열에서 term/code를 추출해 표준화.

    지원 입력:
      - "Term (1234)"
      - "1234: Term", "1234 - Term", "1234|Term"
      - 그 외는 (원문정리, 코드없음)
    """
    s = re.sub(r"\s+", " ", str(raw or "").strip())
    if not s:
        return "", ""

    m = re.match(r"^(?P<term>.+?)\s*\((?P<code>\d{2,6})\)\s*$", s)
    if m:
        return re.sub(r"\s+", " ", m.group("term")).strip(), m.group("code")

    m = re.match(r"^(?P<code>\d{2,6})\s*[:\-\|]\s*(?P<term>.+)$", s)
    if m:
        return re.sub(r"\s+", " ", m.group("term")).strip(), m.group("code")

    return s, ""


@st.cache_data(show_spinner=False)
def infer_problem_code_map_from_db(mtime: float) -> Dict[str, Dict[str, str]]:
    """DB 내 문제 코드 문자열에서 term->code를 자동 추론."""
    _ = mtime
    inferred: Dict[str, Dict[str, Counter[str]]] = {
        "patient": defaultdict(Counter),
        "device": defaultdict(Counter),
    }
    domain_cols = [("patient", "patient_problems"), ("device", "product_problems")]
    with _connect() as conn:
        for domain, col in domain_cols:
            sql = f"""
                SELECT DISTINCT {col}
                FROM maude_reports
                WHERE {col} IS NOT NULL
                  AND TRIM({col}) <> ''
                  AND (
                        INSTR({col}, '(') > 0
                     OR {col} GLOB '[0-9][0-9]*:*'
                     OR {col} GLOB '[0-9][0-9]* -*'
                     OR {col} GLOB '[0-9][0-9]*|*'
                  )
            """
            for (raw_text,) in conn.execute(sql):
                for token in _split_problem_terms(raw_text):
                    term, code = _extract_problem_term_code(token)
                    if not term or not code:
                        continue
                    inferred[domain][_normalize_problem_term(term)][code] += 1

    def _freeze(votes: Dict[str, Counter[str]]) -> Dict[str, str]:
        out: Dict[str, str] = {}
        for norm_term, counter in votes.items():
            if not counter:
                continue
            out[norm_term] = counter.most_common(1)[0][0]
        return out

    patient_map = _freeze(inferred["patient"])
    device_map = _freeze(inferred["device"])
    common_map: Dict[str, str] = {}
    for norm_term, p_code in patient_map.items():
        d_code = device_map.get(norm_term, "")
        if d_code and d_code == p_code:
            common_map[norm_term] = p_code

    return {"common": common_map, "patient": patient_map, "device": device_map}


def _annex_xlsx_mtime() -> float:
    if ANNEX_XLSX_PATH.exists():
        return ANNEX_XLSX_PATH.stat().st_mtime
    return 0.0


_ANNEX_DOMAIN_BY_LETTER = {
    "A": "device",   # Medical Device Problem
    "E": "patient",  # Health Effects - Clinical Signs and Symptoms or Conditions
}


@st.cache_data(show_spinner=False)
def load_annex_code_map(mtime: float) -> Dict[str, object]:
    """`fda-annexes-a-g-2025.xlsx` 의 Combined 시트를 읽어 코드 매핑을 만든다.

    반환:
      {
        "by_domain": {"patient": {NORM_TERM: code}, "device": {...}, "common": {}},
        "by_annex":  {"A": {NORM_TERM: code}, "B": {...}, ...},
        "all_terms_with_comma": [원문 Term, ...]  # 쉼표 split 보호용 (길이 내림차순)
      }
    """
    empty = {
        "by_domain": {"common": {}, "patient": {}, "device": {}},
        "by_annex": {},
        "all_terms_with_comma": [],
    }
    _ = mtime
    if not ANNEX_XLSX_PATH.exists():
        return empty
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(ANNEX_XLSX_PATH, read_only=True, data_only=True)
    except Exception:
        return empty
    if "Combined" not in wb.sheetnames:
        return empty
    ws = wb["Combined"]

    by_annex: Dict[str, Dict[str, str]] = {}
    by_domain: Dict[str, Dict[str, str]] = {"common": {}, "patient": {}, "device": {}}
    terms_with_comma: List[str] = []

    header = None
    idx_annex = idx_fda = idx_term = -1
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        if header is None:
            if row and row[0] == "Annex":
                header = [str(c).strip() if c is not None else "" for c in row]
                try:
                    idx_annex = header.index("Annex")
                    idx_fda = header.index("FDA Code")
                    idx_term = header.index("Term")
                except ValueError:
                    return empty
            continue

        annex = (str(row[idx_annex]).strip() if idx_annex < len(row) and row[idx_annex] is not None else "")
        term = (str(row[idx_term]).strip() if idx_term < len(row) and row[idx_term] is not None else "")
        fda = row[idx_fda] if idx_fda < len(row) else None
        if not annex or not term:
            continue
        code = str(fda).strip() if fda not in (None, "") else ""
        norm = _normalize_problem_term(term)
        if not norm:
            continue

        by_annex.setdefault(annex, {})[norm] = code
        domain = _ANNEX_DOMAIN_BY_LETTER.get(annex)
        if domain and code:
            by_domain[domain][norm] = code
        if "," in term:
            terms_with_comma.append(term)

    # 길이 내림차순(긴 표현 먼저 매치)
    terms_with_comma.sort(key=lambda s: -len(s))
    return {
        "by_domain": by_domain,
        "by_annex": by_annex,
        "all_terms_with_comma": terms_with_comma,
    }


@st.cache_data(show_spinner=False)
def merged_problem_code_map(db_mtime: float, map_mtime: float) -> Dict[str, Dict[str, str]]:
    """DB 자동 추론 + 수동 JSON 매핑 + Annex 사전 병합. 수동 > Annex > 자동 추론."""
    inferred = infer_problem_code_map_from_db(db_mtime)
    manual = load_problem_code_map(map_mtime)
    annex = load_annex_code_map(_annex_xlsx_mtime())
    annex_domain = annex.get("by_domain", {}) if isinstance(annex, dict) else {}

    merged: Dict[str, Dict[str, str]] = {"common": {}, "patient": {}, "device": {}}
    for scope in ("common", "patient", "device"):
        merged[scope].update(inferred.get(scope, {}))
        merged[scope].update(annex_domain.get(scope, {}))
        merged[scope].update(manual.get(scope, {}))
    return merged


def _format_problem_term_with_code(raw: object, domain: Optional[str] = None) -> str:
    """가능하면 `용어 (코드)` 형식으로 표준화."""
    term, explicit_code = _extract_problem_term_code(raw)
    if not term:
        return term
    if explicit_code:
        return f"{term} ({explicit_code})"

    # 로컬 매핑 파일이 있으면 코드 보강
    code_maps = merged_problem_code_map(_db_last_modified(), _problem_code_map_mtime())
    norm = _normalize_problem_term(term)
    code = ""
    if domain in ("patient", "device"):
        code = code_maps.get(domain, {}).get(norm, "")
    if not code:
        code = code_maps.get("common", {}).get(norm, "")
    if code:
        return f"{term} ({code})"

    return term


def _format_problem_field_text(raw: object, domain: Optional[str] = None) -> object:
    """문자열 필드(`a; b; c`)를 `용어 (코드)` 포맷으로 정규화."""
    if raw is None:
        return raw
    s = str(raw).strip()
    if not s:
        return raw
    parts = [p.strip() for p in _split_problem_terms(s) if p and p.strip()]
    if not parts:
        return raw
    return "; ".join(_format_problem_term_with_code(p, domain=domain) for p in parts)


def format_problem_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    for col in (
        "patient_problems", "PATIENT_PROBLEMS", "HEALTH_EFFECT_CLINICAL_CODE",
    ):
        if col in out.columns:
            out[col] = out[col].apply(lambda v: _format_problem_field_text(v, domain="patient"))
    for col in (
        "product_problems", "PRODUCT_PROBLEMS", "MEDICAL_DEVICE_PROBLEM_CODE",
    ):
        if col in out.columns:
            out[col] = out[col].apply(lambda v: _format_problem_field_text(v, domain="device"))
    return out


def _split_problem_terms(s: str) -> List[str]:
    """문제 코드 문자열 분리.

    기본 구분자는 세미콜론(;). 과거 수집본(쉼표 구분)도 호환 처리한다.
    """
    text = str(s or "").strip()
    if not text:
        return []
    if ";" in text:
        return [p.strip() for p in re.split(r"\s*;\s*", text) if p and p.strip()]

    # "Term (1234), Term2 (5678)" 형태는 ")" 경계를 기준으로 우선 분리한다.
    if re.search(r"\(\d{2,6}\)", text):
        chunks = re.split(r"\)\s*,\s*", text)
        terms: List[str] = []
        for i, chunk in enumerate(chunks):
            p = chunk.strip().strip(",")
            if not p:
                continue
            if not p.endswith(")") and i < len(chunks) - 1:
                p = f"{p})"
            terms.append(p)
        if len(terms) >= 2:
            return terms

    # DB 표기 호환: openFDA가 항목을 쉼표로 join 하므로, 이름 자체에 쉼표가
    # 들어간 Annex 용어("Incorrect, Inadequate ..." 등)는 split 전에 보호한다.
    # Annex Excel 에 등재된 모든 쉼표 포함 용어를 자동 수집 (긴 표현 우선).
    annex = load_annex_code_map(_annex_xlsx_mtime())
    annex_comma_terms = (
        annex.get("all_terms_with_comma", []) if isinstance(annex, dict) else []
    )
    fallback_phrases = [
        "Incorrect, Inadequate or Imprecise Result or Readings",
        "Appropriate Clinical Signs, Symptoms, Conditions Term / Code Not Available",
        "No Clinical Signs, Symptoms or Conditions",
    ]
    # Annex 가 비어있을 때 최소 보호 목록 폴백
    phrases = list(annex_comma_terms) if annex_comma_terms else fallback_phrases

    protected = text
    placeholders: Dict[str, str] = {}
    for i, phrase in enumerate(phrases):
        if "," not in phrase:
            continue
        if phrase.lower() in protected.lower():
            token = f"__COMMA_KEEP_{i}__"
            protected = re.sub(re.escape(phrase), token, protected, flags=re.IGNORECASE)
            placeholders[token] = phrase

    parts = [p.strip() for p in re.split(r"\s*,\s*", protected) if p and p.strip()]
    restored: List[str] = []
    for p in parts:
        for token, phrase in placeholders.items():
            p = p.replace(token, phrase)
        restored.append(p)
    return restored


# ---------------------------------------------------------------------------
# 인사이트 탭(탭 7) 전용 쿼리/연산 헬퍼
# ---------------------------------------------------------------------------

@st.cache_data(show_spinner=False)
def query_severe_reports(
    mtime: float,
    brands: List[str],
    date_from: Optional[date],
    date_to: Optional[date],
    keyword: str,
    severity_types: Tuple[str, ...] = ("Death", "Injury"),
    limit: int = 5000,
) -> pd.DataFrame:
    """심각 보고(기본: Death+Injury)만 필터된 주요 칼럼 로드.
    event_types 필터는 여기서 고정하므로 사이드바의 EVENT_TYPE 필터는 무시됨 (§1 특성).
    """
    where, params = _build_where(brands, list(severity_types), date_from, date_to, keyword)
    sql = f"""
        SELECT report_number, event_type, date_received, date_of_event,
               brand_name, manufacturer_name, model_number,
               patient_problems, product_problems,
               summary_complaint_kr, summary_conclusion_kr,
               event_description
        FROM maude_reports
        WHERE {where}
        ORDER BY date_received DESC, report_number DESC
        LIMIT {int(limit)}
    """
    with _connect() as conn:
        df = pd.read_sql_query(sql, conn, params=params)
    return format_problem_columns(apply_brand_aliases(df, MEMBER_TO_CANONICAL))


@st.cache_data(show_spinner=False)
def query_escalation_by_model(
    mtime: float,
    brands: List[str],
    date_from: Optional[date],
    date_to: Optional[date],
    keyword: str,
    min_total: int = 20,
) -> pd.DataFrame:
    """model_number 별 Malfunction/Injury/Death/Other 건수 집계.
    (Injury+Death)/(전체) 위험도 스코어 계산. 최소 min_total 건 이상인 모델만.
    """
    where, params = _build_where(brands, [], date_from, date_to, keyword)
    sql = f"""
        SELECT COALESCE(NULLIF(brand_name,''), '미상')      AS brand_name,
               COALESCE(NULLIF(model_number,''),'(unknown)') AS model_number,
               SUM(CASE WHEN event_type='Death'       THEN 1 ELSE 0 END) AS death_cnt,
               SUM(CASE WHEN event_type='Injury'      THEN 1 ELSE 0 END) AS injury_cnt,
               SUM(CASE WHEN event_type='Malfunction' THEN 1 ELSE 0 END) AS malf_cnt,
               SUM(CASE WHEN event_type='Other'       THEN 1 ELSE 0 END) AS other_cnt,
               COUNT(*) AS total_cnt
        FROM maude_reports
        WHERE {where}
        GROUP BY brand_name, model_number
        HAVING COUNT(*) >= {int(min_total)}
        ORDER BY total_cnt DESC
    """
    with _connect() as conn:
        df = pd.read_sql_query(sql, conn, params=params)
    df = apply_brand_aliases(df, MEMBER_TO_CANONICAL)
    if df.empty:
        return df
    df["harm_ratio"] = (df["death_cnt"] + df["injury_cnt"]) / df["total_cnt"]
    df["harm_ratio"] = df["harm_ratio"].round(3)
    return df


@st.cache_data(show_spinner=False)
def query_source_type_raw(
    mtime: float,
    brands: List[str],
    event_types: List[str],
    date_from: Optional[date],
    date_to: Optional[date],
    keyword: str,
) -> pd.DataFrame:
    """source_type + patient_problems/product_problems 만 로드 — §3 갭 분석용."""
    where, params = _build_where(brands, event_types, date_from, date_to, keyword)
    sql = f"""
        SELECT source_type, patient_problems, product_problems
        FROM maude_reports
        WHERE {where}
    """
    with _connect() as conn:
        return pd.read_sql_query(sql, conn, params=params)


def compute_source_type_gap(
    df: pd.DataFrame,
    code_col: str = "patient_problems",
    top_n: int = 15,
) -> pd.DataFrame:
    """source_type × code 교차에서 '소비자 비율 - 제조사 비율' 갭 계산.
    반환: code, consumer_pct, manufacturer_pct, hcp_pct, gap, total_reports
    gap 이 큰 순서로 top_n 개 반환 (소비자 우세 = 제조사가 놓치는 이슈).
    """
    if df.empty or code_col not in df.columns:
        return pd.DataFrame()
    df = df.dropna(subset=[code_col]).copy()
    if df.empty:
        return pd.DataFrame()
    # source_type 정규화 — MAUDE 에는 다양한 값이 섞여 있음
    src_raw = df["source_type"].fillna("Unknown").astype(str).str.strip()
    df["_src"] = src_raw.apply(_normalize_source_type)
    # 코드 explode
    df["_code"] = df[code_col].astype(str).apply(_split_problem_terms)
    df = df.explode("_code")
    df["_code"] = df["_code"].str.strip()
    df = df[df["_code"] != ""]
    if df.empty:
        return pd.DataFrame()

    # 각 src 별 전체 건수
    totals = df["_src"].value_counts().to_dict()
    consumer_total = totals.get("Consumer", 0) + totals.get("User facility", 0)  # 사용자 측
    mfr_total = totals.get("Manufacturer", 0)
    hcp_total = totals.get("Health Professional", 0)

    # code × src 피벗
    pv = (
        df.groupby(["_code", "_src"]).size().unstack(fill_value=0)
    )
    pv.columns.name = None
    # 비율 계산 (해당 src 내에서 이 코드가 차지하는 비중)
    result = pd.DataFrame(index=pv.index)
    result["consumer_pct"] = (
        (pv.get("Consumer", 0) + pv.get("User facility", 0))
        / max(consumer_total, 1) * 100
    )
    result["manufacturer_pct"] = pv.get("Manufacturer", 0) / max(mfr_total, 1) * 100
    result["hcp_pct"] = pv.get("Health Professional", 0) / max(hcp_total, 1) * 100
    result["gap"] = result["consumer_pct"] - result["manufacturer_pct"]
    # 총 노출도 — 너무 희귀한 코드는 제외
    result["total_reports"] = pv.sum(axis=1)
    result = result[result["total_reports"] >= 5]
    # gap 큰 순으로 정렬 + Top N
    result = result.sort_values("gap", ascending=False).head(top_n)
    return result.round(2).reset_index().rename(columns={"_code": "code"})


def _normalize_source_type(s: str) -> str:
    """MAUDE 의 다양한 source_type 값을 4개 카테고리로 정규화."""
    s_lower = (s or "").lower()
    if "manufacturer" in s_lower:
        return "Manufacturer"
    if "user facility" in s_lower or "hospital" in s_lower or "clinic" in s_lower:
        return "User facility"
    if "consumer" in s_lower or "patient" in s_lower or "individual" in s_lower or "voluntary" in s_lower:
        return "Consumer"
    if "health professional" in s_lower or "physician" in s_lower or "nurse" in s_lower or "hcp" in s_lower:
        return "Health Professional"
    if "distributor" in s_lower or "importer" in s_lower:
        return "Distributor"
    return "Unknown"


@st.cache_data(show_spinner=False)
def query_monthly_severity(
    mtime: float,
    brands: List[str],
    date_from: Optional[date],
    date_to: Optional[date],
    keyword: str,
    num_types: Tuple[str, ...],
    den_types: Tuple[str, ...],
) -> pd.DataFrame:
    """월별 심각도 비율 시계열. 분자/분모 event_type 집합을 사용자가 지정."""
    where, params = _build_where(brands, [], date_from, date_to, keyword)
    # 동적 CASE WHEN IN(...)
    num_sql = ",".join(["?"] * len(num_types)) if num_types else "''"
    den_sql = ",".join(["?"] * len(den_types)) if den_types else "''"
    # SQLite 의 ? 는 SQL 텍스트 내 위치 순서로 바인딩된다.
    # SELECT 절의 CASE WHEN IN(...) 이 WHERE 보다 먼저 등장하므로,
    # 파라미터 순서는 num_types → den_types → where_params 여야 한다.
    params_full = list(num_types) + list(den_types) + list(params)
    sql = f"""
        SELECT SUBSTR(REPLACE(date_received, '-', ''), 1, 4) || '-' || SUBSTR(REPLACE(date_received, '-', ''), 5, 2) AS month,
               SUM(CASE WHEN event_type IN ({num_sql}) THEN 1 ELSE 0 END) AS num_cnt,
               SUM(CASE WHEN event_type IN ({den_sql}) THEN 1 ELSE 0 END) AS den_cnt,
               COUNT(*) AS total_cnt
        FROM maude_reports
        WHERE {where} AND date_received IS NOT NULL AND date_received <> ''
        GROUP BY month
        ORDER BY month
    """
    with _connect() as conn:
        df = pd.read_sql_query(sql, conn, params=params_full)
    if df.empty:
        return df
    df["ratio_pct"] = (df["num_cnt"] / df["den_cnt"].replace(0, pd.NA) * 100).round(2)
    return df


@st.cache_data(show_spinner=False)
def query_monthly_source_mix(
    mtime: float,
    brands: List[str],
    event_types: List[str],
    date_from: Optional[date],
    date_to: Optional[date],
    keyword: str,
) -> pd.DataFrame:
    """월별 source_type 구성 비율 — HCP 비중 증가 감지용."""
    where, params = _build_where(brands, event_types, date_from, date_to, keyword)
    sql = f"""
        SELECT SUBSTR(REPLACE(date_received, '-', ''), 1, 4) || '-' || SUBSTR(REPLACE(date_received, '-', ''), 5, 2) AS month,
               source_type,
               COUNT(*) AS cnt
        FROM maude_reports
        WHERE {where} AND date_received IS NOT NULL AND date_received <> ''
        GROUP BY month, source_type
        ORDER BY month
    """
    with _connect() as conn:
        df = pd.read_sql_query(sql, conn, params=params)
    if df.empty:
        return df
    df["src_norm"] = df["source_type"].fillna("Unknown").apply(_normalize_source_type)
    pv = df.pivot_table(
        index="month", columns="src_norm", values="cnt", aggfunc="sum", fill_value=0
    )
    total_by_month = pv.sum(axis=1).replace(0, pd.NA)
    pct = pv.div(total_by_month, axis=0) * 100
    pct = pct.round(2)
    pct["_total"] = pv.sum(axis=1).astype(int)
    return pct.reset_index()


_INVESTIGATION_KEYWORDS = [
    "under investigation",
    "root cause",
    "capa",
    "corrective action",
    "field safety",
    "recall",
    "510(k)",
    "complaint investigation",
]


@st.cache_data(show_spinner=False)
def query_investigation_density(
    mtime: float,
    brands: List[str],
    event_types: List[str],
    date_from: Optional[date],
    date_to: Optional[date],
    keyword: str,
) -> pd.DataFrame:
    """제조사 서술에서 'under investigation' 류 키워드 언급 비율 (월별)."""
    where, params = _build_where(brands, event_types, date_from, date_to, keyword)
    pattern = "|".join(_INVESTIGATION_KEYWORDS)
    # SQLite 의 REGEXP 가 기본 지원 안 되므로, Python 에서 필터 후 집계
    sql = f"""
        SELECT SUBSTR(REPLACE(date_received, '-', ''), 1, 4) || '-' || SUBSTR(REPLACE(date_received, '-', ''), 5, 2) AS month,
               manufacturer_narrative
        FROM maude_reports
        WHERE {where} AND date_received IS NOT NULL AND date_received <> ''
    """
    with _connect() as conn:
        df = pd.read_sql_query(sql, conn, params=params)
    if df.empty:
        return df
    import re as _re
    regex = _re.compile(pattern, _re.IGNORECASE)
    df["has_investigation"] = (
        df["manufacturer_narrative"].fillna("").apply(lambda t: bool(regex.search(t)))
    )
    monthly = df.groupby("month").agg(
        total_cnt=("has_investigation", "size"),
        invest_cnt=("has_investigation", "sum"),
    ).reset_index()
    monthly["invest_pct"] = (monthly["invest_cnt"] / monthly["total_cnt"].replace(0, pd.NA) * 100).round(2)
    return monthly


@st.cache_data(show_spinner=False)
def query_new_codes(
    mtime: float,
    brands: List[str],
    event_types: List[str],
    date_to: Optional[date],
    keyword: str,
    window_days: int,
    baseline_days: int = 365,
) -> pd.DataFrame:
    """최근 window_days 일 동안 처음 등장한 problem code.
    기준: 그 이전 baseline_days 일 구간에는 나타나지 않았던 코드만."""
    if date_to is None:
        date_to = date.today()
    recent_from = date_to - timedelta(days=window_days)
    baseline_from = recent_from - timedelta(days=baseline_days)
    where_recent, p_recent = _build_where(brands, event_types, recent_from, date_to, keyword)
    where_base, p_base = _build_where(brands, event_types, baseline_from, recent_from, keyword)

    def _fetch(where_clause, params):
        sql = f"""
            SELECT date_received, patient_problems, product_problems
            FROM maude_reports WHERE {where_clause}
        """
        with _connect() as conn:
            return pd.read_sql_query(sql, conn, params=params)

    df_recent = _fetch(where_recent, p_recent)
    df_base = _fetch(where_base, p_base)

    def _codes_with_dates(df, col):
        if df.empty or col not in df.columns:
            return pd.DataFrame(columns=["code", "date_received", "kind"])
        d = df[["date_received", col]].dropna()
        d = d.assign(code=d[col].astype(str).apply(_split_problem_terms))
        d = d.explode("code")
        d["code"] = d["code"].str.strip()
        d = d[d["code"] != ""]
        return d[["date_received", "code"]]

    pat_recent = _codes_with_dates(df_recent, "patient_problems")
    pat_recent["kind"] = "patient_problem"
    prod_recent = _codes_with_dates(df_recent, "product_problems")
    prod_recent["kind"] = "product_problem"
    all_recent = pd.concat([pat_recent, prod_recent], ignore_index=True)

    pat_base = set(_codes_with_dates(df_base, "patient_problems")["code"])
    prod_base = set(_codes_with_dates(df_base, "product_problems")["code"])

    def _is_new(row):
        if row["kind"] == "patient_problem":
            return row["code"] not in pat_base
        return row["code"] not in prod_base

    if all_recent.empty:
        return pd.DataFrame(columns=["kind", "code", "first_seen", "count"])
    new = all_recent[all_recent.apply(_is_new, axis=1)]
    if new.empty:
        return pd.DataFrame(columns=["kind", "code", "first_seen", "count"])
    summary = (
        new.groupby(["kind", "code"])
        .agg(first_seen=("date_received", "min"), count=("code", "size"))
        .reset_index()
        .sort_values(["count", "first_seen"], ascending=[False, False])
    )
    return summary


@st.cache_data(show_spinner=False)
def query_spike_signals(
    mtime: float,
    brands: List[str],
    event_types: List[str],
    date_to: Optional[date],
    keyword: str,
    window_days: int,
    baseline_weeks: int = 12,
) -> pd.DataFrame:
    """(브랜드, event_type) 조합별 최근 window_days 보고 수 vs
    직전 baseline_weeks 주의 같은 길이 평균 비교.
    z-score > 2 또는 ratio > 2 를 급증 신호로 분류."""
    if date_to is None:
        date_to = date.today()
    recent_from = date_to - timedelta(days=window_days)
    baseline_from = recent_from - timedelta(days=window_days * baseline_weeks)
    where_recent, p_recent = _build_where(brands, event_types, recent_from, date_to, keyword)
    where_base, p_base = _build_where(brands, event_types, baseline_from, recent_from, keyword)

    with _connect() as conn:
        df_recent = pd.read_sql_query(
            f"SELECT brand_name, event_type, date_received FROM maude_reports WHERE {where_recent}",
            conn, params=p_recent,
        )
        df_base = pd.read_sql_query(
            f"SELECT brand_name, event_type, date_received FROM maude_reports WHERE {where_base}",
            conn, params=p_base,
        )
    df_recent = apply_brand_aliases(df_recent, MEMBER_TO_CANONICAL)
    df_base = apply_brand_aliases(df_base, MEMBER_TO_CANONICAL)

    if df_recent.empty and df_base.empty:
        return pd.DataFrame(columns=["brand", "event_type", "current", "baseline_avg", "ratio"])

    recent_counts = df_recent.groupby(["brand_name", "event_type"]).size().rename("current")

    # 직전 baseline_weeks 구간을 window_days 단위로 쪼개 평균 + std 계산
    if not df_base.empty:
        df_base["bucket"] = pd.to_datetime(df_base["date_received"], errors="coerce")
        df_base = df_base.dropna(subset=["bucket"])
        if not df_base.empty:
            # 각 레코드를 recent_from 으로부터 몇 window_days 전인지 버킷 인덱스
            delta_days = (recent_from - df_base["bucket"].dt.date).apply(lambda d: d.days if hasattr(d, "days") else 0)
            df_base["bucket_idx"] = (delta_days // window_days).astype(int)
            bucket_counts = (
                df_base.groupby(["brand_name", "event_type", "bucket_idx"]).size().rename("cnt").reset_index()
            )
            baseline_stats = (
                bucket_counts.groupby(["brand_name", "event_type"])
                .agg(baseline_avg=("cnt", "mean"), baseline_std=("cnt", "std"))
                .reset_index()
            )
        else:
            baseline_stats = pd.DataFrame(columns=["brand_name", "event_type", "baseline_avg", "baseline_std"])
    else:
        baseline_stats = pd.DataFrame(columns=["brand_name", "event_type", "baseline_avg", "baseline_std"])

    merged = pd.DataFrame(recent_counts).reset_index()
    if not baseline_stats.empty:
        merged = merged.merge(baseline_stats, on=["brand_name", "event_type"], how="left")
    else:
        merged["baseline_avg"] = 0.0
        merged["baseline_std"] = 0.0
    merged["baseline_avg"] = merged["baseline_avg"].fillna(0.0)
    merged["baseline_std"] = merged["baseline_std"].fillna(0.0)
    merged["ratio"] = merged["current"] / merged["baseline_avg"].replace(0, pd.NA)
    merged["z_score"] = (merged["current"] - merged["baseline_avg"]) / merged["baseline_std"].replace(0, pd.NA)
    # 상위 신호만 — current >= 3 최소치 + (z>=2 or ratio>=2)
    alert = merged[(merged["current"] >= 3) & ((merged["z_score"] >= 2) | (merged["ratio"] >= 2))]
    return alert.sort_values(["z_score", "ratio"], ascending=False).reset_index(drop=True).rename(
        columns={"brand_name": "brand"}
    ).round(2)


@st.cache_data(show_spinner=False)
def query_hcp_reports(
    mtime: float,
    brands: List[str],
    event_types: List[str],
    date_from: Optional[date],
    date_to: Optional[date],
    keyword: str,
    limit: int = 500,
) -> pd.DataFrame:
    """HCP(의료진) 가 신고자로 잡힌 MAUDE 번호 목록.
    source_type 에 'Health Professional / Physician / Nurse / HCP' 가 포함된 건만.
    §4(b) 의 '어떤 보고가 HCP 신고인지' 를 사용자가 직접 확인할 수 있게 표로 제공.
    """
    where, params = _build_where(brands, event_types, date_from, date_to, keyword)
    hcp_filter = (
        "(LOWER(COALESCE(source_type,'')) LIKE '%health professional%' "
        " OR LOWER(COALESCE(source_type,'')) LIKE '%physician%' "
        " OR LOWER(COALESCE(source_type,'')) LIKE '%nurse%' "
        " OR LOWER(COALESCE(source_type,'')) LIKE '%hcp%')"
    )
    sql = f"""
        SELECT report_number, date_received, brand_name, event_type, source_type,
               patient_problems, product_problems,
               summary_complaint_kr, summary_conclusion_kr,
               event_description
        FROM maude_reports
        WHERE {where} AND {hcp_filter}
        ORDER BY date_received DESC, report_number DESC
        LIMIT {int(limit)}
    """
    with _connect() as conn:
        df = pd.read_sql_query(sql, conn, params=params)
    return format_problem_columns(apply_brand_aliases(df, MEMBER_TO_CANONICAL))


@st.cache_data(show_spinner=False)
def query_code_matches(
    mtime: float,
    brands: List[str],
    date_to: Optional[date],
    window_days: int,
    code: str,
    kind: str,
    limit: int = 200,
) -> pd.DataFrame:
    """특정 problem code 가 등장한 보고서 목록.
    §5(b) 신규 코드 클릭 시 관련 MAUDE 를 오른쪽 뷰에 띄우는 용도.

    kind = 'patient_problem' → patient_problems 칼럼 매칭
    kind = 'product_problem' → product_problems 칼럼 매칭
    """
    if not code:
        return pd.DataFrame()
    if date_to is None:
        date_to = date.today()
    date_from = date_to - timedelta(days=window_days)
    col = "patient_problems" if kind == "patient_problem" else "product_problems"
    where, params = _build_where(brands, [], date_from, date_to, "")
    sql = f"""
        SELECT report_number, date_received, brand_name, event_type, source_type,
               patient_problems, product_problems,
               summary_complaint_kr, summary_conclusion_kr,
               event_description
        FROM maude_reports
        WHERE {where} AND {col} LIKE ?
        ORDER BY date_received DESC
        LIMIT {int(limit)}
    """
    params_full = list(params) + [f"%{code}%"]
    with _connect() as conn:
        df = pd.read_sql_query(sql, conn, params=params_full)
    return format_problem_columns(apply_brand_aliases(df, MEMBER_TO_CANONICAL))


# ---------------------------------------------------------------------------
# Altair hover-linked 차트 + 표 (사용자 요구 1)
# ---------------------------------------------------------------------------

def linked_line_and_table(
    df: pd.DataFrame,
    x_col: str,
    line_cols: List[str],
    table_cols: Optional[List[str]] = None,
    y_title: str = "",
    height: int = 280,
    key: str = "",
) -> None:
    """hover-연동된 라인 차트 + 원본 표.

    동작:
      - 위쪽: line_cols 각 지표를 라인차트로 표시 (point=True).
      - 마우스가 라인 위에 오면 세로 rule 이 그 x(월) 위치를 하이라이트.
      - 아래쪽: table_cols(없으면 line_cols 전부) × month 히트맵 스타일 표.
        hover 한 월의 모든 셀이 파스텔(살구색) 으로 물들어 한눈에 확인 가능.
      - Streamlit native line_chart 와 달리 차트와 표가 같은 Altair selection 으로 묶여 있음.
    """
    if df is None or df.empty or x_col not in df.columns:
        st.caption("데이터 없음")
        return
    # 숫자 칼럼만 살리기
    line_cols = [c for c in line_cols if c in df.columns]
    if not line_cols:
        st.caption("데이터 없음")
        return
    if table_cols is None:
        table_cols = line_cols
    table_cols = [c for c in table_cols if c in df.columns]

    data = df[[x_col] + list(dict.fromkeys(line_cols + table_cols))].copy()
    # x 축은 월 문자열("YYYY-MM") 이면 그대로 ordinal 사용
    x_field = alt.X(f"{x_col}:O", title="월", axis=alt.Axis(labelAngle=-45))

    # 라인용 long 데이터
    long_line = data.melt(id_vars=[x_col], value_vars=line_cols,
                          var_name="지표", value_name="값")

    hover = alt.selection_point(
        on="mouseover",
        nearest=True,
        fields=[x_col],
        empty=False,
        clear="mouseout",
    )

    line = alt.Chart(long_line).mark_line(point=True).encode(
        x=x_field,
        y=alt.Y("값:Q", title=y_title),
        color=alt.Color("지표:N", legend=alt.Legend(orient="top")),
        tooltip=[x_col, "지표", alt.Tooltip("값:Q", format=".2f")],
    )
    rule = alt.Chart(data).mark_rule(color="#ff9999", strokeWidth=2).encode(
        x=x_field,
        opacity=alt.condition(hover, alt.value(0.7), alt.value(0)),
        tooltip=[alt.Tooltip(f"{x_col}:O", title="월")] + [
            alt.Tooltip(f"{c}:Q", format=".2f") for c in line_cols
        ],
    ).add_params(hover)
    line_layer = (line + rule).properties(height=height)

    # 표: 히트맵 스타일 (행=지표, 열=x), 같은 hover selection 으로 파스텔 하이라이트
    long_tbl = data.melt(id_vars=[x_col], value_vars=table_cols,
                         var_name="항목", value_name="값")
    cells = alt.Chart(long_tbl).mark_rect(
        stroke="#e0e0e0", strokeWidth=0.5
    ).encode(
        x=x_field,
        y=alt.Y("항목:N", title=None),
        # hover 된 x 와 같은 컬럼 전체를 파스텔색으로
        color=alt.condition(hover, alt.value("#FFE4B5"), alt.value("#ffffff")),
        tooltip=[x_col, "항목", alt.Tooltip("값:Q", format=".2f")],
    )
    text = alt.Chart(long_tbl).mark_text(fontSize=10).encode(
        x=x_field,
        y=alt.Y("항목:N"),
        text=alt.Text("값:Q", format=".2f"),
        color=alt.value("#333"),
    )
    table_layer = (cells + text).properties(
        height=max(80, len(table_cols) * 30 + 30)
    )

    chart = alt.vconcat(line_layer, table_layer).resolve_scale(
        x="shared", color="independent"
    ).configure_view(stroke=None)
    st.altair_chart(chart, use_container_width=True, key=f"altair_linked_{key}" if key else None)


# ---------------------------------------------------------------------------
# Excel 다운로드
# ---------------------------------------------------------------------------

def df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    """DataFrame 을 Excel (xlsx) 바이트로. openpyxl write_only 모드 스트리밍."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("전체 보고서")
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        # NaN / NaT / None 을 빈 문자열로 치환 (openpyxl 허용 타입 맞추기)
        ws.append([("" if (v is None or (isinstance(v, float) and pd.isna(v))) else v) for v in row])

    if len(df) > 0:
        last_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_sheet(wb, name: str, df: pd.DataFrame) -> None:
    """인사이트 리포트용 헬퍼: DataFrame 을 write_only 워크북에 한 시트로 추가."""
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet(name[:31])  # Excel 시트명 31자 제한
    if df is None or df.empty:
        ws.append(["(데이터 없음)"])
        return
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([
            ("" if (v is None or (isinstance(v, float) and pd.isna(v))) else v)
            for v in row
        ])
    last_col = get_column_letter(len(df.columns))
    ws.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"
    ws.freeze_panes = "A2"


def generate_insight_report(
    brands: List[str],
    event_types: List[str],
    date_from: Optional[date],
    date_to: Optional[date],
    keyword: str,
    window_days: int,
    sev_num: List[str],
    sev_den: List[str],
    min_total: int,
    insights_max_rows: int,
    mtime: float,
) -> Path:
    """인사이트 탭의 모든 섹션을 한 xlsx 파일로 스냅샷 저장.

    현재 사이드바 필터와 분석 설정이 그대로 반영된 상태로 저장된다.
    파일명에 타임스탬프(YYYYMMDD_HHMM)가 들어가 덮어쓰기 없음.
    반환: 저장된 파일의 절대 경로."""
    from openpyxl import Workbook

    wb = Workbook(write_only=True)
    keyword_text, code_filters, report_stage, id_filters = _split_search_keyword(keyword)

    # 0) 요약 메타 시트
    meta_rows = [
        ["생성 시각", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["DB 파일",   str(DB_PATH.name)],
        ["전체 DB 건수", conn_count_all()],
        ["브랜드 필터", ", ".join(brands) if brands else "(전체)"],
        ["EVENT_TYPE 필터", ", ".join(event_types) if event_types else "(전체)"],
        ["날짜 범위", f"{date_from} ~ {date_to}"],
        ["키워드", keyword_text or "(없음)"],
        ["Report/MDR ID", ", ".join(id_filters) if id_filters else "(없음)"],
        ["문제 코드 번호", ", ".join(code_filters) if code_filters else "(없음)"],
        ["MDR 보고 구분", "초기 보고" if report_stage == "initial" else ("후속 수정" if report_stage == "followup" else "(전체)")],
        ["최근 윈도우", f"{window_days}일"],
        ["심각도 수식", f"{'+'.join(sev_num) or '∅'} / {'+'.join(sev_den) or '∅'}"],
        ["에스컬레이션 최소 건수", min_total],
    ]
    meta_df = pd.DataFrame(meta_rows, columns=["항목", "값"])
    _write_sheet(wb, "0_메타", meta_df)

    # 1) §1 Death/Injury 드릴다운
    df_severe = query_severe_reports(
        mtime, brands, date_from, date_to, keyword,
        severity_types=("Death", "Injury"), limit=insights_max_rows,
    )
    _write_sheet(wb, "1_Death_Injury_원본", df_severe)
    # §1 서브: 브랜드×심각도 + Top 코드들
    if not df_severe.empty:
        bx = df_severe.groupby(["brand_name", "event_type"]).size().unstack(fill_value=0).reset_index()
        _write_sheet(wb, "1b_브랜드x심각도", bx)

        def _top_codes(sub, col, label):
            domain = "patient" if col == "patient_problems" else "device"
            ser = _explode_codes(sub[col], domain=domain)
            if ser.empty:
                return pd.DataFrame(columns=[label, "건수"])
            return ser.value_counts().head(20).rename_axis(label).reset_index(name="건수")

        _write_sheet(wb, "1c_Death_Patient_Top20",
                     _top_codes(df_severe[df_severe["event_type"] == "Death"], "patient_problems", PATIENT_PROBLEM_LABEL))
        _write_sheet(wb, "1d_Death_Device_Top20",
                     _top_codes(df_severe[df_severe["event_type"] == "Death"], "product_problems", DEVICE_PROBLEM_LABEL))
        _write_sheet(wb, "1e_Injury_Patient_Top20",
                     _top_codes(df_severe[df_severe["event_type"] == "Injury"], "patient_problems", PATIENT_PROBLEM_LABEL))
        _write_sheet(wb, "1f_Injury_Device_Top20",
                     _top_codes(df_severe[df_severe["event_type"] == "Injury"], "product_problems", DEVICE_PROBLEM_LABEL))

    # 2) §2 에스컬레이션
    df_esc = query_escalation_by_model(
        mtime, brands, date_from, date_to, keyword, min_total=min_total,
    )
    if not df_esc.empty:
        df_esc_out = df_esc.sort_values("harm_ratio", ascending=False).copy()
        df_esc_out["harm_ratio_pct"] = (df_esc_out["harm_ratio"] * 100).round(1)
    else:
        df_esc_out = df_esc
    _write_sheet(wb, "2_에스컬레이션", df_esc_out)

    # 3) §3 제조사 사각지대
    df_src_raw = query_source_type_raw(
        mtime, brands, event_types, date_from, date_to, keyword,
    )
    gap_pat = compute_source_type_gap(df_src_raw, code_col="patient_problems", top_n=50)
    gap_prod = compute_source_type_gap(df_src_raw, code_col="product_problems", top_n=50)
    _write_sheet(wb, "3a_갭_Patient", gap_pat)
    _write_sheet(wb, "3b_갭_Device", gap_prod)

    # 4) §4 규제 선행 지표
    df_sev = query_monthly_severity(
        mtime, brands, date_from, date_to, keyword, tuple(sev_num), tuple(sev_den),
    )
    _write_sheet(wb, "4a_월별심각도", df_sev)

    df_src_mon = query_monthly_source_mix(
        mtime, brands, event_types, date_from, date_to, keyword,
    )
    _write_sheet(wb, "4b_월별source구성", df_src_mon)

    df_inv = query_investigation_density(
        mtime, brands, event_types, date_from, date_to, keyword,
    )
    _write_sheet(wb, "4c_조사키워드밀도", df_inv)

    # 5) §5 급증 신호 + 신규 코드
    df_spike = query_spike_signals(
        mtime, brands, event_types, date_to, keyword, window_days=window_days,
    )
    _write_sheet(wb, "5a_급증신호", df_spike)

    df_newc = query_new_codes(
        mtime, brands, event_types, date_to, keyword,
        window_days=window_days, baseline_days=365,
    )
    _write_sheet(wb, "5b_신규코드", df_newc)

    # 저장
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = BASE_DIR / f"maude_insight_report_{ts}.xlsx"
    wb.save(out_path)
    return out_path


def conn_count_all() -> int:
    """전체 DB 건수 — 리포트 메타용. 캐시 없이 즉시 쿼리."""
    try:
        with _connect() as conn:
            return int(conn.execute("SELECT COUNT(*) FROM maude_reports").fetchone()[0])
    except Exception:
        return 0


# ---------------------------------------------------------------------------
# 전체 대시보드 덤프 (모든 탭 + 네이티브 Excel 차트)
# ---------------------------------------------------------------------------

def _write_df_to_ws(ws, df: pd.DataFrame, add_filter: bool = True) -> int:
    """일반(write_only=아님) 워크시트에 DataFrame 을 써넣음. 반환: 쓴 데이터 행 수(헤더 제외)."""
    from openpyxl.utils import get_column_letter
    if df is None or df.empty:
        ws.append(["(데이터 없음)"])
        return 0
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([
            ("" if (v is None or (isinstance(v, float) and pd.isna(v))) else v)
            for v in row
        ])
    n = len(df)
    if add_filter and n > 0:
        last_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A1:{last_col}{n + 1}"
    ws.freeze_panes = "A2"
    return n


def _add_bar_chart(ws, title: str, df: pd.DataFrame, anchor: str = "D2",
                   cat_col_idx: int = 1, val_col_idx: int = 2,
                   width: float = 18, height: float = 10) -> None:
    """df 가 이미 ws 에 [헤더, 데이터...] 로 써진 상태에서 네이티브 BarChart 를 추가.
    cat_col_idx / val_col_idx 는 1-based 열 인덱스."""
    from openpyxl.chart import BarChart, Reference
    if df is None or df.empty:
        return
    n_rows = len(df)
    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = title
    chart.y_axis.title = str(df.columns[cat_col_idx - 1])
    chart.x_axis.title = str(df.columns[val_col_idx - 1])
    data = Reference(ws, min_col=val_col_idx, min_row=1,
                     max_col=val_col_idx, max_row=n_rows + 1)
    cats = Reference(ws, min_col=cat_col_idx, min_row=2,
                     max_col=cat_col_idx, max_row=n_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = width
    chart.height = height
    ws.add_chart(chart, anchor)


def _add_line_chart(ws, title: str, df: pd.DataFrame, anchor: str = "I2",
                    cat_col_idx: int = 1,
                    val_col_start: int = 2, val_col_end: Optional[int] = None,
                    width: float = 22, height: float = 12) -> None:
    """df 가 이미 ws 에 [헤더, 데이터...] 로 써진 상태에서 네이티브 LineChart 를 추가.
    val_col_start~val_col_end 까지 여러 시리즈를 그림."""
    from openpyxl.chart import LineChart, Reference
    if df is None or df.empty:
        return
    n_rows = len(df)
    if val_col_end is None:
        val_col_end = len(df.columns)
    chart = LineChart()
    chart.title = title
    chart.style = 12
    chart.y_axis.title = "건수"
    chart.x_axis.title = str(df.columns[cat_col_idx - 1])
    data = Reference(ws, min_col=val_col_start, min_row=1,
                     max_col=val_col_end, max_row=n_rows + 1)
    cats = Reference(ws, min_col=cat_col_idx, min_row=2,
                     max_col=cat_col_idx, max_row=n_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = width
    chart.height = height
    ws.add_chart(chart, anchor)


def generate_full_dashboard_report(
    brands: List[str],
    event_types: List[str],
    date_from: Optional[date],
    date_to: Optional[date],
    keyword: str,
    window_days: int,
    sev_num: List[str],
    sev_den: List[str],
    min_total: int,
    mtime: float,
    max_rows: int = 5000,
) -> Path:
    """모든 탭(인사이트 + 전체보고서 + EVENT + 문제코드 + 인구통계 + 제조사/국가 + 월별추이)을
    한 xlsx 파일로 스냅샷 저장. 표 + 네이티브(편집 가능한) Excel 차트 포함.

    인사이트 리포트와 달리 write_only 를 쓰지 않는다(네이티브 차트 때문).
    현재 사이드바 필터와 분석 설정이 그대로 반영된다.
    """
    from openpyxl import Workbook

    wb = Workbook()
    keyword_text, code_filters, report_stage, id_filters = _split_search_keyword(keyword)
    # 기본 빈 시트 제거
    default_ws = wb.active
    wb.remove(default_ws)

    available = detect_columns(mtime)

    # ------------------------------------------------------------------
    # 0) 메타
    # ------------------------------------------------------------------
    ws = wb.create_sheet("0_메타")
    meta_rows = [
        ["생성 시각", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["DB 파일", str(DB_PATH.name)],
        ["전체 DB 건수", conn_count_all()],
        ["브랜드 필터", ", ".join(brands) if brands else "(전체)"],
        ["EVENT_TYPE 필터", ", ".join(event_types) if event_types else "(전체)"],
        ["날짜 범위", f"{date_from} ~ {date_to}"],
        ["키워드", keyword_text or "(없음)"],
        ["Report/MDR ID", ", ".join(id_filters) if id_filters else "(없음)"],
        ["문제 코드 번호", ", ".join(code_filters) if code_filters else "(없음)"],
        ["MDR 보고 구분", "초기 보고" if report_stage == "initial" else ("후속 수정" if report_stage == "followup" else "(전체)")],
        ["최근 윈도우", f"{window_days}일"],
        ["심각도 수식(분자/분모)", f"{'+'.join(sev_num) or '∅'} / {'+'.join(sev_den) or '∅'}"],
        ["에스컬레이션 최소 건수", min_total],
        ["전체 보고서 최대 행수", max_rows],
    ]
    meta_df = pd.DataFrame(meta_rows, columns=["항목", "값"])
    _write_df_to_ws(ws, meta_df, add_filter=False)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60

    # ------------------------------------------------------------------
    # 1) 인사이트 §1 Death/Injury
    # ------------------------------------------------------------------
    df_severe = query_severe_reports(
        mtime, brands, date_from, date_to, keyword,
        severity_types=("Death", "Injury"), limit=max_rows,
    )
    _write_df_to_ws(wb.create_sheet("1_Death_Injury_원본"), df_severe)

    if not df_severe.empty:
        bx = df_severe.groupby(["brand_name", "event_type"]).size().unstack(fill_value=0).reset_index()
        _write_df_to_ws(wb.create_sheet("1b_브랜드x심각도"), bx)

        def _top_codes(sub, col, label):
            domain = "patient" if col == "patient_problems" else "device"
            ser = _explode_codes(sub[col], domain=domain)
            if ser.empty:
                return pd.DataFrame(columns=[label, "건수"])
            return ser.value_counts().head(20).rename_axis(label).reset_index(name="건수")

        _write_df_to_ws(wb.create_sheet("1c_Death_Patient_Top20"),
                        _top_codes(df_severe[df_severe["event_type"] == "Death"], "patient_problems", PATIENT_PROBLEM_LABEL))
        _write_df_to_ws(wb.create_sheet("1d_Death_Device_Top20"),
                        _top_codes(df_severe[df_severe["event_type"] == "Death"], "product_problems", DEVICE_PROBLEM_LABEL))
        _write_df_to_ws(wb.create_sheet("1e_Injury_Patient_Top20"),
                        _top_codes(df_severe[df_severe["event_type"] == "Injury"], "patient_problems", PATIENT_PROBLEM_LABEL))
        _write_df_to_ws(wb.create_sheet("1f_Injury_Device_Top20"),
                        _top_codes(df_severe[df_severe["event_type"] == "Injury"], "product_problems", DEVICE_PROBLEM_LABEL))

    # ------------------------------------------------------------------
    # 2) 인사이트 §2 에스컬레이션
    # ------------------------------------------------------------------
    df_esc = query_escalation_by_model(
        mtime, brands, date_from, date_to, keyword, min_total=min_total,
    )
    if not df_esc.empty:
        df_esc_out = df_esc.sort_values("harm_ratio", ascending=False).copy()
        df_esc_out["harm_ratio_pct"] = (df_esc_out["harm_ratio"] * 100).round(1)
    else:
        df_esc_out = df_esc
    _write_df_to_ws(wb.create_sheet("2_에스컬레이션"), df_esc_out)

    # ------------------------------------------------------------------
    # 3) 인사이트 §3 제조사 사각지대
    # ------------------------------------------------------------------
    df_src_raw = query_source_type_raw(
        mtime, brands, event_types, date_from, date_to, keyword,
    )
    gap_pat = compute_source_type_gap(df_src_raw, code_col="patient_problems", top_n=50)
    gap_prod = compute_source_type_gap(df_src_raw, code_col="product_problems", top_n=50)
    _write_df_to_ws(wb.create_sheet("3a_갭_Patient"), gap_pat)
    _write_df_to_ws(wb.create_sheet("3b_갭_Device"), gap_prod)

    # ------------------------------------------------------------------
    # 4) 인사이트 §4 선행 지표
    # ------------------------------------------------------------------
    df_sev = query_monthly_severity(
        mtime, brands, date_from, date_to, keyword, tuple(sev_num), tuple(sev_den),
    )
    ws_sev = wb.create_sheet("4a_월별심각도")
    n_sev = _write_df_to_ws(ws_sev, df_sev)
    # 월 vs ratio_pct 선 그래프
    if n_sev > 0 and {"month", "ratio_pct"}.issubset(df_sev.columns):
        cols = list(df_sev.columns)
        try:
            cat_idx = cols.index("month") + 1
            val_idx = cols.index("ratio_pct") + 1
            _add_line_chart(
                ws_sev, "월별 심각도 비율 (%)", df_sev,
                anchor=f"{chr(ord('A') + len(cols) + 1)}2",
                cat_col_idx=cat_idx, val_col_start=val_idx, val_col_end=val_idx,
            )
        except ValueError:
            pass

    df_src_mon = query_monthly_source_mix(
        mtime, brands, event_types, date_from, date_to, keyword,
    )
    ws_src = wb.create_sheet("4b_월별source구성")
    n_src = _write_df_to_ws(ws_src, df_src_mon)
    if n_src > 0 and len(df_src_mon.columns) >= 3:
        # 첫 칼럼이 month, 나머지는 source 별 건수/비율
        cat_idx = 1
        val_start = 2
        val_end = len(df_src_mon.columns)
        _add_line_chart(
            ws_src, "월별 신고자 유형(Source Type) 구성", df_src_mon,
            anchor=f"{chr(ord('A') + val_end + 1)}2",
            cat_col_idx=cat_idx, val_col_start=val_start, val_col_end=val_end,
        )

    df_inv = query_investigation_density(
        mtime, brands, event_types, date_from, date_to, keyword,
    )
    _write_df_to_ws(wb.create_sheet("4c_조사키워드밀도"), df_inv)

    # ------------------------------------------------------------------
    # 5) 인사이트 §5 급증 + 신규 코드
    # ------------------------------------------------------------------
    df_spike = query_spike_signals(
        mtime, brands, event_types, date_to, keyword, window_days=window_days,
    )
    _write_df_to_ws(wb.create_sheet("5a_급증신호"), df_spike)

    df_newc = query_new_codes(
        mtime, brands, event_types, date_to, keyword,
        window_days=window_days, baseline_days=365,
    )
    _write_df_to_ws(wb.create_sheet("5b_신규코드"), df_newc)

    # ------------------------------------------------------------------
    # 6) 전체 보고서 (최대 max_rows)
    # ------------------------------------------------------------------
    df_rep = query_reports(
        mtime, brands, event_types, date_from, date_to, keyword, limit=max_rows
    )
    _write_df_to_ws(wb.create_sheet("6_전체보고서"), df_rep)

    # ------------------------------------------------------------------
    # 7) EVENT_TYPE 분포 + 브랜드 × EVENT
    # ------------------------------------------------------------------
    df_evt = query_group(
        mtime, brands, event_types, date_from, date_to, keyword,
        "SELECT event_type, COUNT(*) FROM maude_reports WHERE {where} "
        "GROUP BY event_type ORDER BY COUNT(*) DESC",
        ["EVENT_TYPE", "건수"],
    )
    ws_evt = wb.create_sheet("7_EVENT_TYPE")
    n_evt = _write_df_to_ws(ws_evt, df_evt)
    if n_evt > 0:
        _add_bar_chart(ws_evt, "EVENT_TYPE 별 건수", df_evt, anchor="D2")

    df_bx = query_group(
        mtime, brands, event_types, date_from, date_to, keyword,
        "SELECT brand_name, event_type, COUNT(*) FROM maude_reports WHERE {where} "
        "GROUP BY brand_name, event_type ORDER BY COUNT(*) DESC",
        ["BRAND_NAME", "EVENT_TYPE", "건수"],
    )
    if not df_bx.empty:
        pivot_bx = (
            df_bx.pivot_table(index="BRAND_NAME", columns="EVENT_TYPE",
                              values="건수", aggfunc="sum", fill_value=0)
            .reset_index()
        )
    else:
        pivot_bx = pd.DataFrame()
    _write_df_to_ws(wb.create_sheet("7b_브랜드xEVENT"), pivot_bx)

    # ------------------------------------------------------------------
    # 8) 문제 코드 (patient / device) Top 20 + 브랜드 × Health Effect - Clinical Code
    # ------------------------------------------------------------------
    df_codes = query_code_fields(
        mtime, brands, event_types, date_from, date_to, keyword
    )

    pat_codes = _explode_codes(df_codes.get("patient_problems", pd.Series(dtype=object)), domain="patient")
    if pat_codes.empty:
        top_pat = pd.DataFrame(columns=[PATIENT_PROBLEM_LABEL, "건수"])
    else:
        top_pat = (pat_codes.value_counts().head(20)
                   .rename_axis(PATIENT_PROBLEM_LABEL).reset_index(name="건수"))
    ws_pat = wb.create_sheet("8a_환자문제Top20")
    n_pat = _write_df_to_ws(ws_pat, top_pat)
    if n_pat > 0:
        _add_bar_chart(ws_pat, f"{PATIENT_PROBLEM_LABEL} Top 20", top_pat, anchor="D2",
                       height=14)

    prod_codes = _explode_codes(df_codes.get("product_problems", pd.Series(dtype=object)), domain="device")
    if prod_codes.empty:
        top_prod = pd.DataFrame(columns=[DEVICE_PROBLEM_LABEL, "건수"])
    else:
        top_prod = (prod_codes.value_counts().head(20)
                    .rename_axis(DEVICE_PROBLEM_LABEL).reset_index(name="건수"))
    ws_prod = wb.create_sheet("8b_기기문제Top20")
    n_prod = _write_df_to_ws(ws_prod, top_prod)
    if n_prod > 0:
        _add_bar_chart(ws_prod, f"{DEVICE_PROBLEM_LABEL} Top 20", top_prod, anchor="D2",
                       height=14)

    # 브랜드 × Health Effect - Clinical Code (상위 10 × 상위 10) pivot
    if not df_codes.empty and df_codes["patient_problems"].dropna().any():
        df_exp = df_codes.dropna(subset=["patient_problems"]).copy()
        df_exp = df_exp[df_exp["brand_name"].notna() & (df_exp["brand_name"] != "")]
        df_exp["_codes"] = df_exp["patient_problems"].astype(str).apply(_split_problem_terms)
        df_exp = df_exp.explode("_codes")
        df_exp["_codes"] = df_exp["_codes"].str.strip()
        df_exp = df_exp[df_exp["_codes"] != ""]
        if not df_exp.empty:
            top_brands = df_exp["brand_name"].value_counts().head(10).index
            top_codes_idx = df_exp["_codes"].value_counts().head(10).index
            df_sub = df_exp[df_exp["brand_name"].isin(top_brands)
                            & df_exp["_codes"].isin(top_codes_idx)]
            if df_sub.empty:
                pivot_code = pd.DataFrame()
            else:
                pivot_code = (
                    df_sub.groupby(["brand_name", "_codes"]).size().unstack(fill_value=0)
                    .reindex(columns=top_codes_idx, fill_value=0)
                    .reindex(index=top_brands, fill_value=0)
                    .reset_index()
                    .rename(columns={"brand_name": "BRAND_NAME"})
                )
        else:
            pivot_code = pd.DataFrame()
    else:
        pivot_code = pd.DataFrame()
    _write_df_to_ws(wb.create_sheet("8c_브랜드xHealthEffect"), pivot_code)

    # ------------------------------------------------------------------
    # 9) 환자 인구통계
    # ------------------------------------------------------------------
    if "patient_sex" in available:
        df_sex = query_group(
            mtime, brands, event_types, date_from, date_to, keyword,
            "SELECT COALESCE(NULLIF(patient_sex,''),'미상'), COUNT(*) "
            "FROM maude_reports WHERE {where} GROUP BY 1 ORDER BY 2 DESC",
            ["PATIENT_SEX", "건수"],
        )
        ws_sex = wb.create_sheet("9a_성별")
        if _write_df_to_ws(ws_sex, df_sex) > 0:
            _add_bar_chart(ws_sex, "성별 분포", df_sex, anchor="D2")
    if "patient_race" in available:
        df_race = query_group(
            mtime, brands, event_types, date_from, date_to, keyword,
            "SELECT COALESCE(NULLIF(patient_race,''),'미상'), COUNT(*) "
            "FROM maude_reports WHERE {where} GROUP BY 1 ORDER BY 2 DESC LIMIT 20",
            ["PATIENT_RACE", "건수"],
        )
        ws_race = wb.create_sheet("9b_인종")
        if _write_df_to_ws(ws_race, df_race) > 0:
            _add_bar_chart(ws_race, "인종(Race) 분포", df_race, anchor="D2",
                           height=14)
    if "patient_ethnicity" in available:
        df_eth = query_group(
            mtime, brands, event_types, date_from, date_to, keyword,
            "SELECT COALESCE(NULLIF(patient_ethnicity,''),'미상'), COUNT(*) "
            "FROM maude_reports WHERE {where} GROUP BY 1 ORDER BY 2 DESC LIMIT 20",
            ["PATIENT_ETHNICITY", "건수"],
        )
        ws_eth = wb.create_sheet("9c_민족")
        if _write_df_to_ws(ws_eth, df_eth) > 0:
            _add_bar_chart(ws_eth, "민족(Ethnicity) 분포", df_eth, anchor="D2",
                           height=14)
    if "patient_age" in available:
        df_age = query_group(
            mtime, brands, event_types, date_from, date_to, keyword,
            "SELECT COALESCE(NULLIF(patient_age,''),'미상'), COUNT(*) "
            "FROM maude_reports WHERE {where} GROUP BY 1 ORDER BY 2 DESC LIMIT 20",
            ["PATIENT_AGE", "건수"],
        )
        ws_age = wb.create_sheet("9d_연령")
        if _write_df_to_ws(ws_age, df_age) > 0:
            _add_bar_chart(ws_age, "연령 분포 Top 20", df_age, anchor="D2",
                           height=14)

    # ------------------------------------------------------------------
    # 10) 제조사 · 국가 · 제품 코드
    # ------------------------------------------------------------------
    df_mfr = query_group(
        mtime, brands, event_types, date_from, date_to, keyword,
        "SELECT manufacturer_name, COUNT(*) FROM maude_reports WHERE {where} "
        "GROUP BY manufacturer_name ORDER BY COUNT(*) DESC LIMIT 15",
        ["MANUFACTURER", "건수"],
    )
    ws_mfr = wb.create_sheet("10a_제조사Top15")
    if _write_df_to_ws(ws_mfr, df_mfr) > 0:
        _add_bar_chart(ws_mfr, "제조사 Top 15", df_mfr, anchor="D2", height=12)

    if "manufacturer_country" in available:
        df_ctry = query_group(
            mtime, brands, event_types, date_from, date_to, keyword,
            "SELECT COALESCE(NULLIF(manufacturer_country,''),'미상'), COUNT(*) "
            "FROM maude_reports WHERE {where} GROUP BY 1 ORDER BY 2 DESC LIMIT 15",
            ["COUNTRY", "건수"],
        )
        ws_ctry = wb.create_sheet("10b_제조국Top15")
        if _write_df_to_ws(ws_ctry, df_ctry) > 0:
            _add_bar_chart(ws_ctry, "제조사 소재 국가 Top 15", df_ctry, anchor="D2",
                           height=12)

    df_pc = query_group(
        mtime, brands, event_types, date_from, date_to, keyword,
        "SELECT COALESCE(NULLIF(product_code,''),'미상'), COUNT(*) "
        "FROM maude_reports WHERE {where} GROUP BY 1 ORDER BY 2 DESC LIMIT 15",
        ["PRODUCT_CODE", "건수"],
    )
    ws_pc = wb.create_sheet("10c_제품코드Top15")
    if _write_df_to_ws(ws_pc, df_pc) > 0:
        _add_bar_chart(ws_pc, "PRODUCT_CODE Top 15", df_pc, anchor="D2", height=12)

    # ------------------------------------------------------------------
    # 11) 월별 추이 (LineChart + BarChart)
    # ------------------------------------------------------------------
    df_mon = query_group(
        mtime, brands, event_types, date_from, date_to, keyword,
        """
        SELECT SUBSTR(REPLACE(date_received, '-', ''), 1, 4) || '-' || SUBSTR(REPLACE(date_received, '-', ''), 5, 2) AS 월,
               SUM(CASE WHEN event_type='Death'       THEN 1 ELSE 0 END) AS 사망,
               SUM(CASE WHEN event_type='Injury'      THEN 1 ELSE 0 END) AS 상해,
               SUM(CASE WHEN event_type='Malfunction' THEN 1 ELSE 0 END) AS 오작동,
               SUM(CASE WHEN event_type='Other'       THEN 1 ELSE 0 END) AS 기타,
               COUNT(*) AS 합계
        FROM maude_reports
        WHERE {where} AND date_received IS NOT NULL AND date_received <> ''
        GROUP BY 1
        ORDER BY 1
        """,
        ["월", "사망", "상해", "오작동", "기타", "합계"],
    )
    ws_mon = wb.create_sheet("11_월별추이")
    n_mon = _write_df_to_ws(ws_mon, df_mon)
    if n_mon > 0:
        # LineChart — 사망/상해/오작동/기타 (cols 2..5)
        _add_line_chart(
            ws_mon, "월별 이벤트 유형 추이", df_mon,
            anchor="H2",
            cat_col_idx=1, val_col_start=2, val_col_end=5,
            width=24, height=12,
        )
        # BarChart — 합계 (col 6)
        _add_bar_chart(
            ws_mon, "월별 합계", df_mon, anchor="H26",
            cat_col_idx=1, val_col_idx=6,
            width=24, height=10,
        )

    # ------------------------------------------------------------------
    # 저장
    # ------------------------------------------------------------------
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = BASE_DIR / f"maude_full_dashboard_{ts}.xlsx"
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# UI — 사이드바 필터
# ---------------------------------------------------------------------------

if not DB_PATH.exists():
    st.error(
        f"DB 파일이 없습니다: {DB_PATH.name}\n\n"
        "먼저 `run_collector.bat` (또는 `python fda_maude_collector.py`) 을 실행해 "
        "`fda_maude_cgm.db` 를 생성하세요."
    )
    st.stop()

mtime = _db_last_modified()
all_brands, all_event_types, min_date, max_date, total_reports = load_metadata(mtime)
BRAND_GROUPS = load_brand_groups()
CANONICAL_TO_MEMBERS, MEMBER_TO_CANONICAL = build_brand_alias_maps(BRAND_GROUPS)
available_cols = detect_columns(mtime)
# 필터용 표시 브랜드: 그룹 대표명이 있으면 대표명으로 표시
display_brands = sorted({str(normalize_brand_value(b, MEMBER_TO_CANONICAL)) for b in all_brands if str(b).strip()})
# 구 버전 DB 대응: 신규 칼럼이 없는지 확인
_missing_new_cols = [c for c in (
    "manufacturer_country", "patient_sex", "patient_age",
    "patient_race", "patient_ethnicity",
) if c not in available_cols]

if total_reports == 0:
    st.warning(
        "DB 에 보고서가 0건입니다. 수집기를 먼저 실행해 데이터를 넣어주세요.\n\n"
        "`python fda_maude_collector.py --initial`"
    )
    st.stop()

with st.sidebar:
    st.title("🔎 필터")
    st.caption(f"DB: `{DB_PATH.name}` · 총 {total_reports:,} 건 수집됨")

    with st.expander("🧩 브랜드 그룹(대표명) 설정", expanded=False):
        st.markdown(
            """
            <style>
            /* checkbox 선택(checked)된 항목을 파스텔 톤으로 강조 */
            :root{
                --brandMemberBg: #f9e7f6;
                --brandMemberBorder: #ebc7e8;
                --brandMemberAccent: #b65fc0;
            }
            .stCheckbox input[type="checkbox"]:checked{
                accent-color: var(--brandMemberAccent) !important;
            }
            /* DOM 구조가 버전에 따라 달라서, 여러 패턴을 함께 지정 */
            .stCheckbox input[type="checkbox"]:checked + div,
            .stCheckbox input[type="checkbox"]:checked + span,
            .stCheckbox input[type="checkbox"]:checked ~ div,
            .stCheckbox input[type="checkbox"]:checked ~ span{
                background: var(--brandMemberBg) !important;
                border: 1px solid var(--brandMemberBorder) !important;
                border-radius: 10px !important;
                padding: 2px 8px !important;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )
        st.caption(
            "DB 에 있는 브랜드명을 여러 개 선택해 1개의 대표명으로 묶을 수 있습니다. "
            "저장 후 모든 통계 집계는 대표명 기준으로 처리됩니다."
        )

        # 새 그룹을 저장한 직후에는 다음 rerun에서 자동으로
        # "(새 그룹 만들기)"로 전환하고 입력값(대표명/멤버)을 초기 상태로 맞춥니다.
        # (중요) 버튼 핸들러 안에서 위젯 키를 직접 바꾸면 StreamlitAPIException 이 날 수 있어
        # 여기서는 플래그만 처리합니다.
        if st.session_state.get("_brand_group_after_save_new_mode"):
            st.session_state["brand_group_edit_target"] = "(새 그룹 만들기)"
            st.session_state["_brand_group_loaded_target"] = "__after_save_new__"
            st.session_state["_brand_group_after_save_new_mode"] = False

        edit_target = st.selectbox(
            "수정할 그룹",
            options=["(새 그룹 만들기)"] + sorted(BRAND_GROUPS.keys()),
            index=0,
            key="brand_group_edit_target",
        )
        default_name = "" if edit_target == "(새 그룹 만들기)" else edit_target
        default_members = [] if edit_target == "(새 그룹 만들기)" else BRAND_GROUPS.get(edit_target, [])

        # edit_target 이 바뀌면 입력칸 값도 해당 그룹 값으로 동기화.
        # (Streamlit session_state 가 이전 입력을 유지해 기본값이 안 보이는 문제 방지)
        loaded_target = st.session_state.get("_brand_group_loaded_target")
        if loaded_target != edit_target:
            st.session_state["brand_group_canonical_name"] = default_name
            st.session_state["brand_group_members"] = list(default_members)
            # checkbox 키들도 같이 동기화(버튼 저장 후 '새 그룹 만들기'로 갈 때 초기화 보장)
            selected_set = set(default_members)
            for b in all_brands:
                st.session_state[f"brand_group_member_cb::{b}"] = b in selected_set
            st.session_state["_brand_group_loaded_target"] = edit_target

        canonical_name = st.text_input(
            "대표 브랜드명",
            key="brand_group_canonical_name",
            placeholder="예: DEXCOM G7",
        ).strip()

        # 그룹별(대표명별)로 다른 파스텔 색 적용
        seed = canonical_name or ("NEW_GROUP" if edit_target == "(새 그룹 만들기)" else edit_target)
        hue = int(hashlib.md5(seed.encode("utf-8")).hexdigest(), 16) % 360
        st.markdown(
            f"<style>:root{{--brandMemberBg: hsl({hue}, 70%, 90%);--brandMemberBorder: hsl({hue}, 65%, 80%);--brandMemberAccent: hsl({hue}, 55%, 55%);}}</style>",
            unsafe_allow_html=True,
        )

        member_search = st.text_input(
            "멤버 검색 (부분일치)",
            value="",
            key="brand_group_member_search",
            placeholder="예: G7 / SENSOR / LIBRE",
        ).strip()
        member_max = st.slider(
            "표시 최대(성능/가독성용)",
            min_value=50,
            max_value=2000,
            value=600,
            step=50,
            key="brand_group_member_max",
        )

        visible_members = [
            b for b in all_brands
            if (not member_search) or (member_search.lower() in str(b).lower())
        ][: int(member_max)]

        def _toggle_brand_member(brand: str) -> None:
            sel = set(st.session_state.get("brand_group_members", []))
            cb_key = f"brand_group_member_cb::{brand}"
            if st.session_state.get(cb_key, False):
                sel.add(brand)
            else:
                sel.discard(brand)
            st.session_state["brand_group_members"] = sorted(sel)

        # 멀티셀렉트 대신 checkbox 목록을 써서,
        # '이미 선택된 항목'도 목록에 보이면서 파스텔로 강조되도록 구현합니다.
        for b in visible_members:
            cb_key = f"brand_group_member_cb::{b}"
            if cb_key not in st.session_state:
                st.session_state[cb_key] = b in set(st.session_state.get("brand_group_members", []))
            st.checkbox(
                b,
                value=st.session_state.get(cb_key, False),
                key=cb_key,
                label_visibility="visible",
                on_change=lambda bb=b: _toggle_brand_member(bb),
            )

        selected_members = st.session_state.get("brand_group_members", [])
        c_save, c_delete = st.columns(2)
        with c_save:
            if st.button("그룹 저장/갱신", use_container_width=True, key="btn_save_brand_group"):
                if not canonical_name:
                    st.warning("대표 브랜드명을 입력해 주세요.")
                elif not selected_members:
                    st.warning("그룹 멤버를 1개 이상 선택해 주세요.")
                else:
                    is_new_mode = edit_target == "(새 그룹 만들기)"
                    updated = dict(BRAND_GROUPS)
                    if edit_target != "(새 그룹 만들기)" and edit_target != canonical_name:
                        updated.pop(edit_target, None)
                    updated[canonical_name] = sorted(set(selected_members))
                    save_brand_groups(updated)
                    st.cache_data.clear()
                    if is_new_mode:
                        # 위젯 키를 여기서 직접 바꾸지 말고, 다음 rerun에서 초기화합니다.
                        st.session_state["_brand_group_after_save_new_mode"] = True
                    st.success(f"저장 완료: {canonical_name} ({len(selected_members)}개 멤버)")
                    st.rerun()
        with c_delete:
            if st.button("선택 그룹 삭제", use_container_width=True, key="btn_delete_brand_group"):
                if edit_target == "(새 그룹 만들기)":
                    st.info("삭제할 기존 그룹을 먼저 선택해 주세요.")
                else:
                    updated = dict(BRAND_GROUPS)
                    updated.pop(edit_target, None)
                    save_brand_groups(updated)
                    st.cache_data.clear()
                    st.success(f"삭제 완료: {edit_target}")
                    st.rerun()

        if BRAND_GROUPS:
            st.markdown("**저장된 그룹**")
            for canon in sorted(BRAND_GROUPS.keys()):
                members_preview = ", ".join(BRAND_GROUPS[canon][:5])
                suffix = " ..." if len(BRAND_GROUPS[canon]) > 5 else ""
                st.caption(f"- {canon} ← {members_preview}{suffix}")

    # 날짜 범위 기본값/선택 가능 범위
    default_from = min_date or (date.today() - timedelta(days=365 * 2))
    default_to = max_date or date.today()
    today = date.today()
    picker_min = min(min_date, date(2000, 1, 1)) if min_date else date(2000, 1, 1)
    picker_max = max(max_date, today) if max_date else today
    picker_max = max(picker_max, today + timedelta(days=365))

    # 필터는 form submit 시에만 반영
    if "applied_filters" not in st.session_state:
        st.session_state.applied_filters = {
            "brands": [],
            "event_types": [],
            "date_from": default_from,
            "date_to": default_to,
            "keyword": "",
            "id_search": "",
            "code_search": "",
            "report_stage": "all",
            "max_rows": 5000,
        }
        st.session_state.filter_applied = False

    applied = st.session_state.applied_filters
    applied_brands = [b for b in applied.get("brands", []) if b in display_brands]
    applied_events = [e for e in applied.get("event_types", []) if e in all_event_types]
    applied_from = applied.get("date_from") or default_from
    applied_to = applied.get("date_to") or default_to
    applied_kw = str(applied.get("keyword", "")).strip()
    applied_id_search = str(applied.get("id_search", "")).strip()
    applied_code_search = str(applied.get("code_search", "")).strip()
    applied_report_stage = _normalize_report_stage(applied.get("report_stage", "all"))
    applied_max_rows = int(applied.get("max_rows", 5000))
    if applied_max_rows < 100:
        applied_max_rows = 100
    if applied_max_rows > 50000:
        applied_max_rows = 50000

    with st.form("filter_form", clear_on_submit=False):
        form_brands = st.multiselect(
            "브랜드 (여러 개 선택 가능, 비우면 전체)",
            options=display_brands,
            default=applied_brands,
            help="그룹 대표명을 설정한 경우 대표명 기준으로 표시됩니다.",
        )
        form_events = st.multiselect(
            "EVENT_TYPE",
            options=all_event_types,
            default=applied_events,
            help="Death / Injury / Malfunction / Other — 비우면 전체",
        )
        form_date_range = st.date_input(
            "DATE_RECEIVED 범위",
            value=(applied_from, applied_to),
            min_value=picker_min,
            max_value=picker_max,
            help="FDA 접수일 기준. 캘린더에서 두 날짜를 클릭해 범위 지정. "
                 "DB 에 없는 구간을 선택해도 OK — 결과가 0건일 뿐 에러가 아닙니다.",
        )
        form_keyword = st.text_input(
            "키워드 검색",
            value=applied_kw,
            help="event_description / manufacturer_narrative / 요약 / patient·product problems / report_number / mdr_report_key 에서 부분일치 검색",
        )
        form_id_search = st.text_input(
            "Report Number / MDR Report Key 검색",
            value=applied_id_search,
            help="쉼표/공백/줄바꿈으로 여러 값 입력 가능. 숫자형(예: 2954323-2024-14003, 19171008)은 자동으로 정확일치, 그 외는 부분일치로 검색합니다.",
        )
        form_code_search = st.text_input(
            "문제 코드 번호 검색",
            value=applied_code_search,
            help="예: 3191 또는 3191, 2602. 기간/브랜드/EVENT_TYPE 필터와 함께 적용되며 "
                 "Health Effect - Clinical Code / Medical Device Problem Code 양쪽에서 매칭합니다.",
        )
        report_stage_options = {
            "전체": "all",
            "초기 보고만 (Initial submission)": "initial",
            "후속 수정만 (Followup)": "followup",
        }
        stage_labels = list(report_stage_options.keys())
        inv_stage = {v: k for k, v in report_stage_options.items()}
        form_report_stage_label = st.selectbox(
            "MDR 보고 구분",
            options=stage_labels,
            index=stage_labels.index(inv_stage.get(applied_report_stage, "전체")),
            help="초기 보고(Initial submission)와 후속 수정(Followup)을 분리해서 볼 수 있습니다.",
        )
        form_report_stage = report_stage_options[form_report_stage_label]
        form_max_rows = st.slider(
            "최대 표시 건수 (성능용 제한)",
            min_value=100, max_value=50000, step=500, value=applied_max_rows,
            help="필터 결과가 많을 때 표시/다운로드할 최대 행. 다운로드 엑셀은 이 값까지.",
        )
        submitted = st.form_submit_button("🔍 필터 적용", type="primary", use_container_width=True)

    if submitted:
        if isinstance(form_date_range, tuple):
            if len(form_date_range) == 2:
                d_from, d_to = form_date_range
            elif len(form_date_range) == 1:
                d_from = d_to = form_date_range[0]
            else:
                d_from, d_to = default_from, default_to
        elif isinstance(form_date_range, date):
            d_from = d_to = form_date_range
        else:
            d_from, d_to = default_from, default_to
        st.session_state.applied_filters = {
            "brands": form_brands,
            "event_types": form_events,
            "date_from": d_from,
            "date_to": d_to,
            "keyword": form_keyword.strip(),
            "id_search": form_id_search.strip(),
            "code_search": form_code_search.strip(),
            "report_stage": _normalize_report_stage(form_report_stage),
            "max_rows": int(form_max_rows),
        }
        st.session_state.filter_applied = True

    if not st.session_state.get("filter_applied", False):
        st.info("🔍 필터를 설정한 후 '필터 적용' 버튼을 클릭하세요")
        st.stop()

    applied = st.session_state.applied_filters
    sel_brands = list(applied.get("brands", []))
    sel_events = list(applied.get("event_types", []))
    date_from = applied.get("date_from") or default_from
    date_to = applied.get("date_to") or default_to
    keyword = str(applied.get("keyword", "")).strip()
    keyword_display = keyword
    id_search = str(applied.get("id_search", "")).strip()
    code_search = str(applied.get("code_search", "")).strip()
    report_stage = _normalize_report_stage(applied.get("report_stage", "all"))
    query_keyword = _compose_search_keyword(
        keyword,
        code_search,
        report_stage=report_stage,
        id_search=id_search,
    )
    max_rows = int(applied.get("max_rows", 5000))

    # 선택된 대표명 -> 실제 DB 브랜드 목록으로 확장 (필터 쿼리용)
    sel_brands_expanded: List[str] = []
    for chosen in sel_brands:
        if chosen in CANONICAL_TO_MEMBERS:
            sel_brands_expanded.extend(CANONICAL_TO_MEMBERS[chosen])
        else:
            sel_brands_expanded.append(chosen)
    seen_expand: Set[str] = set()
    sel_brands_query = [b for b in sel_brands_expanded if not (b in seen_expand or seen_expand.add(b))]

    st.session_state.run_queries = True

    with st.expander("📊 현재 필터 상태", expanded=False):
        st.write(f"**브랜드:** {', '.join(sel_brands) if sel_brands else '(전체)'}")
        st.write(f"**EVENT_TYPE:** {', '.join(sel_events) if sel_events else '(전체)'}")
        st.write(f"**날짜 범위:** {date_from} ~ {date_to}")
        st.write(f"**키워드:** '{keyword_display}'" if keyword_display else "(없음)")
        st.write(f"**Report/MDR ID:** '{id_search}'" if id_search else "(없음)")
        if id_search:
            parsed_ids = _parse_id_search_terms(id_search)
            st.write(f"**인식된 ID:** {', '.join(parsed_ids) if parsed_ids else '(없음)'}")
        st.write(f"**문제 코드 번호:** '{code_search}'" if code_search else "(없음)")
        if code_search:
            parsed_codes = _parse_code_numbers(code_search)
            st.write(f"**인식된 코드:** {', '.join(parsed_codes) if parsed_codes else '(없음)'}")
        st.write(
            "**MDR 보고 구분:** "
            + ("초기 보고" if report_stage == "initial" else ("후속 수정" if report_stage == "followup" else "(전체)"))
        )
        st.write(f"**최대 표시 건수:** {max_rows:,}건")

    # 이후 쿼리 함수에는 코드 필터를 포함한 keyword 를 공통 전달.
    keyword = query_keyword

# ---------------------------------------------------------------------------
# 상단 헤더 + 요약 지표
# ---------------------------------------------------------------------------

st.title("📊 FDA MAUDE CGM 부작용 대시보드")
st.caption(
    "Dexcom / FreeStyle Libre 등 CGM 관련 부작용 보고를 필드별로 검색·집계합니다. "
    "데이터는 `fda_maude_cgm.db` (FDA openFDA API 수집) 에서 실시간으로 읽습니다."
)

filtered_total = query_filtered_count(
    mtime, sel_brands_query, sel_events, date_from, date_to, keyword
)

# 지표 카드
with st.container():
    cols = st.columns(4)
    cols[0].metric("전체 DB 건수", f"{total_reports:,}")
    cols[1].metric("필터 결과 건수", f"{filtered_total:,}")
    cols[2].metric("날짜 범위", f"{date_from} ~ {date_to}")
    cols[3].metric("선택 브랜드", f"{len(sel_brands) or '전체'}")

if _missing_new_cols:
    st.info(
        "⚠️ 현재 DB 는 업그레이드 이전에 수집된 것이어서 다음 칼럼이 비어 있습니다: "
        f"`{', '.join(_missing_new_cols)}`. \n"
        "인구통계 / 국가 분포 탭은 `run_collector.bat` 을 한 번 더 실행해 최신 "
        "스키마로 마이그레이션 + 데이터를 업데이트하면 채워집니다 (기존 레코드는 "
        "유지되고 새로 받는 건만 인구통계가 포함됩니다)."
    )

if filtered_total == 0:
    st.warning(
        "필터 결과가 0건입니다. 필터를 완화하거나 기간을 넓혀보세요. "
        "참고: 브랜드는 **부분 일치** 로 매칭됩니다 — 예) 'DEXCOM' 을 선택하면 "
        "'DEXCOM G7 CONTINUOUS GLUCOSE MONITORING SYSTEM' 같은 긴 이름도 포함됩니다."
    )
    st.stop()

# ---------------------------------------------------------------------------
# 탭
# ---------------------------------------------------------------------------

if st.session_state.run_queries:
        page = st.sidebar.radio(
            "화면 선택",
            [
                "🔍 인사이트",
                "📋 전체 보고서",
                "⚠️ EVENT_TYPE 분포",
                "🏥 문제 코드",
                "👤 환자 인구통계",
                "🏭 제조사 · 국가",
                "📈 월별 추이",
            ],
            key="page_select",
        )

        # ---------------- 탭 1: 🔍 인사이트 (A+C: 개발 리스크 회피 + 규제 선제 대응)
        # 다른 탭들은 descriptive(현상 기술)이지만 이 탭은 diagnostic(진단/예측).
        # 경쟁사(Dexcom / Libre) 의 실패 패턴을 "우리 제품 설계·규제 대응" 으로 옮기는 것이 목적.
        if page == "🔍 인사이트":
            st.caption(
                "💡 **이 탭은 경쟁사 FDA MAUDE 데이터로부터 우리 제품의 설계 리스크와 "
                "규제 대응을 선제적으로 준비하기 위한 진단 뷰입니다.** 좌측 사이드바의 브랜드 / "
                "날짜 / 키워드 필터가 아래 모든 섹션에 적용됩니다 (§1 의 EVENT_TYPE 은 예외 — "
                "설계상 Death+Injury 고정)."
            )

            # ==== 분석 설정 (접을 수 있는 expander) ====
            with st.expander("⚙️ 분석 설정 (최근 윈도우 · 심각도 수식)", expanded=False):
                c_a, c_b = st.columns(2)
                with c_a:
                    st.markdown("**최근 윈도우** (급증 탐지 · 신규 코드 기준)")
                    window_preset = st.radio(
                        "프리셋",
                        options=["7일", "14일", "30일", "사용자 정의"],
                        index=2, horizontal=True, key="ins_window_preset",
                    )
                    if window_preset == "사용자 정의":
                        window_days = int(st.number_input(
                            "일 수 (1-365)", min_value=1, max_value=365, value=30, step=1,
                            key="ins_window_custom",
                        ))
                    else:
                        window_days = int(window_preset.replace("일", ""))

                with c_b:
                    st.markdown("**심각도 비율 수식** (§4 규제 선행 지표)")
                    sev_preset = st.radio(
                        "프리셋",
                        options=[
                            "보수: Death / (Death+Injury+Malfunction)",
                            "민감: Death / (Death+Injury)",
                            "사용자 정의",
                        ],
                        index=0, key="ins_sev_preset",
                    )
                    if sev_preset.startswith("보수"):
                        sev_num = ["Death"]
                        sev_den = ["Death", "Injury", "Malfunction"]
                    elif sev_preset.startswith("민감"):
                        sev_num = ["Death"]
                        sev_den = ["Death", "Injury"]
                    else:
                        sev_num = st.multiselect(
                            "분자 (더하기)",
                            options=["Death", "Injury", "Malfunction", "Other"],
                            default=["Death"], key="ins_sev_num",
                        )
                        sev_den = st.multiselect(
                            "분모 (더하기)",
                            options=["Death", "Injury", "Malfunction", "Other"],
                            default=["Death", "Injury"], key="ins_sev_den",
                        )

            st.caption(
                f"현재 설정: **최근 {window_days}일** · 심각도 비율 = "
                f"**{' + '.join(sev_num) or '∅'} / {' + '.join(sev_den) or '∅'}**"
            )

            # ==== §1. Death / Injury 심층 드릴다운 =========================
            st.markdown("---")
            st.subheader("💀 §1. Death / Injury 심층 드릴다운")
            st.caption(
                "사망·중상 보고만 추려 **어떤 patient/device 문제가 실제 환자 피해로 이어졌는가**를 "
                "분석합니다. 우리 제품 설계 리뷰에서 방어해야 할 실패 목록으로 활용."
            )
            df_severe = query_severe_reports(
                mtime, sel_brands_query, date_from, date_to, keyword,
                severity_types=("Death", "Injury"), limit=max_rows,
            )
            if df_severe.empty:
                st.info("현재 필터 범위에 Death/Injury 보고가 없습니다.")
            else:
                death_n = int((df_severe["event_type"] == "Death").sum())
                injury_n = int((df_severe["event_type"] == "Injury").sum())
                m1, m2, m3 = st.columns(3)
                m1.metric("Death 보고", f"{death_n:,}")
                m2.metric("Injury 보고", f"{injury_n:,}")
                m3.metric("합계", f"{death_n + injury_n:,}")

                # 브랜드 × 심각도 매트릭스
                st.markdown("#### 브랜드 × 심각도 (Death/Injury)")
                bx = (
                    df_severe.groupby(["brand_name", "event_type"]).size().unstack(fill_value=0)
                )
                st.dataframe(bx, use_container_width=True)

                # Death 전용 / Injury 전용 탑 코드
                cL, cR = st.columns(2)
                with cL:
                    st.markdown(f"##### Death 케이스 — {PATIENT_PROBLEM_LABEL} Top 10")
                    death_pat = _explode_codes(
                        df_severe.loc[df_severe["event_type"] == "Death", "patient_problems"]
                    , domain="patient")
                    if death_pat.empty:
                        st.caption("데이터 없음")
                    else:
                        st.dataframe(
                            death_pat.value_counts().head(10).rename_axis(PATIENT_PROBLEM_LABEL).reset_index(name="건수"),
                            use_container_width=True, hide_index=True,
                        )
                    st.markdown(f"##### Death 케이스 — {DEVICE_PROBLEM_LABEL} Top 10")
                    death_prod = _explode_codes(
                        df_severe.loc[df_severe["event_type"] == "Death", "product_problems"]
                    , domain="device")
                    if death_prod.empty:
                        st.caption("데이터 없음")
                    else:
                        st.dataframe(
                            death_prod.value_counts().head(10).rename_axis(DEVICE_PROBLEM_LABEL).reset_index(name="건수"),
                            use_container_width=True, hide_index=True,
                        )
                with cR:
                    st.markdown(f"##### Injury 케이스 — {PATIENT_PROBLEM_LABEL} Top 10")
                    inj_pat = _explode_codes(
                        df_severe.loc[df_severe["event_type"] == "Injury", "patient_problems"]
                    , domain="patient")
                    if inj_pat.empty:
                        st.caption("데이터 없음")
                    else:
                        st.dataframe(
                            inj_pat.value_counts().head(10).rename_axis(PATIENT_PROBLEM_LABEL).reset_index(name="건수"),
                            use_container_width=True, hide_index=True,
                        )
                    st.markdown(f"##### Injury 케이스 — {DEVICE_PROBLEM_LABEL} Top 10")
                    inj_prod = _explode_codes(
                        df_severe.loc[df_severe["event_type"] == "Injury", "product_problems"]
                    , domain="device")
                    if inj_prod.empty:
                        st.caption("데이터 없음")
                    else:
                        st.dataframe(
                            inj_prod.value_counts().head(10).rename_axis(DEVICE_PROBLEM_LABEL).reset_index(name="건수"),
                            use_container_width=True, hide_index=True,
                        )

                # 최근 Death 케이스 — 좌우 split view
                death_df = (
                    df_severe[df_severe["event_type"] == "Death"]
                    .head(50)
                    .reset_index(drop=True)
                )
                if not death_df.empty:
                    st.markdown("##### 📄 최근 Death 케이스 상세")
                    st.caption(
                        "좌측 표에서 MAUDE 번호 행을 클릭하면 오른쪽에 상세 내용(Clinical/Device Problem Code, "
                        "한글 요약, 결론, 원문 전문)이 펼쳐집니다."
                    )
                    with st.container(border=True):
                        cL, cR = st.columns([1, 2])
                        with cL:
                            sel = st.dataframe(
                                death_df[["report_number", "brand_name", "date_received"]]
                                .rename(columns={
                                    "report_number": "MAUDE",
                                    "brand_name": "브랜드",
                                    "date_received": "접수일",
                                }),
                                use_container_width=True,
                                hide_index=True,
                                height=400,
                                on_select="rerun",
                                selection_mode="single-row",
                                key="ins_death_select",
                            )
                        with cR:
                            rows = (
                                sel.selection.rows
                                if hasattr(sel, "selection") and hasattr(sel.selection, "rows")
                                else (sel.get("selection", {}).get("rows", []) if isinstance(sel, dict) else [])
                            )
                            if rows:
                                r = death_df.iloc[rows[0]]
                                st.markdown(
                                    f"### `{r['report_number']}`  \n"
                                    f"**브랜드:** {r['brand_name'] or '(미상)'}  "
                                    f"· **접수일:** {r['date_received'] or '?'}  "
                                    f"· **발생일:** {r['date_of_event'] or '?'}"
                                )
                                st.markdown("---")
                                st.markdown(
                                    f"**🧑 {PATIENT_PROBLEM_LABEL}:**  \n"
                                    f"{r['patient_problems'] or '(없음)'}"
                                )
                                st.markdown(
                                    f"**🛠️ {DEVICE_PROBLEM_LABEL}:**  \n"
                                    f"{r['product_problems'] or '(없음)'}"
                                )
                                st.markdown(
                                    f"**🇰🇷 [요약] 소비자 불만:**  \n"
                                    f"{r['summary_complaint_kr'] or '(없음)'}"
                                )
                                st.markdown(
                                    f"**🇰🇷 [요약] 결론:**  \n"
                                    f"{r['summary_conclusion_kr'] or '(없음)'}"
                                )
                                st.markdown("**📜 event_description (원문 전문):**")
                                st.text_area(
                                    label="event_description",
                                    value=(r["event_description"] or "(없음)"),
                                    height=260,
                                    label_visibility="collapsed",
                                    key=f"ins_death_desc_{r['report_number']}",
                                )
                            else:
                                st.info("⬅️ 좌측 표에서 MAUDE 번호 행을 클릭하면 상세가 여기에 표시됩니다.")

            # ==== §2. 에스컬레이션 위험도 ==================================
            st.markdown("---")
            st.subheader("📈 §2. 에스컬레이션 위험도 (model_number 단위)")
            st.caption(
                "`harm_ratio = (Death + Injury) / 전체 보고` — 단순 오작동이 아니라 "
                "실제 환자 피해로 이어지는 비율. 높은 모델 = 설계 결함이 의심되는 경쟁사 제품."
            )
            min_total = st.slider(
                "최소 보고 건수 (랭킹 대상)", min_value=5, max_value=200, value=20, step=5,
                key="ins_min_total",
                help="너무 희귀한 모델 제외. 통계적 안정성 확보.",
            )
            df_esc = query_escalation_by_model(
                mtime, sel_brands_query, date_from, date_to, keyword, min_total=min_total,
            )
            if df_esc.empty:
                st.info("조건을 만족하는 모델이 없습니다. 최소 건수를 낮춰보세요.")
            else:
                df_top = df_esc.sort_values("harm_ratio", ascending=False).head(15).copy()
                df_top["harm_ratio_pct"] = (df_top["harm_ratio"] * 100).round(1)
                df_show = df_top[[
                    "brand_name", "model_number", "total_cnt",
                    "death_cnt", "injury_cnt", "malf_cnt", "other_cnt",
                    "harm_ratio_pct",
                ]].rename(columns={
                    "brand_name": "BRAND", "model_number": "MODEL", "total_cnt": "전체",
                    "death_cnt": "Death", "injury_cnt": "Injury", "malf_cnt": "Malf",
                    "other_cnt": "Other", "harm_ratio_pct": "위험도(%)",
                })
                st.dataframe(df_show, use_container_width=True, hide_index=True)
                # 바 차트 — 위험도만
                st.bar_chart(
                    df_show.set_index("MODEL")[["위험도(%)"]].sort_values("위험도(%)", ascending=False)
                )

            # ==== §3. 제조사 사각지대 (Source_type 갭) ======================
            st.markdown("---")
            st.subheader("🎯 §3. 제조사 사각지대 — 소비자 vs 제조사 보고 갭")
            st.caption(
                "동일 problem code 에 대해 **소비자·사용시설 신고 비율 − 제조사 신고 비율**. "
                "갭이 큼 = 소비자는 호소하는데 제조사 공식 신고에는 잘 담기지 않는 이슈. "
                "→ 우리 제품의 PMS/CAPA 설계 시 선제 고려 / FDA 실사 시 취약 영역."
            )
            df_src = query_source_type_raw(
                mtime, sel_brands_query, sel_events, date_from, date_to, keyword,
            )
            if df_src.empty:
                st.info("데이터가 없습니다.")
            else:
                cL, cR = st.columns(2)
                with cL:
                    st.markdown(f"##### {PATIENT_PROBLEM_LABEL} 갭 Top 15")
                    gap_pat = compute_source_type_gap(df_src, code_col="patient_problems", top_n=15)
                    if gap_pat.empty:
                        st.caption("데이터 없음")
                    else:
                        st.dataframe(
                            gap_pat.rename(columns={
                                "code": PATIENT_PROBLEM_LABEL,
                                "consumer_pct": "소비자%",
                                "manufacturer_pct": "제조사%",
                                "hcp_pct": "의료진%",
                                "gap": "갭(pp)",
                                "total_reports": "노출수",
                            }),
                            use_container_width=True, hide_index=True,
                        )
                with cR:
                    st.markdown(f"##### {DEVICE_PROBLEM_LABEL} 갭 Top 15")
                    gap_prod = compute_source_type_gap(df_src, code_col="product_problems", top_n=15)
                    if gap_prod.empty:
                        st.caption("데이터 없음")
                    else:
                        st.dataframe(
                            gap_prod.rename(columns={
                                "code": DEVICE_PROBLEM_LABEL,
                                "consumer_pct": "소비자%",
                                "manufacturer_pct": "제조사%",
                                "hcp_pct": "의료진%",
                                "gap": "갭(pp)",
                                "total_reports": "노출수",
                            }),
                            use_container_width=True, hide_index=True,
                        )

            # ==== §4. 규제 선행 지표 ========================================
            st.markdown("---")
            st.subheader("📋 §4. 규제 선행 지표 (FDA 행동을 선행할 수 있는 시그널)")
            st.caption(
                "FDA Warning Letter / Recall 같은 규제 조치에 **선행**하는 것으로 알려진 시그널들. "
                "하나만 보지 말고 세 가지를 겹쳐 보면 신뢰도 상승."
            )

            # (a) 심각도 비율 월별
            st.markdown("##### (a) 월별 심각도 비율")
            st.caption(
                f"`{' + '.join(sev_num) or '∅'} / {' + '.join(sev_den) or '∅'}` 비율 월 단위 추이. "
                "건수가 아니라 비율이 중요 — 보고 활동량의 영향을 배제하고 심각성 변화만 추적."
            )
            df_sev = query_monthly_severity(
                mtime, sel_brands_query, date_from, date_to, keyword,
                tuple(sev_num), tuple(sev_den),
            )
            if df_sev.empty or df_sev["ratio_pct"].dropna().empty:
                st.caption("데이터 없음")
            else:
                st.caption("💡 라인 위에 마우스를 올리면 아래 표의 해당 월 셀이 파스텔로 하이라이트됩니다.")
                linked_line_and_table(
                    df_sev,
                    x_col="month",
                    line_cols=["ratio_pct"],
                    table_cols=["ratio_pct", "num_cnt", "den_cnt", "total_cnt"],
                    y_title="심각도 %",
                    key="4a_sev",
                )

            # (b) HCP 신고 비중 — 위(hover 연동 차트+표) → 아래(HCP 신고 MAUDE 목록)
            st.markdown("##### (b) 월별 HCP(의료진) 신고 비중")
            st.caption(
                "의료진 신고 비중이 증가 = 임상 현장에서 이슈 가시화 → FDA 관심 상승 타이밍. "
                "**아래에 이 기간 HCP 가 신고한 실제 MAUDE 번호 목록** 이 함께 제공됩니다."
            )
            df_src_mon = query_monthly_source_mix(
                mtime, sel_brands_query, sel_events, date_from, date_to, keyword,
            )
            if df_src_mon.empty:
                st.caption("데이터 없음")
            else:
                cols_to_plot = [
                    c for c in ("Health Professional", "Consumer", "Manufacturer", "User facility")
                    if c in df_src_mon.columns
                ]
                if cols_to_plot:
                    linked_line_and_table(
                        df_src_mon,
                        x_col="month",
                        line_cols=cols_to_plot,
                        table_cols=cols_to_plot + (["_total"] if "_total" in df_src_mon.columns else []),
                        y_title="신고자 구성 %",
                        key="4b_src",
                    )

            # HCP 신고 MAUDE 목록 — 차트+표 아래에 전체 폭으로
            st.markdown("**🩺 이 기간 HCP 신고 MAUDE 목록**")
            df_hcp = query_hcp_reports(
                mtime, sel_brands_query, sel_events, date_from, date_to, keyword, limit=500
            )
            st.caption(
                f"HCP(의료진) 신고 {len(df_hcp):,}건 · `source_type` 이 "
                "`Health Professional / Physician / Nurse / HCP` 류인 건만. "
                "(최근 접수일 순, 최대 500건)"
            )
            if df_hcp.empty:
                st.info("이 기간 HCP 신고 보고 없음.")
            else:
                st.dataframe(
                    df_hcp[[
                        "report_number", "date_received", "brand_name",
                        "event_type", "source_type",
                        "patient_problems", "product_problems",
                        "summary_complaint_kr",
                    ]].rename(columns={
                        "report_number": "MAUDE",
                        "date_received": "접수일",
                        "brand_name": "브랜드",
                        "event_type": "유형",
                        "source_type": "신고자",
                        "patient_problems": PATIENT_PROBLEM_LABEL,
                        "product_problems": DEVICE_PROBLEM_LABEL,
                        "summary_complaint_kr": "한글 요약",
                    }),
                    use_container_width=True,
                    hide_index=True,
                    height=360,
                )

            # (c) 제조사 서술 조사 키워드 밀도 — hover 연동
            st.markdown("##### (c) 제조사 서술 — \"under investigation\" 류 키워드 밀도")
            st.caption(
                "`" + " / ".join(_INVESTIGATION_KEYWORDS[:4]) + "`… 언급 비율. "
                "제조사가 내부적으로 이슈를 인지·조사 중임을 서술에서 드러내는 빈도. "
                "💡 라인 위에 마우스를 올리면 아래 표의 해당 월 셀이 파스텔로 하이라이트됩니다."
            )
            df_inv = query_investigation_density(
                mtime, sel_brands_query, sel_events, date_from, date_to, keyword,
            )
            if df_inv.empty or df_inv["invest_pct"].dropna().empty:
                st.caption("데이터 없음")
            else:
                linked_line_and_table(
                    df_inv,
                    x_col="month",
                    line_cols=["invest_pct"],
                    table_cols=["invest_pct", "invest_cnt", "total_cnt"],
                    y_title="조사 키워드 %",
                    key="4c_inv",
                )

            # ==== §5. 급증 신호 (boost) + 신규 코드 =========================
            st.markdown("---")
            st.subheader(f"🚨 §5. 최근 {window_days}일 급증 신호 & 신규 problem code")

            # (a) 급증 신호
            st.markdown(f"##### (a) 급증 신호 — 최근 {window_days}일 vs 직전 12개 윈도우 평균")
            st.caption(
                "(브랜드 × event_type) 단위로 최근 구간 보고 수를 직전 기간과 비교. "
                "`z-score >= 2` 또는 `ratio >= 2` 이고 최소 3건 이상인 조합만 표시."
            )
            df_spike = query_spike_signals(
                mtime, sel_brands_query, sel_events, date_to, keyword, window_days=window_days,
            )
            if df_spike.empty:
                st.success("✅ 현재 급증 신호 없음 (필터 범위 기준).")
            else:
                st.warning(f"⚠️ 급증 신호 {len(df_spike)}개 감지")
                st.dataframe(
                    df_spike.rename(columns={
                        "brand": "BRAND", "event_type": "EVENT", "current": "최근",
                        "baseline_avg": "기준평균", "baseline_std": "기준표준편차",
                        "ratio": "배수", "z_score": "z_score",
                    }),
                    use_container_width=True, hide_index=True,
                )

            # (b) 신규 코드 — 좌우 split view
            st.markdown(f"##### (b) 최근 {window_days}일 내 처음 등장한 problem code")
            st.caption(
                "직전 12개월 기간에는 없었는데 최근에 나타난 코드 = 새로운 이슈 유형 조기 감지. "
                "`kind` = patient_problem(임상) / product_problem(기기). "
                "좌측 표에서 행을 클릭하면 오른쪽에 해당 코드를 담은 MAUDE 목록과 상세가 펼쳐집니다."
            )
            df_newc = query_new_codes(
                mtime, sel_brands_query, sel_events, date_to, keyword,
                window_days=window_days, baseline_days=365,
            )
            if df_newc.empty:
                st.info("신규 코드 없음 (이전 12개월 대비).")
            else:
                df_newc_view = df_newc.reset_index(drop=True)
                with st.container(border=True):
                    cL, cR = st.columns([1, 2])
                    with cL:
                        sel_nc = st.dataframe(
                            df_newc_view.rename(columns={
                                "kind": "종류", "code": "코드",
                                "first_seen": "첫 등장", "count": "건수",
                            }),
                            use_container_width=True,
                            hide_index=True,
                            height=400,
                            on_select="rerun",
                            selection_mode="single-row",
                            key="ins_newcode_select",
                        )
                    with cR:
                        rows_nc = (
                            sel_nc.selection.rows
                            if hasattr(sel_nc, "selection") and hasattr(sel_nc.selection, "rows")
                            else (
                                sel_nc.get("selection", {}).get("rows", [])
                                if isinstance(sel_nc, dict)
                                else []
                            )
                        )
                        if rows_nc:
                            r_nc = df_newc_view.iloc[rows_nc[0]]
                            code = str(r_nc["code"])
                            kind = str(r_nc["kind"])
                            st.markdown(
                                f"### 🆕 `{code}`  \n"
                                f"**종류:** {kind}  · **첫 등장:** {r_nc['first_seen']}  "
                                f"· **건수:** {int(r_nc['count'])}"
                            )
                            df_matches = query_code_matches(
                                mtime, sel_brands_query, date_to, window_days, code, kind, limit=200,
                            )
                            st.caption(f"이 코드를 담은 MAUDE {len(df_matches):,}건")
                            if df_matches.empty:
                                st.info("매칭 MAUDE 없음.")
                            else:
                                st.dataframe(
                                    df_matches[[
                                        "report_number", "date_received", "brand_name",
                                        "event_type", "source_type",
                                    ]].rename(columns={
                                        "report_number": "MAUDE",
                                        "date_received": "접수일",
                                        "brand_name": "브랜드",
                                        "event_type": "유형",
                                        "source_type": "신고자",
                                    }),
                                    use_container_width=True,
                                    hide_index=True,
                                    height=220,
                                )
                                # 드릴다운: 특정 MAUDE 선택 → 상세
                                picked = st.selectbox(
                                    "📄 상세 조회할 MAUDE 번호",
                                    options=df_matches["report_number"].tolist(),
                                    key=f"ins_newcode_pick_{code}",
                                )
                                if picked:
                                    det = df_matches[df_matches["report_number"] == picked].iloc[0]
                                    st.markdown(
                                        f"**브랜드:** {det['brand_name'] or '(미상)'} · "
                                        f"**유형:** {det['event_type']} · **신고자:** {det['source_type']}"
                                    )
                                    st.markdown(
                                        f"**{PATIENT_PROBLEM_LABEL}:** {det['patient_problems'] or '(없음)'}  \n"
                                        f"**{DEVICE_PROBLEM_LABEL}:** {det['product_problems'] or '(없음)'}  \n"
                                        f"**[요약] 소비자 불만:** {det['summary_complaint_kr'] or '(없음)'}  \n"
                                        f"**[요약] 결론:** {det['summary_conclusion_kr'] or '(없음)'}"
                                    )
                                    st.text_area(
                                        label="event_description",
                                        value=(det["event_description"] or "(없음)"),
                                        height=200,
                                        label_visibility="collapsed",
                                        key=f"ins_newcode_desc_{picked}",
                                    )
                        else:
                            st.info("⬅️ 좌측 표에서 코드 행을 클릭하면 관련 MAUDE 들이 여기에 표시됩니다.")

            # ==== 리포트 발행 버튼 ==========================================
            st.markdown("---")
            st.subheader("📄 리포트 발행")
            st.caption(
                "현재 필터·설정 상태의 스냅샷을 xlsx 로 저장합니다. 자동 실행되지 않고 "
                "**버튼 클릭 시에만** 생성됩니다. 저장 위치는 수집기와 같은 폴더 "
                f"(`{BASE_DIR.name}`). 파일명에 타임스탬프가 들어가 덮어쓰기 없음."
            )

            col_rpt1, col_rpt2 = st.columns(2)

            with col_rpt1:
                st.markdown("**(a) 인사이트 탭만** — 표 중심 (write_only 스트리밍, 가벼움)")
                if st.button("📄 인사이트 리포트 (xlsx)", type="primary", key="ins_report_btn"):
                    with st.spinner("인사이트 리포트 생성 중..."):
                        rpt_path = generate_insight_report(
                            brands=sel_brands_query,
                            event_types=sel_events,
                            date_from=date_from,
                            date_to=date_to,
                            keyword=keyword,
                            window_days=window_days,
                            sev_num=sev_num,
                            sev_den=sev_den,
                            min_total=min_total,
                            insights_max_rows=max_rows,
                            mtime=mtime,
                        )
                    st.success(f"생성 완료: `{rpt_path.name}`")
                    with open(rpt_path, "rb") as f:
                        st.download_button(
                            label=f"⬇️ 다운로드 ({rpt_path.stat().st_size // 1024} KB)",
                            data=f.read(),
                            file_name=rpt_path.name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="ins_report_dl",
                        )
                    st.caption(f"경로: `{rpt_path}`")

            with col_rpt2:
                st.markdown(
                    "**(b) 전체 대시보드** — 모든 탭 + **엑셀 네이티브 차트** "
                    "(BarChart / LineChart, 엑셀에서 바로 편집 가능)"
                )
                if st.button("📦 전체 대시보드 덤프 (xlsx)", type="primary", key="full_report_btn"):
                    with st.spinner("전체 대시보드 덤프 생성 중 (차트 포함)..."):
                        full_path = generate_full_dashboard_report(
                            brands=sel_brands_query,
                            event_types=sel_events,
                            date_from=date_from,
                            date_to=date_to,
                            keyword=keyword,
                            window_days=window_days,
                            sev_num=sev_num,
                            sev_den=sev_den,
                            min_total=min_total,
                            mtime=mtime,
                            max_rows=max_rows,
                        )
                    st.success(f"생성 완료: `{full_path.name}`")
                    with open(full_path, "rb") as f:
                        st.download_button(
                            label=f"⬇️ 다운로드 ({full_path.stat().st_size // 1024} KB)",
                            data=f.read(),
                            file_name=full_path.name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="full_report_dl",
                        )
                    st.caption(f"경로: `{full_path}`")
                    st.caption(
                        "포함 시트: 0_메타 · 1~5 인사이트 · 6_전체보고서 · 7_EVENT_TYPE · "
                        "8_문제코드 · 9_인구통계 · 10_제조사국가 · 11_월별추이"
                    )


        # ---------------- 탭 2: 전체 보고서
        if page == "📋 전체 보고서":
            st.subheader(f"필터 결과 ({min(filtered_total, max_rows):,} / {filtered_total:,} 건 표시)")

            df = query_reports(
                mtime, sel_brands_query, sel_events, date_from, date_to, keyword, limit=max_rows
            )

            st.dataframe(
                df,
                use_container_width=True,
                height=600,
                hide_index=True,
            )

            # Excel 다운로드 버튼
            if len(df) > 0:
                excel_bytes = df_to_excel_bytes(df)
                fname = (
                    f"maude_filtered_{date_from}_{date_to}_"
                    f"{(sel_brands[0] if sel_brands else 'ALL').replace(' ', '_')}.xlsx"
                )
                st.download_button(
                    label=f"⬇️ Excel 다운로드 ({len(df):,} 행)",
                    data=excel_bytes,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            # 개별 보고서 상세 보기
            if len(df) > 0:
                st.markdown("#### 개별 보고서 상세")
                options = df["MAUDE 번호"].astype(str).tolist()
                pick = st.selectbox("MAUDE 번호 선택", options=options, index=0)
                if pick:
                    row = df[df["MAUDE 번호"].astype(str) == pick].iloc[0]
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(f"**EVENT_TYPE:** {row['EVENT_TYPE']}")
                        st.markdown(f"**BRAND:** {row['BRAND_NAME']}")
                        st.markdown(f"**MANUFACTURER:** {row['MANUFACTURER']} ({row['COUNTRY'] or 'N/A'})")
                        st.markdown(f"**DATE_RECEIVED:** {row['DATE_RECEIVED']} · DATE_OF_EVENT: {row['DATE_OF_EVENT']}")
                        st.markdown(
                            f"**환자:** {row['SEX'] or '?'} · "
                            f"{row['AGE'] or '?'} · "
                            f"{row['RACE'] or '?'} · "
                            f"{row['ETHNICITY'] or '?'}"
                        )
                    with c2:
                        st.markdown(f"**[요약] 소비자 불만:** {row['[요약] 소비자 불만']}")
                        st.markdown(f"**[요약] 제조사 대응:** {row['[요약] 제조사 대응']}")
                        st.markdown(f"**[요약] 결론:** {row['[요약] 결론']}")

                    # Problem code (Annex A / E) — 라벨과 코드 함께 표기
                    he_val = row.get("HEALTH_EFFECT_CLINICAL_CODE") if hasattr(row, "get") else row["HEALTH_EFFECT_CLINICAL_CODE"]
                    md_val = row.get("MEDICAL_DEVICE_PROBLEM_CODE") if hasattr(row, "get") else row["MEDICAL_DEVICE_PROBLEM_CODE"]
                    st.markdown(
                        f"**🧑 {PATIENT_PROBLEM_LABEL} (Annex E):** {he_val or '(없음)'}  \n"
                        f"**⚙️ {DEVICE_PROBLEM_LABEL} (Annex A):** {md_val or '(없음)'}"
                    )

                    sections = query_report_narrative_sections(mtime, str(pick))
                    with st.expander("EVENT_DESCRIPTION (원문)"):
                        event_desc = sections.get("event_description") or row["EVENT_DESCRIPTION (원문)"]
                        st.text(event_desc or "(없음)")
                    with st.expander("MANUFACTURER_NARRATIVE (원문)"):
                        mfr_narr = sections.get("manufacturer_narrative") or row["MANUFACTURER_NARRATIVE (원문)"]
                        st.text(mfr_narr or "(없음)")
                    with st.expander("ADDITIONAL_MANUFACTURER_NARRATIVE (원문)"):
                        st.text(sections.get("additional_manufacturer_narrative") or "(없음)")

            # 코드별 보고서 탐색 (드릴다운) — Insight (b) 와 동일한 split-view 패턴
            if len(df) > 0:
                st.markdown("---")
                st.subheader("📊 코드별 보고서 탐색")
                st.caption(
                    "현재 필터 결과(상단 표)의 problem code 별 빈도. "
                    "좌측 표에서 코드를 클릭하면 우측에 해당 코드를 포함한 보고서 목록과 상세가 표시됩니다."
                )

                def _render_code_drilldown(
                    label: str, col_name: str, kind_key: str, annex_letter: str
                ) -> None:
                    if col_name not in df.columns:
                        st.info("데이터 없음")
                        return
                    rows_acc: List[Tuple[str, str]] = []
                    for v in df[col_name].dropna().astype(str):
                        if not v.strip():
                            continue
                        for p in _split_problem_terms(v):
                            term, code = _extract_problem_term_code(p)
                            if not term:
                                continue
                            rows_acc.append((code or "(코드없음)", term))
                    if not rows_acc:
                        st.info("코드 데이터가 없습니다.")
                        return
                    summary = (
                        pd.DataFrame(rows_acc, columns=["코드", "용어"])
                        .groupby(["코드", "용어"], as_index=False)
                        .size()
                        .rename(columns={"size": "건수"})
                        .sort_values(["건수", "코드"], ascending=[False, True])
                        .reset_index(drop=True)
                    )
                    st.caption(
                        f"고유 코드 {len(summary):,}개 · 총 출현 {int(summary['건수'].sum()):,} 회 "
                        f"(보고서당 다중 코드 포함)"
                    )
                    with st.container(border=True):
                        cL, cR = st.columns([1, 2])
                        with cL:
                            sel = st.dataframe(
                                summary,
                                use_container_width=True,
                                hide_index=True,
                                height=420,
                                on_select="rerun",
                                selection_mode="single-row",
                                key=f"code_dd_{kind_key}_table",
                            )
                        with cR:
                            rows_sel = (
                                sel.selection.rows
                                if hasattr(sel, "selection") and hasattr(sel.selection, "rows")
                                else (
                                    sel.get("selection", {}).get("rows", [])
                                    if isinstance(sel, dict)
                                    else []
                                )
                            )
                            if not rows_sel:
                                st.info("⬅️ 좌측 표에서 코드 행을 클릭하면 관련 보고서가 표시됩니다.")
                                return
                            r_sel = summary.iloc[rows_sel[0]]
                            code = str(r_sel["코드"])
                            term = str(r_sel["용어"])
                            st.markdown(
                                f"### `{code}` — {term}  \n"
                                f"**Annex {annex_letter}** · "
                                f"**현재 필터 내 매칭:** {int(r_sel['건수'])} 회"
                            )
                            # df 행 중 해당 코드 포함 행 필터
                            if code == "(코드없음)":
                                pattern = re.escape(term)
                            else:
                                pattern = re.escape(f"({code})")
                            mask = (
                                df[col_name].fillna("").astype(str)
                                .str.contains(pattern, regex=True, na=False)
                            )
                            sub = df.loc[mask].copy()
                            if sub.empty:
                                st.warning("매칭 보고서가 없습니다.")
                                return
                            st.caption(f"이 코드를 담은 MAUDE {len(sub):,}건")
                            table_cols = [
                                c for c in [
                                    "MAUDE 번호", "DATE_RECEIVED", "BRAND_NAME",
                                    "EVENT_TYPE", "SEX", "AGE",
                                ] if c in sub.columns
                            ]
                            st.dataframe(
                                sub[table_cols],
                                use_container_width=True,
                                hide_index=True,
                                height=220,
                            )
                            picked = st.selectbox(
                                "📄 상세 조회할 MAUDE 번호",
                                options=sub["MAUDE 번호"].astype(str).tolist(),
                                key=f"code_dd_{kind_key}_pick_{code}",
                            )
                            if picked:
                                d = sub[sub["MAUDE 번호"].astype(str) == picked].iloc[0]
                                st.markdown(
                                    f"**브랜드:** {d.get('BRAND_NAME','') or '(미상)'} · "
                                    f"**EVENT_TYPE:** {d.get('EVENT_TYPE','')} · "
                                    f"**DATE_RECEIVED:** {d.get('DATE_RECEIVED','')}"
                                )
                                st.markdown(
                                    f"**🧑 {PATIENT_PROBLEM_LABEL}:** "
                                    f"{d.get('HEALTH_EFFECT_CLINICAL_CODE','') or '(없음)'}  \n"
                                    f"**⚙️ {DEVICE_PROBLEM_LABEL}:** "
                                    f"{d.get('MEDICAL_DEVICE_PROBLEM_CODE','') or '(없음)'}  \n"
                                    f"**[요약] 소비자 불만:** "
                                    f"{d.get('[요약] 소비자 불만','') or '(없음)'}  \n"
                                    f"**[요약] 결론:** "
                                    f"{d.get('[요약] 결론','') or '(없음)'}"
                                )
                                st.text_area(
                                    label="event_description",
                                    value=str(d.get("EVENT_DESCRIPTION (원문)", "") or "(없음)"),
                                    height=180,
                                    label_visibility="collapsed",
                                    key=f"code_dd_{kind_key}_desc_{picked}",
                                )

                tab_he, tab_md = st.tabs([
                    f"🧑 {PATIENT_PROBLEM_LABEL}  (Annex E)",
                    f"⚙️ {DEVICE_PROBLEM_LABEL}  (Annex A)",
                ])
                with tab_he:
                    _render_code_drilldown(
                        PATIENT_PROBLEM_LABEL, "HEALTH_EFFECT_CLINICAL_CODE",
                        kind_key="patient", annex_letter="E",
                    )
                with tab_md:
                    _render_code_drilldown(
                        DEVICE_PROBLEM_LABEL, "MEDICAL_DEVICE_PROBLEM_CODE",
                        kind_key="device", annex_letter="A",
                    )

        # ---------------- 탭 3: EVENT_TYPE 분포
        if page == "⚠️ EVENT_TYPE 분포":
            st.subheader("EVENT_TYPE 별 건수")
            df_evt = query_group(
                mtime, sel_brands_query, sel_events, date_from, date_to, keyword,
                "SELECT event_type, COUNT(*) FROM maude_reports WHERE {where} "
                "GROUP BY event_type ORDER BY COUNT(*) DESC",
                ["EVENT_TYPE", "건수"],
            )
            if df_evt.empty:
                st.info("표시할 데이터가 없습니다.")
            else:
                c1, c2 = st.columns([1, 2])
                with c1:
                    st.dataframe(df_evt, use_container_width=True, hide_index=True)
                with c2:
                    st.bar_chart(df_evt.set_index("EVENT_TYPE"))

            st.markdown("#### 브랜드 × EVENT_TYPE 교차")
            df_bx = query_group(
                mtime, sel_brands_query, sel_events, date_from, date_to, keyword,
                "SELECT brand_name, event_type, COUNT(*) FROM maude_reports WHERE {where} "
                "GROUP BY brand_name, event_type ORDER BY COUNT(*) DESC",
                ["BRAND_NAME", "EVENT_TYPE", "건수"],
            )
            if not df_bx.empty:
                pivot = df_bx.pivot_table(
                    index="BRAND_NAME", columns="EVENT_TYPE", values="건수", aggfunc="sum", fill_value=0
                )
                st.dataframe(pivot, use_container_width=True)

        # ---------------- 탭 4: 문제 코드 (Health Effect - Clinical Code / Medical Device Problem Code)
        if page == "🏥 문제 코드":
            st.caption(
                "FDA MAUDE 의 표준 코드셋 필드입니다. 신고 접수 시 FDA CDRH 표준 리스트에서 "
                "선택한 값이라 정형화되어 통계 활용성이 높습니다."
            )

            df_codes = query_code_fields(
                mtime, sel_brands_query, sel_events, date_from, date_to, keyword
            )
            st.caption(f"집계 대상 보고서: **{len(df_codes):,} 건** (현재 필터 기준)")

            st.subheader(f"🏥 {PATIENT_PROBLEM_LABEL} Top 20")
            pat_codes = _explode_codes(df_codes.get("patient_problems", pd.Series(dtype=object)), domain="patient")
            if pat_codes.empty:
                st.info("현재 필터 범위에 patient_problems 데이터가 없습니다.")
            else:
                top_pat = (
                    pat_codes.value_counts()
                    .head(20)
                    .rename_axis(PATIENT_PROBLEM_LABEL)
                    .reset_index(name="건수")
                )
                c1, c2 = st.columns([1, 2])
                with c1:
                    st.dataframe(top_pat, use_container_width=True, hide_index=True)
                with c2:
                    st.bar_chart(top_pat.set_index(PATIENT_PROBLEM_LABEL))

            st.markdown("---")
            st.subheader(f"⚙️ {DEVICE_PROBLEM_LABEL} Top 20")
            prod_codes = _explode_codes(df_codes.get("product_problems", pd.Series(dtype=object)), domain="device")
            if prod_codes.empty:
                st.info("현재 필터 범위에 product_problems 데이터가 없습니다.")
            else:
                top_prod = (
                    prod_codes.value_counts()
                    .head(20)
                    .rename_axis(DEVICE_PROBLEM_LABEL)
                    .reset_index(name="건수")
                )
                c1, c2 = st.columns([1, 2])
                with c1:
                    st.dataframe(top_prod, use_container_width=True, hide_index=True)
                with c2:
                    st.bar_chart(top_prod.set_index(DEVICE_PROBLEM_LABEL))

            st.markdown("---")
            st.subheader(f"🔬 브랜드 × {PATIENT_PROBLEM_LABEL} 교차 (상위 10 × 상위 10)")
            if df_codes.empty or df_codes["patient_problems"].dropna().empty:
                st.info("교차 분석 가능한 데이터가 없습니다.")
            else:
                df_exp = df_codes.dropna(subset=["patient_problems"]).copy()
                df_exp = df_exp[df_exp["brand_name"].notna() & (df_exp["brand_name"] != "")]
                df_exp["_codes"] = df_exp["patient_problems"].astype(str).apply(_split_problem_terms)
                df_exp = df_exp.explode("_codes")
                df_exp["_codes"] = df_exp["_codes"].str.strip()
                df_exp = df_exp[df_exp["_codes"] != ""]

                if df_exp.empty:
                    st.info("교차 분석 가능한 데이터가 없습니다.")
                else:
                    top_brands = df_exp["brand_name"].value_counts().head(10).index
                    top_codes = df_exp["_codes"].value_counts().head(10).index
                    df_sub = df_exp[
                        df_exp["brand_name"].isin(top_brands)
                        & df_exp["_codes"].isin(top_codes)
                    ]
                    if df_sub.empty:
                        st.info("상위 10 × 10 교차 범위에 데이터가 없습니다.")
                    else:
                        pivot = (
                            df_sub.groupby(["brand_name", "_codes"])
                            .size()
                            .unstack(fill_value=0)
                        )
                        pivot = pivot.reindex(columns=top_codes, fill_value=0)
                        pivot = pivot.reindex(index=top_brands, fill_value=0)
                        pivot.columns.name = PATIENT_PROBLEM_LABEL
                        pivot.index.name = "BRAND_NAME"
                        st.dataframe(pivot, use_container_width=True)

        # ---------------- 탭 5: 환자 인구통계
        if page == "👤 환자 인구통계":
            demo_cols = {"patient_sex", "patient_race", "patient_ethnicity", "patient_age"}
            available_demo = demo_cols & set(available_cols)
            if not available_demo:
                st.info(
                    "현재 DB 에는 환자 인구통계 칼럼(`patient_sex` / `patient_race` / "
                    "`patient_ethnicity` / `patient_age`)이 없습니다."
                )
            else:
                if "patient_sex" in available_cols:
                    st.subheader("성별 분포")
                    df_sex = query_group(
                        mtime, sel_brands_query, sel_events, date_from, date_to, keyword,
                        "SELECT COALESCE(NULLIF(patient_sex,''),'미상'), COUNT(*) "
                        "FROM maude_reports WHERE {where} GROUP BY 1 ORDER BY 2 DESC",
                        ["PATIENT_SEX", "건수"],
                    )
                    c1, c2 = st.columns(2)
                    with c1:
                        st.dataframe(df_sex, use_container_width=True, hide_index=True)
                    with c2:
                        if not df_sex.empty:
                            st.bar_chart(df_sex.set_index("PATIENT_SEX"))

                if "patient_race" in available_cols:
                    st.subheader("인종(Race) 분포")
                    df_race = query_group(
                        mtime, sel_brands_query, sel_events, date_from, date_to, keyword,
                        "SELECT COALESCE(NULLIF(patient_race,''),'미상'), COUNT(*) "
                        "FROM maude_reports WHERE {where} GROUP BY 1 ORDER BY 2 DESC LIMIT 20",
                        ["PATIENT_RACE", "건수"],
                    )
                    c1, c2 = st.columns(2)
                    with c1:
                        st.dataframe(df_race, use_container_width=True, hide_index=True)
                    with c2:
                        if not df_race.empty:
                            st.bar_chart(df_race.set_index("PATIENT_RACE"))

                if "patient_ethnicity" in available_cols:
                    st.subheader("민족(Ethnicity) 분포")
                    df_eth = query_group(
                        mtime, sel_brands_query, sel_events, date_from, date_to, keyword,
                        "SELECT COALESCE(NULLIF(patient_ethnicity,''),'미상'), COUNT(*) "
                        "FROM maude_reports WHERE {where} GROUP BY 1 ORDER BY 2 DESC LIMIT 20",
                        ["PATIENT_ETHNICITY", "건수"],
                    )
                    c1, c2 = st.columns(2)
                    with c1:
                        st.dataframe(df_eth, use_container_width=True, hide_index=True)
                    with c2:
                        if not df_eth.empty:
                            st.bar_chart(df_eth.set_index("PATIENT_ETHNICITY"))

                if "patient_age" in available_cols:
                    st.subheader("연령대 분포")
                    df_age = query_group(
                        mtime, sel_brands_query, sel_events, date_from, date_to, keyword,
                        "SELECT COALESCE(NULLIF(patient_age,''),'미상'), COUNT(*) "
                        "FROM maude_reports WHERE {where} GROUP BY 1 ORDER BY 2 DESC LIMIT 20",
                        ["PATIENT_AGE", "건수"],
                    )
                    st.dataframe(df_age, use_container_width=True, hide_index=True)

        # ---------------- 탭 6: 제조사 · 국가
        if page == "🏭 제조사 · 국가":
            st.subheader("제조사 Top 15")
            df_mfr = query_group(
                mtime, sel_brands_query, sel_events, date_from, date_to, keyword,
                "SELECT manufacturer_name, COUNT(*) FROM maude_reports WHERE {where} "
                "GROUP BY manufacturer_name ORDER BY COUNT(*) DESC LIMIT 15",
                ["MANUFACTURER", "건수"],
            )
            c1, c2 = st.columns([1, 2])
            with c1:
                st.dataframe(df_mfr, use_container_width=True, hide_index=True)
            with c2:
                if not df_mfr.empty:
                    st.bar_chart(df_mfr.set_index("MANUFACTURER"))

            st.subheader("제조사 소재 국가 Top 15")
            if "manufacturer_country" not in available_cols:
                st.info("현재 DB 에는 `manufacturer_country` 칼럼이 없습니다.")
            else:
                df_ctry = query_group(
                    mtime, sel_brands_query, sel_events, date_from, date_to, keyword,
                    "SELECT COALESCE(NULLIF(manufacturer_country,''),'미상'), COUNT(*) "
                    "FROM maude_reports WHERE {where} GROUP BY 1 ORDER BY 2 DESC LIMIT 15",
                    ["COUNTRY", "건수"],
                )
                c1, c2 = st.columns([1, 2])
                with c1:
                    st.dataframe(df_ctry, use_container_width=True, hide_index=True)
                with c2:
                    if not df_ctry.empty:
                        st.bar_chart(df_ctry.set_index("COUNTRY"))

            st.subheader("제품 코드(PRODUCT_CODE) 별 Top 15")
            df_pc = query_group(
                mtime, sel_brands_query, sel_events, date_from, date_to, keyword,
                "SELECT COALESCE(NULLIF(product_code,''),'미상'), COUNT(*) "
                "FROM maude_reports WHERE {where} GROUP BY 1 ORDER BY 2 DESC LIMIT 15",
                ["PRODUCT_CODE", "건수"],
            )
            st.dataframe(df_pc, use_container_width=True, hide_index=True)

        # ---------------- 탭 7: 월별 추이
        if page == "📈 월별 추이":
            st.subheader("월별 이벤트 유형 추이")
            df_mon = query_group(
                mtime, sel_brands_query, sel_events, date_from, date_to, keyword,
                """
                SELECT substr(date_received,1,7) AS 월,
                       SUM(CASE WHEN event_type='Death'       THEN 1 ELSE 0 END) AS 사망,
                       SUM(CASE WHEN event_type='Injury'      THEN 1 ELSE 0 END) AS 상해,
                       SUM(CASE WHEN event_type='Malfunction' THEN 1 ELSE 0 END) AS 오작동,
                       SUM(CASE WHEN event_type='Other'       THEN 1 ELSE 0 END) AS 기타,
                       COUNT(*) AS 합계
                FROM maude_reports
                WHERE {where} AND date_received IS NOT NULL AND date_received <> ''
                GROUP BY 1
                ORDER BY 1
                """,
                ["월", "사망", "상해", "오작동", "기타", "합계"],
            )
            if df_mon.empty:
                st.info("표시할 데이터가 없습니다.")
            else:
                df_trend = df_mon.set_index("월")
                st.line_chart(df_trend[["사망", "상해", "오작동", "기타"]])
                st.bar_chart(df_trend[["합계"]])
                with st.expander("월별 원본 표"):
                    st.dataframe(df_trend, use_container_width=True)
